# -*- coding: utf-8 -*-
"""
گزارش زنده‌ی SVA — مستقیم از گوگل‌شیت «group Asli» می‌خونه (تب‌های
backtest(sva M15) و Backtest(sva 1H))، آمار رو حساب می‌کنه و HTML می‌سازه.

نکته‌ی مهم: این نسخه‌ی اول (v1) فقط آمار کلی (تعداد، وین‌ریت و expectancy
به‌ازای هر سطح R، و شکاف معمولی/مخفی برای M15) رو حساب می‌کنه — همون چیزهایی
که مستقیم از داده‌ی خام و بدون نیاز به قضاوت دستی اضافه (مثل «داخل تایم/
خارج تایم») قابل استخراجه. جدول‌های ریزتر (رشته باخت، برگشت به فری‌ریسک،
ترکیب ارز×تایم×واگرایی) بعداً که قانون دقیق «داخل/خارج تایم» مشخص شد اضافه
می‌شن.
"""

import io
import requests
import openpyxl
from datetime import datetime
import pytz

SHEET_ID = "1tsq6cyRjwi8z36QsreQX80v-8WXdj2vXujd0HKlmmQ4"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TEHRAN = pytz.timezone("Asia/Tehran")

R_LEVELS = [("1R", 1.0, "1R ✓"), ("1.5R", 1.5, "1.5R ✓"), ("2R", 2.0, "2R ✓"), ("3R", 3.0, "3R ✓")]


def fetch_workbook():
    """دانلود مستقیم فایل خام گوگل‌شیت (چون شیت عمومیه، نیازی به کلید نیست)."""
    r = requests.get(EXPORT_URL, timeout=30)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)


def _cell(ws, r, c):
    return ws.cell(r, c).value


def parse_trades(ws, has_divergence_col):
    """
    ستون‌ها (مشترک بین M15 و 1H):
    A=#, B=Symbol, C=Direction, D=EntryDate, E=EntryTime, F=EntryPrice,
    G=StopLoss, H=Result, I=کلوز دستی, J=Stop($),
    K=1R✓, L=1.5R✓, M=2R✓, N=3R✓, O=Timeframe, P=Sender, Q=Notes,
    [فقط M15] R=برگشت به فری, S=نوع واگرایی, بعدش Chart Img ...
    """
    trades = []
    for r in range(3, ws.max_row + 1):
        symbol = _cell(ws, r, 2)
        if not symbol:
            continue
        result = _cell(ws, r, 8)
        trade = {
            "symbol": symbol,
            "direction": _cell(ws, r, 3),
            "result": result,
            "sender": _cell(ws, r, 16),
            "ticks": {
                "1R": _cell(ws, r, 11),
                "1.5R": _cell(ws, r, 12),
                "2R": _cell(ws, r, 13),
                "3R": _cell(ws, r, 14),
            },
        }
        if has_divergence_col:
            trade["divergence"] = _cell(ws, r, 19)  # نوع واگرایی؟
        trades.append(trade)
    return trades


def is_valid_trade(t):
    """ردیف‌هایی که هنوز نتیجه ندارن (باز/ناقص) رو از حساب کنار می‌ذاریم."""
    return t["result"] in ("Win", "Loss", "BE", "Be", "be")


def stats_for_r_level(trades, tick_key, r_value):
    valid = [t for t in trades if is_valid_trade(t)]
    n = len(valid)
    if n == 0:
        return {"n": 0, "win_rate": 0.0, "expectancy": 0.0, "wins": 0}
    wins = sum(1 for t in valid if t["ticks"].get(tick_key) == 1)
    win_rate = wins / n
    expectancy = win_rate * r_value - (1 - win_rate) * 1.0
    return {"n": n, "win_rate": win_rate * 100, "expectancy": expectancy, "wins": wins}


def compute_overview(trades):
    valid = [t for t in trades if is_valid_trade(t)]
    total = len(valid)
    wins = sum(1 for t in valid if t["result"] == "Win")
    losses = sum(1 for t in valid if t["result"] == "Loss")
    be = total - wins - losses
    per_r = {}
    for label, rv, key in R_LEVELS:
        per_r[label] = stats_for_r_level(trades, label, rv)
    return {"total": total, "wins": wins, "losses": losses, "be": be, "per_r": per_r}


def compute_by_divergence(trades):
    out = {}
    for kind in ("معمولی", "مخفی"):
        subset = [t for t in trades if t.get("divergence") == kind]
        per_r = {label: stats_for_r_level(subset, label, rv) for label, rv, key in R_LEVELS}
        valid = [t for t in subset if is_valid_trade(t)]
        out[kind] = {"n": len(valid), "per_r": per_r}
    return out


def fmt_pct(x):
    return f"{x:.1f}%"


def fmt_exp(x):
    sign = "+" if x >= 0 else ""
    return f"{sign}{x:.2f}"


def cls_for(x):
    if x > 0.02:
        return "pos"
    if x < -0.02:
        return "neg"
    return "neu"


def render_r_row(label, s):
    if s["n"] == 0:
        return f"<td class='neu'>—</td><td class='neu'>—</td>"
    return (f"<td class='{cls_for(s['win_rate']-50)}'>{fmt_pct(s['win_rate'])}</td>"
            f"<td class='{cls_for(s['expectancy'])}'>{fmt_exp(s['expectancy'])}</td>")


def render_overview_table(title, ov):
    rows = "".join(render_r_row(label, ov["per_r"][label]) for label, rv, key in R_LEVELS)
    return f"""
    <div class="panel"><h3>{title}</h3>
    <div class="tscroll"><table>
    <thead><tr><th>n</th><th colspan="2">1R</th><th colspan="2">1.5R</th><th colspan="2">2R</th><th colspan="2">3R</th></tr>
    <tr class="subh"><th></th><th>WR</th><th>exp</th><th>WR</th><th>exp</th><th>WR</th><th>exp</th><th>WR</th><th>exp</th></tr>
    </thead><tbody><tr><td class='n'>{ov['total']}</td>{rows}</tr></tbody></table></div>
    <p style="font-size:11px;color:var(--muted2);font-family:var(--fmono);margin-top:8px">
    برد {ov['wins']} · باخت {ov['losses']} · سر‌به‌سر {ov['be']}</p>
    </div>"""


def render_divergence_table(div):
    rows = ""
    for kind, d in div.items():
        cells = "".join(render_r_row(kind, d["per_r"][label]) for label, rv, key in R_LEVELS)
        rows += f"<tr><td>{kind}</td><td class='n'>{d['n']}</td>{cells}</tr>"
    return f"""
    <div class="panel"><h3>تفکیک نوع واگرایی (فقط M15)</h3>
    <div class="tscroll"><table>
    <thead><tr><th>نوع</th><th>n</th><th colspan="2">1R</th><th colspan="2">1.5R</th><th colspan="2">2R</th><th colspan="2">3R</th></tr>
    <tr class="subh"><th></th><th></th><th>WR</th><th>exp</th><th>WR</th><th>exp</th><th>WR</th><th>exp</th><th>WR</th><th>exp</th></tr>
    </thead><tbody>{rows}</tbody></table></div>
    </div>"""


PAGE_CSS = """
:root{--bg:#080C10;--panel:#0D1520;--panel2:#121D2A;--border:#1C2A38;--bordersoft:#162030;
--text:#DDE5EE;--muted:#7C8EA0;--muted2:#4E6070;--win:#1FCE6B;--loss:#F04060;--gold:#E8B340;--blue:#5B9BFF;
--r:12px;--fm:'Vazirmatn',sans-serif;--fmono:'JetBrains Mono',monospace;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:radial-gradient(ellipse 1200px 500px at 15% -10%,rgba(232,179,64,.06),transparent),var(--bg);color:var(--text);font-family:var(--fm);line-height:1.7;}
.wrap{max-width:1080px;margin:0 auto;padding:24px 16px 50px;}
.hero{margin-bottom:22px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;}
.hero .eye{font-family:var(--fmono);font-size:11px;letter-spacing:.12em;color:var(--gold);margin-bottom:8px;}
.hero h1{font-size:22px;font-weight:900;margin-bottom:6px;}
.hero p{color:var(--muted);font-size:12.5px;}
.btn{background:var(--gold);color:#080C10;font-weight:800;border:none;border-radius:10px;padding:10px 18px;font-family:var(--fm);font-size:13px;cursor:pointer;}
.btn:hover{opacity:.9;}
.panel{background:var(--panel);border:1px solid var(--border);border-radius:var(--r);padding:14px;margin-bottom:14px;}
.panel h3{font-size:13px;font-weight:700;margin-bottom:8px;color:var(--muted);}
table{width:100%;border-collapse:collapse;font-size:12px;}
thead th{text-align:center;padding:6px;color:var(--muted);font-weight:600;font-size:10px;border-bottom:1px solid var(--border);background:var(--panel2);}
thead tr.subh th{font-size:9px;color:var(--muted2);padding:3px;}
tbody td{padding:7px 6px;border-bottom:1px solid var(--bordersoft);font-family:var(--fmono);text-align:center;}
tbody td:first-child{font-family:var(--fm);font-weight:600;text-align:right;}
td.n{color:var(--muted);}
td.pos{color:var(--win);} td.neg{color:var(--loss);} td.neu{color:var(--muted);}
.tscroll{overflow-x:auto;}
footer{text-align:center;color:var(--muted2);font-size:11px;font-family:var(--fmono);padding:18px;}
.note{background:rgba(232,179,64,.09);border:1px solid rgba(232,179,64,.35);border-radius:10px;padding:12px 14px;font-size:12.5px;margin-bottom:14px;}
"""


def render_page(ov_m15, ov_1h, div_m15, generated_at):
    return f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>گزارش زنده SVA</title>
<link href="https://fonts.googleapis.com/css2?family=Vazirmatn:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>{PAGE_CSS}</style>
</head>
<body>
<div class="wrap">
  <div class="hero">
    <div>
      <div class="eye">SVA · گزارش زنده</div>
      <h1>گزارش آماری <span style="color:var(--gold)">SVA</span></h1>
      <p>مستقیم از گوگل‌شیت خونده می‌شه · آخرین آپدیت: {generated_at}</p>
    </div>
    <button class="btn" onclick="location.href=location.pathname+'?t='+Date.now()">🔄 بروزرسانی</button>
  </div>
  <div class="note">
    <b style="color:var(--gold);display:block;margin-bottom:6px">نسخه‌ی اول (v1)</b>
    این صفحه فقط آمار کلی (وین‌ریت و expectancy هر سطح R، و تفکیک نوع واگرایی) رو مستقیم و زنده از شیت حساب می‌کنه.
    جدول‌های ریزتر گزارش قبلی (داخل/خارج تایم، رشته باخت، برگشت به فری‌ریسک) بعد از مشخص‌شدن قانون دقیق «داخل/خارج تایم» اضافه می‌شن.
  </div>
  {render_overview_table("آمار کلی — M15", ov_m15)}
  {render_divergence_table(div_m15)}
  {render_overview_table("آمار کلی — H1", ov_1h)}
  <footer>گزارش زنده‌ی SVA · منبع: شیت «group Asli» (backtest(sva M15) + Backtest(sva 1H))</footer>
</div>
</body>
</html>"""


def build_report_html():
    wb = fetch_workbook()
    ws_m15 = wb["backtest(sva M15)"]
    ws_1h = wb["Backtest(sva 1H)"]
    trades_m15 = parse_trades(ws_m15, has_divergence_col=True)
    trades_1h = parse_trades(ws_1h, has_divergence_col=False)
    ov_m15 = compute_overview(trades_m15)
    ov_1h = compute_overview(trades_1h)
    div_m15 = compute_by_divergence(trades_m15)
    generated_at = datetime.now(TEHRAN).strftime("%Y-%m-%d %H:%M") + " (تهران)"
    return render_page(ov_m15, ov_1h, div_m15, generated_at)
