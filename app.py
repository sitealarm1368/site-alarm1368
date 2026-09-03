import os, json, time, threading, requests, pytz, secrets
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from flask import Flask, request, jsonify, send_from_directory
from datetime import datetime, timedelta
import re

app = Flask(__name__, static_folder='static')
VERSION = "8.0"

TEHRAN = pytz.timezone("Asia/Tehran")

# ==================== متغیرهای محیطی ====================
ALERTS_FILE = "alerts.json"
FIRED_BACKUP_FILE = "fired_backup.json"  # لایه‌ی دفاعی دوم — جلوگیری از فایر مجدد بعد از ری‌استارت اگه Supabase قطع بود
BOT_TOKEN_ENV = os.environ.get("BOT_TOKEN", "")
YOUR_CHAT_ID = "109419675"
BROADCAST_MODE = os.environ.get("BROADCAST_MODE", "false").lower() == "true"

_cache_alerts = None
# قفل محافظ کش آلارم‌ها — جلوگیری از race condition وقتی چند thread هم‌زمان
# (مثلاً check_alerts و add_alert) با فاصله‌ی کم روی _cache_alerts کار می‌کنن
_alerts_cache_lock = threading.RLock()

def now_teh():
    return datetime.now(TEHRAN).strftime("%Y-%m-%d %H:%M:%S")

def now_pretty():
    return now_teh()  # alias برای سازگاری

def get_pip_multiplier(symbol):
    sym_up = symbol.upper()
    crypto_list = ['BTC','ETH','SOL','BNB','XRP','ADA','DOGE','TRX','TON','AVAX','MATIC','DOT','LINK','UNI','ATOM','LTC','SHIB','OP','ARB','NEAR','FTM','SAND','MANA']
    if any(x in sym_up for x in crypto_list):
        return 1
    if "XAU" in sym_up or "XAG" in sym_up:
        return 10
    if "JPY" in sym_up:
        return 100
    return 10000

def is_crypto_symbol(sym):
    sym_up = sym.upper()
    crypto_list = ['BTC','ETH','SOL','BNB','XRP','ADA','DOGE','TRX','TON','AVAX','MATIC','DOT','LINK','UNI','ATOM','LTC','SHIB','OP','ARB','NEAR','FTM','SAND','MANA']
    return any(c in sym_up for c in crypto_list)

def _empty_alerts():
    return {
        "alerts": [], "archive": [], "telegram": {"bot_token": "", "chat_ids": []},
        "users": [], "errors": [], "last_update": None
    }

def fix_alerts(data):
    e = _empty_alerts()
    for k in e:
        if k not in data:
            data[k] = e[k]
    return data

# =====================================================================
# Supabase — جدول alerts (هر آلارم یه ردیف + یه ردیف config با id='__config__')
# =====================================================================

def _sb_upsert_alert(a):
    """یه آلارم رو upsert کن"""
    if not SUPABASE_KEY: return
    try:
        # فیلدهای پایه که حتماً توی جدول هستن
        payload = {
            "id":           a["id"],
            "symbol":       a.get("symbol",""),
            "type":         a.get("type","forex"),
            "condition":    a.get("condition","above"),
            "target_price": float(a.get("target_price",0)),
            "status":       "expired" if a.get("expired_at") else ("fired" if a.get("fired_at") else ("active" if a.get("active") else "cancelled")),
            "created_by":   a.get("created_by",""),
            "created_at":   a.get("created_at", now_teh()),
            "comment":      a.get("comment",""),
            "is_private":   bool(a.get("is_private", False)),
            "notify_only":  str(a.get("notify_only","")) if a.get("notify_only") else None,
            "active":       bool(a.get("active", True)),
            "last_price":   float(a["last_price"]) if a.get("last_price") is not None else None,
            "last_checked": a.get("last_checked"),
            "fired_at":     a.get("fired_at"),
            "fired_price":  float(a["fired_price"]) if a.get("fired_price") is not None else None,
        }
        # فیلدهای اختیاری — فقط اگه توی جدول وجود داشت اضافه میشن
        if a.get("tag"):
            payload["tag"] = a["tag"]
        if a.get("private_cid") is not None:
            payload["private_cid"] = str(a["private_cid"]) if a["private_cid"] else None
        if a.get("expires_at") is not None:
            payload["expires_at"] = a["expires_at"]
        if a.get("expired_at") is not None:
            payload["expired_at"] = a["expired_at"]

        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/alerts",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload, timeout=10)
        if r.status_code not in (200,201,204):
            err_text = r.text
            # اگه ستون پیدا نشد، بدون اون فیلد دوباره تلاش کن
            if r.status_code == 400 and "PGRST204" in err_text:
                import re as _re
                missing = _re.search(r"Could not find the '([^']+)'", err_text)
                if missing:
                    col = missing.group(1)
                    print(f"[alerts] ستون '{col}' توی DB نیست — بدون اون retry میکنیم")
                    payload.pop(col, None)
                    r2 = requests.post(
                        f"{SUPABASE_URL}/rest/v1/alerts",
                        headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
                        json=payload, timeout=10)
                    if r2.status_code not in (200,201,204):
                        print(f"[alerts] upsert {a['id']} retry failed: {r2.status_code} {r2.text[:120]}")
                    return
            print(f"[alerts] upsert {a['id']}: {r.status_code} {r.text[:120]}")
    except Exception as e:
        print(f"[alerts] upsert error: {e}")

def _sb_upsert_config(tg, users, errors):
    """config (token + chat_ids + users) رو توی یه ردیف ثابت ذخیره کن"""
    if not SUPABASE_KEY: return
    try:
        payload = {
            "id": "__config__",
            "symbol": "__config__",
            "type": "config",
            "condition": "none",
            "target_price": 0,
            "telegram_token": tg.get("bot_token",""),
            "chat_ids": tg.get("chat_ids", []),
            "users": users,
            "active": False,
        }
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/alerts",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload, timeout=10)
        if r.status_code not in (200,201,204):
            print(f"[alerts] config save: {r.status_code} {r.text[:80]}")
    except Exception as e:
        print(f"[alerts] config save error: {e}")

def _sb_load_deprioritize_masoud() -> bool:
    """می‌خونه آیا اولویت پایین مسعود فعاله یا نه — پیش‌فرض True اگه چیزی پیدا نشد"""
    if not SUPABASE_KEY:
        return True
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__&select=deprioritize_masoud",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200:
            rows = r.json()
            if rows and rows[0].get("deprioritize_masoud") is not None:
                return bool(rows[0]["deprioritize_masoud"])
    except Exception as e:
        print(f"[assign] load deprioritize_masoud exc: {e}")
    return True

def _sb_save_deprioritize_masoud(value: bool):
    """فقط همین یک فیلد رو patch می‌کنه — با upsert_config تداخل نداره"""
    if not SUPABASE_KEY:
        return
    try:
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            json={"deprioritize_masoud": value}, timeout=8)
    except Exception as e:
        print(f"[assign] save deprioritize_masoud exc: {e}")

def _sb_load_news_reminder_state() -> dict:
    """وضعیت هشدار ۱۵ دقیقه‌ای اخبار قرمز رو می‌خونه — {date, events, sent} —
    نیازمند یه ستون jsonb به اسم news_reminder_state روی ردیف __config__ جدول alerts.
    اگه ستون هنوز ساخته نشده یا خالیه، دیکشنری خالی برمی‌گردونه (fallback به حافظه‌ی موقت)."""
    if not SUPABASE_KEY:
        return {}
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__&select=news_reminder_state",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200:
            rows = r.json()
            if rows and rows[0].get("news_reminder_state") is not None:
                val = rows[0]["news_reminder_state"]
                return val if isinstance(val, dict) else json.loads(val)
    except Exception as e:
        print(f"[news] load news_reminder_state exc: {e}")
    return {}

def _sb_save_news_reminder_state(state: dict):
    """فقط همین یک فیلد رو patch می‌کنه — با upsert_config تداخل نداره"""
    if not SUPABASE_KEY:
        return
    try:
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            json={"news_reminder_state": state}, timeout=8)
    except Exception as e:
        print(f"[news] save news_reminder_state exc: {e}")

def _sb_load_unavailable_members() -> list:
    """اسم اعضایی که موقتاً از چرخه‌ی تقسیم آلارم کنار گذاشته شدن (مرخصی/غیره) —
    پیش‌فرض لیست خالی (یعنی همه در دسترسن) اگه چیزی پیدا نشد. نیازمند یه ستون
    jsonb به اسم unavailable_members روی ردیف __config__ جدول alerts."""
    if not SUPABASE_KEY:
        return []
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__&select=unavailable_members",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200:
            rows = r.json()
            if rows and rows[0].get("unavailable_members") is not None:
                val = rows[0]["unavailable_members"]
                return val if isinstance(val, list) else json.loads(val)
    except Exception as e:
        print(f"[assign] load unavailable_members exc: {e}")
    return []

def _sb_save_unavailable_members(members: list):
    """فقط همین یک فیلد رو patch می‌کنه — با upsert_config تداخل نداره"""
    if not SUPABASE_KEY:
        return
    try:
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.__config__",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            json={"unavailable_members": members}, timeout=8)
    except Exception as e:
        print(f"[assign] save unavailable_members exc: {e}")

def _sb_load_all_alerts():
    """همه ردیف‌های جدول alerts رو بخون و به فرمت داخلی تبدیل کن"""
    if not SUPABASE_KEY: return None
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?select=*&limit=2000",
            headers=_sb_h(), timeout=10)
        if r.status_code != 200:
            print(f"[alerts] load failed: {r.status_code} {r.text[:80]}")
            return None
        rows = r.json()
        if not rows:
            print("[alerts] Supabase: جدول خالیه")
            return None

        config_row = next((x for x in rows if x["id"] == "__config__"), None)
        tg = {"bot_token": "", "chat_ids": []}
        users = []
        if config_row:
            tg["bot_token"] = config_row.get("telegram_token","") or ""
            raw_cids = config_row.get("chat_ids") or []
            tg["chat_ids"] = raw_cids if isinstance(raw_cids, list) else json.loads(raw_cids)
            raw_users = config_row.get("users") or []
            users = raw_users if isinstance(raw_users, list) else json.loads(raw_users)

        alerts = []
        archive = []
        for row in rows:
            if row["id"] == "__config__": continue
            a = {
                "id":           row["id"],
                "symbol":       row.get("symbol",""),
                "type":         row.get("type","forex"),
                "condition":    row.get("condition","above"),
                "target_price": row.get("target_price",0),
                "created_by":   row.get("created_by",""),
                "created_at":   row.get("created_at",""),
                "comment":      row.get("comment",""),
                "is_private":   row.get("is_private", False),
                "private_cid":  row.get("private_cid") or row.get("notify_only"),
                "notify_only":  row.get("notify_only"),
                "active":       row.get("active", True),
                "last_price":   row.get("last_price"),
                "last_checked": row.get("last_checked"),
                "fired_at":     row.get("fired_at"),
                "fired_price":  row.get("fired_price"),
            }
            status = row.get("status","active")
            if status == "fired" or row.get("fired_at"):
                archive.append(a)
            elif status == "active" and row.get("active"):
                alerts.append(a)

        data = {"alerts": alerts, "archive": archive, "telegram": tg,
                "users": users, "errors": [], "last_update": now_teh()}
        print(f"[alerts] Loaded from Supabase — {len(alerts)} active, {len(archive)} archived")
        return data
    except Exception as e:
        print(f"[alerts] load error: {e}")
        return None



def load_alerts():
    global _cache_alerts
    with _alerts_cache_lock:
        if _cache_alerts is not None:
            return _cache_alerts
        # 1. Supabase
        d = _sb_load_all_alerts()
        if d is not None:
            _cache_alerts = d
            return _cache_alerts
        # 2. local fallback
        if os.path.exists(ALERTS_FILE):
            try:
                with open(ALERTS_FILE, "r", encoding="utf-8") as f:
                    _cache_alerts = fix_alerts(json.load(f))
                    return _cache_alerts
            except: pass
        _cache_alerts = _empty_alerts()
        return _cache_alerts

def _sb_delete_alert(aid):
    """یه آلارم رو از Supabase حذف کن"""
    _deleted_ids.add(str(aid))  # فوری به blacklist اضافه کن
    if not SUPABASE_KEY: return
    try:
        r = requests.delete(
            f"{SUPABASE_URL}/rest/v1/alerts?id=eq.{aid}",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            timeout=8)
        print(f"[alerts] delete {aid}: status={r.status_code} body={r.text[:80]}")
    except Exception as e:
        print(f"[alerts] delete error: {e}")

def save_alerts(data):
    """cache رو آپدیت کن + local backup — Supabase رو در background بزن"""
    global _cache_alerts
    with _alerts_cache_lock:
        _cache_alerts = data
    # local backup سریع
    try:
        with open(ALERTS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[alerts] local backup error: {e}")
    # Supabase در background — بلاک نمیشه
    if SUPABASE_KEY:
        snapshot = (
            data.get("telegram",{}),
            data.get("users",[]),
            data.get("errors",[]),
            list(data.get("alerts",[]))
        )
        def _bg(snap=snapshot):
            tg, users, errors, alerts = snap
            _sb_upsert_config(tg, users, errors)
            for a in alerts:
                _sb_upsert_alert(a)
        threading.Thread(target=_bg, daemon=True).start()

def save_alert_expired(a):
    """
    وقتی آلارم منقضی می‌شه (تاریخ انقضاش رسیده و فایر نشده) — به‌جای حذف کامل
    از دیتابیس، فقط با status='expired' علامت‌گذاری می‌شه (دقیقاً مثل fired،
    ولی با یه وضعیت جدا). یعنی رکوردش برای همیشه تو Supabase می‌مونه و با SQL
    قابل خروجی گرفتنه، فقط دیگه تو لیست «آلارم‌های فعال» و چک قیمت نیست.
    """
    global _cache_alerts
    a["expired_at"] = now_teh()
    a["active"] = False
    ok = _sb_upsert_alert_retry(a, max_tries=3)
    if not ok:
        print(f"[alerts] ⚠️ expired state برای {a.get('id')} تو Supabase ذخیره نشد!")
    with _alerts_cache_lock:
        if _cache_alerts is not None:
            _cache_alerts["alerts"] = [x for x in _cache_alerts.get("alerts", []) if x.get("id") != a.get("id")]

def save_alert_fired(a):
    """
    وقتی آلارم fire میشه — ردیف رو آپدیت کن و cache رو پاک کن تا archive درست لود بشه.
    این ذخیره‌سازی حیاتیه: اگه fail بشه و سرور همون لحظه ری‌استارت/دیپلوی بشه،
    آلارم دوباره fired_at ندار می‌مونه و بعد از بالا اومدن دوباره فایر می‌شه.
    برای همین با چند تلاش (retry) کار می‌کنیم و هم‌زمان یه بک‌آپ محلی فوری هم می‌زنیم.
    """
    global _cache_alerts
    ok = _sb_upsert_alert_retry(a, max_tries=3)
    if not ok:
        print(f"[alerts] ⚠️ CRITICAL: fired state برای {a.get('id')} بعد از چند تلاش تو Supabase ذخیره نشد!")
    # بک‌آپ محلی فوری و مستقل از Supabase — حتی اگه Supabase کامل قطع باشه
    try:
        _local_mark_fired_backup(a)
    except Exception as e:
        print(f"[alerts] local fired backup error: {e}")
    # به‌جای پاک کردن کامل کش (که باعث یه fetch کامل و پرهزینه‌ی بعدی از
    # Supabase می‌شد — شامل کل آرشیو تاریخی، و از هر جای برنامه‌ای که بعدش
    # load_alerts() صدا بزنه، حتی poll_telegram که تقریباً پیوسته در حال اجراست)،
    # مستقیم خودِ کش حافظه رو اصلاح می‌کنیم: این آلارم از فعال به آرشیو منتقل
    # می‌شه، بدون این‌که نیاز به خوندن دوباره از Supabase باشه.
    with _alerts_cache_lock:
        if _cache_alerts is not None:
            _cache_alerts["alerts"] = [x for x in _cache_alerts.get("alerts", []) if x.get("id") != a.get("id")]
            _cache_alerts.setdefault("archive", []).append(a)

def _sb_upsert_alert_retry(a, max_tries=3):
    """
    _sb_upsert_alert رو با چند تلاش صدا بزن و مطمئن شو fired_at واقعاً ثبت شده —
    با یه GET تأیید می‌کنیم، نه فقط status code ارسال.
    """
    if not SUPABASE_KEY:
        return True  # بدون Supabase چیزی برای تایید نیست، local backup کافیه
    for attempt in range(1, max_tries + 1):
        _sb_upsert_alert(a)
        try:
            r = requests.get(
                f"{SUPABASE_URL}/rest/v1/alerts?id=eq.{a['id']}&select=fired_at,active",
                headers=_sb_h(), timeout=8)
            if r.status_code == 200:
                rows = r.json()
                if rows and rows[0].get("fired_at"):
                    return True
        except Exception as e:
            print(f"[alerts] verify fired attempt {attempt} exc: {e}")
        print(f"[alerts] fired upsert تلاش {attempt}/{max_tries} تایید نشد — دوباره امتحان می‌کنیم")
        time.sleep(0.5)
    return False

def _local_mark_fired_backup(a):
    """
    یه فایل کوچیک محلی جدا از alerts.json که فقط آیدی آلارم‌های fired رو نگه می‌داره —
    حتی اگه Supabase کامل قطع باشه، بعد از ری‌استارت این فایل خونده می‌شه تا
    از فایر مجدد جلوگیری بشه (لایه‌ی دفاعی دوم).
    """
    fired_ids = set()
    if os.path.exists(FIRED_BACKUP_FILE):
        try:
            with open(FIRED_BACKUP_FILE, "r", encoding="utf-8") as f:
                fired_ids = set(json.load(f))
        except Exception:
            fired_ids = set()
    fired_ids.add(str(a["id"]))
    try:
        with open(FIRED_BACKUP_FILE, "w", encoding="utf-8") as f:
            json.dump(list(fired_ids), f)
    except Exception as e:
        print(f"[alerts] local fired backup write error: {e}")

# =====================================================================
# Supabase
# =====================================================================
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://erwimqqskkzcsayvhxot.supabase.co")
APP_BASE_URL = os.environ.get("APP_BASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

def _sb_h():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }

def log_error(msg):
    try:
        data = load_alerts()
        errs = data.get("errors", [])
        errs.append({"time": now_teh(), "msg": str(msg)})
        data["errors"] = errs[-20:]
        save_alerts(data)
    except:
        pass
    print(f"[ERR] {msg}")

def is_forex_market_open():
    now_utc = datetime.utcnow()
    wd = now_utc.weekday()
    if wd == 5: return False
    if wd == 6: return now_utc.hour >= 21
    return True

H = {"User-Agent": "Mozilla/5.0 (compatible; PriceBot/1.0)"}
_last_known = {}

def get_forex_prices_batch(symbols):
    if not symbols: return {}
    clean = [s.upper().replace("/", "").replace(" ", "") for s in symbols]
    qs = "&".join(f"symbols={s}" for s in clean)
    url = f"https://biquote.io/api/latest?{qs}"
    try:
        r = requests.get(url, timeout=12, headers=H)
        r.raise_for_status()
        raw = r.json()
        result = {}
        if isinstance(raw, list):
            for item in raw:
                sym = item.get("symbol","").upper().replace("/","")
                bid = item.get("bid") or item.get("price") or item.get("last")
                if sym and bid and float(bid) > 0:
                    result[sym] = float(bid)
                    _last_known[sym] = {"price": float(bid), "ts": now_teh(), "stale": False}
        elif isinstance(raw, dict):
            for sym, data in raw.items():
                if isinstance(data, dict):
                    bid = data.get("bid") or data.get("price") or data.get("last")
                elif isinstance(data, (int, float)):
                    bid = data
                else:
                    bid = None
                if bid and float(bid) > 0:
                    result[sym.upper()] = float(bid)
                    _last_known[sym.upper()] = {"price": float(bid), "ts": now_teh(), "stale": False}
        if result: return result
    except Exception: pass
    result = {}
    for sym in clean:
        if sym in _last_known:
            cached = _last_known[sym]
            _last_known[sym]["stale"] = True
            result[sym] = cached["price"]
        else:
            try:
                base, quote = sym[:3], sym[3:6]
                r3 = requests.get(f"https://api.frankfurter.app/latest?from={base}&to={quote}", timeout=7)
                if r3.ok:
                    rate = r3.json().get("rates", {}).get(quote)
                    if rate:
                        result[sym] = float(rate)
                        _last_known[sym] = {"price": float(rate), "ts": now_teh(), "stale": False}
            except Exception: pass
    return result

def get_forex_price(symbol):
    sym = symbol.upper().replace("/","").replace(" ","")
    batch = get_forex_prices_batch([sym])
    return batch.get(sym)

CG_MAP = {
    "BTC":"bitcoin","ETH":"ethereum","BNB":"binancecoin","SOL":"solana",
    "XRP":"ripple","ADA":"cardano","DOGE":"dogecoin","TRX":"tron",
    "TON":"toncoin","AVAX":"avalanche-2","LINK":"chainlink","DOT":"polkadot",
    "MATIC":"matic-network","UNI":"uniswap","ATOM":"cosmos","LTC":"litecoin",
    "SHIB":"shiba-inu","OP":"optimism","ARB":"arbitrum","NEAR":"near",
}

def _cg_price(base):
    gid = CG_MAP.get(base)
    if not gid: return None
    try:
        d = requests.get(f"https://api.coingecko.com/api/v3/simple/price?ids={gid}&vs_currencies=usd", headers=H, timeout=8).json()
        return float(d[gid]["usd"])
    except:
        return None

def get_crypto_price(symbol):
    base = symbol.upper()
    for s in ["USDT","USDC","USD","BUSD"]:
        base = base.replace(s,"")
    base = base.replace("/","").strip()
    try:
        r = requests.get(f"https://biquote.io/api/latest?symbols={base}USD", timeout=8, headers=H)
        if r.ok:
            raw = r.json()
            bid = None
            if isinstance(raw, list) and raw:
                bid = raw[0].get("bid") or raw[0].get("price") or raw[0].get("last")
            elif isinstance(raw, dict):
                bid = raw.get("bid") or raw.get("price") or raw.get("last")
            if bid and float(bid) > 100:
                return float(bid)
    except Exception: pass
    sources = [
        ("OKX", lambda: float(requests.get(f"https://www.okx.com/api/v5/market/ticker?instId={base}-USDT", headers=H).json()["data"][0]["last"])),
        ("Binance-USDT", lambda: float(requests.get(f"https://api.binance.com/api/v3/ticker/price?symbol={base}USDT", headers=H).json()["price"])),
    ]
    for name, fn in sources:
        try:
            p = fn()
            if p and p > 0:
                return float(p)
        except: pass
    log_error(f"Crypto price failed for {symbol}")
    return None

def get_price(symbol, asset_type):
    if asset_type == "crypto":
        return get_crypto_price(symbol)
    return get_forex_price(symbol)

# ── ردیابی پیام‌های ربات برای پاک‌سازی چت ────────────────────
# { chat_id: [msg_id, msg_id, ...] }  — فقط پیام‌های غیر-fire
_bot_msg_ids: dict = {}
_BOT_MSG_MAX = 200  # حداکثر تعداد id هر چت

# ── ردیابی پیام‌های fired alert در همه چت‌ها ──────────────────
# { alert_id: { chat_id: message_id, ... } }
_fired_msg_ids: dict = {}

def _sb_save_fired_msgs(alert_id: str, cid_to_mid: dict):
    """map چت‌آیدی → مسیج‌آیدی یه آلارم fired رو توی Supabase ذخیره کن"""
    if not SUPABASE_KEY: return
    try:
        payload = {
            "id": alert_id,
            "msg_map": json.dumps(cid_to_mid),
            "created_at": now_teh()
        }
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/fired_msgs",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload, timeout=8)
        if r.status_code not in (200, 201, 204):
            print(f"[fired_msgs] save error: {r.status_code} {r.text[:80]}")
    except Exception as e:
        print(f"[fired_msgs] save exc: {e}")

def _sb_load_fired_msgs():
    """همه fired_msgs رو از Supabase بخون و توی _fired_msg_ids لود کن"""
    if not SUPABASE_KEY: return
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/fired_msgs?select=*&limit=2000",
            headers=_sb_h(), timeout=10)
        if r.status_code == 200:
            for row in r.json():
                aid = row.get("id")
                raw = row.get("msg_map", "{}")
                if aid:
                    _fired_msg_ids[aid] = json.loads(raw) if isinstance(raw, str) else (raw or {})
            print(f"[fired_msgs] لود شد — {len(_fired_msg_ids)} آلارم")
        else:
            print(f"[fired_msgs] load error: {r.status_code} {r.text[:80]}")
    except Exception as e:
        print(f"[fired_msgs] load exc: {e}")

def _sb_delete_fired_msgs(alert_id: str):
    """fired_msgs یه آلارم رو از Supabase پاک کن"""
    if not SUPABASE_KEY: return
    try:
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/fired_msgs?id=eq.{alert_id}",
            headers=_sb_h(), timeout=8)
    except Exception as e:
        print(f"[fired_msgs] delete exc: {e}")

# ── شمارنده هشتگ نماد ────────────────────────────────────────
# { "XAUUSD": 12, "BTCUSDT": 3, ... }  — در حافظه cache
_sym_counters: dict = {}

def _sb_next_sym_counter(sym: str) -> int:
    """
    شمارنده نماد رو یکی افزایش بده و مقدار جدید رو برگردون.
    اگه جدول symbol_counters نداشتیم از حافظه استفاده میکنه.
    """
    global _sym_counters
    if SUPABASE_KEY:
        try:
            # خوندن مقدار فعلی
            r = requests.get(
                f"{SUPABASE_URL}/rest/v1/symbol_counters?id=eq.{sym}&select=counter",
                headers=_sb_h(), timeout=6)
            if r.status_code == 200 and r.json():
                cur = int(r.json()[0]["counter"])
            else:
                cur = _sym_counters.get(sym, 0)
            new_val = cur + 1
            # upsert
            requests.post(
                f"{SUPABASE_URL}/rest/v1/symbol_counters",
                headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
                json={"id": sym, "counter": new_val}, timeout=6)
            _sym_counters[sym] = new_val
            return new_val
        except Exception as e:
            print(f"[counter] exc: {e}")
    # fallback حافظه
    _sym_counters[sym] = _sym_counters.get(sym, 0) + 1
    return _sym_counters[sym]

def _sb_load_sym_counters():
    """همه شمارنده‌ها رو از Supabase لود کن"""
    if not SUPABASE_KEY: return
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/symbol_counters?select=*",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200:
            for row in r.json():
                _sym_counters[row["id"]] = int(row["counter"])
            print(f"[counter] لود شد — {len(_sym_counters)} نماد")
    except Exception as e:
        print(f"[counter] load exc: {e}")

def _make_alarm_tag(sym: str) -> str:
    """هشتگ منحصربه‌فرد برای آلارم — مثلاً #XAUUSD7"""
    n = _sb_next_sym_counter(sym)
    return f"#{sym}{n}"

def _extract_symbol_from_tag(tag: str) -> str:
    """از تگ آلارم (مثلاً #XAUUSD7) نماد رو استخراج کن (XAUUSD)"""
    m = re.match(r'^#?([A-Za-z]+)\d*$', tag or "")
    return m.group(1) if m else ""

# =====================================================================
# 📋 سیستم تقسیم مسئولیت آلارم‌ها بین اعضای تیم (بدون شیفت — رندوم عادلانه)
# =====================================================================

# همه اعضای تیم + آیدی عددی تلگرام‌شون (برای تشخیص دکمه‌ی مخصوص «دیدم»)
TEAM_MEMBERS = ["اتابک", "مهران", "علی", "پیمان", "مسعود"]

TEAM_MEMBER_IDS = {
    "مسعود": "109419675",
    "پیمان": "5089282711",
    "علی":   "138116289",
    "اتابک": "349639401",
    "مهران": "117005892",
}
# معکوس: آیدی → اسم (برای چک سریع از from.id در callback)
TEAM_ID_TO_NAME = {v: k for k, v in TEAM_MEMBER_IDS.items()}

# نام کسی که در حالت فعال «ترجیح داده نمی‌شه» — این خودش همیشه ثابت می‌مونه،
# اما این‌که فعال باشه یا نه از پنل وب قابل تغییره (نگاه کن به _deprioritize_masoud_active)
DEPRIORITIZED_MEMBER = "مسعود"
# بازه ساعتی تهران که این عضو تقریباً هیچ‌وقت نباید بگیره مگه بقیه مشغول باشن
DEPRIORITIZED_BLOCK_START = 8
DEPRIORITIZED_BLOCK_END = 12

# آیا اولویت پایین مسعود فعاله؟ در startup از Supabase لود می‌شه و از پنل وب
# قابل تغییره. پیش‌فرض True تا رفتار فعلی حفظ بشه.
_deprioritize_masoud_active: bool = True
_deprioritize_masoud_lock = threading.Lock()

def _get_deprioritize_masoud() -> bool:
    with _deprioritize_masoud_lock:
        return _deprioritize_masoud_active

def _set_deprioritize_masoud(value: bool):
    global _deprioritize_masoud_active
    with _deprioritize_masoud_lock:
        _deprioritize_masoud_active = value
    _sb_save_deprioritize_masoud(value)

# اسم اعضایی که موقتاً از چرخه‌ی تقسیم آلارم کنار گذاشته شدن (مرخصی/غیره).
# پیش‌فرض خالی (همه در دسترس) تا رفتار فعلی حفظ بشه.
_unavailable_members: set = set()
_unavailable_members_lock = threading.Lock()

def _get_unavailable_members() -> set:
    with _unavailable_members_lock:
        return set(_unavailable_members)

def _set_member_availability(name: str, available: bool):
    global _unavailable_members
    with _unavailable_members_lock:
        if available:
            _unavailable_members.discard(name)
        else:
            _unavailable_members.add(name)
        snapshot = sorted(_unavailable_members)
    _sb_save_unavailable_members(snapshot)

# =====================================================================
# 🌐 پنل ادمین وب — /admin-panel
# =====================================================================
ADMIN_PANEL_PASSWORD = os.environ.get("ADMIN_PANEL_PASSWORD", "")
# session tokenهای معتبر — در حافظه، با انقضای ۱۲ ساعته
_admin_sessions: dict = {}  # token → expiry_timestamp
_admin_sessions_lock = threading.Lock()
ADMIN_SESSION_TTL_SECONDS = 12 * 3600

def _create_admin_session() -> str:
    token = secrets.token_urlsafe(32)
    with _admin_sessions_lock:
        _admin_sessions[token] = time.time() + ADMIN_SESSION_TTL_SECONDS
    return token

def _validate_admin_session(token: str) -> bool:
    if not token:
        return False
    with _admin_sessions_lock:
        expiry = _admin_sessions.get(token)
        if expiry is None:
            return False
        if time.time() > expiry:
            del _admin_sessions[token]
            return False
        return True

def _require_admin_session():
    """چک هدر Authorization برای endpointهای پنل وب — برمی‌گردونه None اگه معتبر بود، وگرنه یه Flask response خطا"""
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.replace("Bearer ", "").strip()
    if not _validate_admin_session(token):
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    return None

def _is_weekend_tehran():
    """شنبه (5) و یکشنبه (6) — فقط برای نمایش/گزارش، دیگه تاثیری در تقسیم نداره"""
    return datetime.now(TEHRAN).weekday() in (5, 6)

# { member_name: count_of_active_assignments }  — in-memory cache
# این شمارش فقط برای false-handling و آمار لحظه‌ای استفاده می‌شه، نه برای تصمیم تقسیم
_active_assign_count: dict = {}

# { member_name: count_of_alarms_received_this_week } — معیار اصلی تقسیم عادلانه.
# هر آلارمی که به کسی داده بشه (چه فعال بمونه چه بعداً false بشه) اینجا شمرده می‌شه
# و تا شنبه‌ی هفته‌ی بعد صفر نمی‌شه. این یعنی کسی که همین الان false کرده، دیگه
# «صفر» حساب نمی‌شه صرفاً چون آلارمش تموم شده — باید صبر کنه تا نوبت واقعی‌ش بشه،
# و اگه یه روز آلارم کمی گیرش اومد، بقیه‌ی هفته جبران می‌شه.
_daily_assign_count: dict = {}
_daily_assign_date: str = ""  # کلید هفته (تاریخ شنبه‌ی همون هفته) که شمارش هفتگی براش معتبره
_daily_assign_lock = threading.Lock()

def _tehran_today_str() -> str:
    return datetime.now(TEHRAN).strftime("%Y-%m-%d")

def _tehran_week_start_str() -> str:
    """تاریخ شنبه‌یِ همین هفته (شروع هفته‌ی ایرانی) به‌عنوان کلید شمارش هفتگی"""
    now = datetime.now(TEHRAN)
    days_since_saturday = (now.weekday() - 5) % 7  # weekday(): دوشنبه=0 ... شنبه=5
    week_start = now - timedelta(days=days_since_saturday)
    return week_start.strftime("%Y-%m-%d")

def _ensure_daily_reset():
    """شمارشِ عادلانه دیگه روزانه نیست، هفتگیه: اگه هفته عوض شده (رسیدیم به شنبه‌ی جدید)
    شمارش صفر می‌شه. تا وقتی هفته عوض نشده، شمارش تجمیعی می‌مونه — همین باعث می‌شه
    کسی که یه روز سبک (آلارم کم) گیرش اومده، بقیه‌ی هفته جبران بشه و در کل هفته
    تقریباً به تعداد مساوی با بقیه آلارم بگیره."""
    global _daily_assign_count, _daily_assign_date
    week_key = _tehran_week_start_str()
    with _daily_assign_lock:
        if _daily_assign_date != week_key:
            _daily_assign_count = {}
            _daily_assign_date = week_key

def _bump_daily_count(name: str):
    _ensure_daily_reset()
    with _daily_assign_lock:
        _daily_assign_count[name] = _daily_assign_count.get(name, 0) + 1

def _get_daily_counts(members: list) -> dict:
    _ensure_daily_reset()
    with _daily_assign_lock:
        return {m: _daily_assign_count.get(m, 0) for m in members}

def _rebuild_daily_assign_count():
    """بعد از ری‌استارت، شمارش هفتگی رو از Supabase بازسازی کن — بدون این
    کار، بعد از هر دیپلوی عدالتِ هفتگی از صفر شروع می‌شد و بی‌معنی می‌شد."""
    global _daily_assign_count, _daily_assign_date
    rows = _sb_load_today_assignments()
    counts = {}
    for row in rows:
        name = row.get("assigned_to", "")
        if name:
            counts[name] = counts.get(name, 0) + 1
    with _daily_assign_lock:
        _daily_assign_count = counts
        _daily_assign_date = _tehran_week_start_str()
    print(f"[assign] شمارش هفتگی بازسازی شد: {counts}")

# جلوگیری از double-handover: startup و scheduler هر کدوم فقط یه بار اجرا کنن
# جلوگیری از race condition در /False — اگه یه آلارم داره false میشه، دیگران صبر کنن
_false_in_progress: set = set()
_false_in_progress_lock = threading.Lock()

# جلوگیری از دابل-ack — اگه دکمه «دیدم» دوبار کلیک بشه (دابل‌تپ یا کندی شبکه)
# قبل از این‌که پیام ادیت و دکمه پاک بشه، این از ثبت تکراری جلوگیری می‌کنه
_ack_done: set = set()
_ack_lock = threading.Lock()

# ─── Supabase helpers ────────────────────────────────────────────────

def _sb_save_assignment(alarm_id: str, alarm_tag: str, assignee: str, shift: str, fired_at: str,
                        symbol: str = "", target_price: float = 0, created_by: str = ""):
    """ذخیره/آپدیت assignment در Supabase"""
    if not SUPABASE_KEY: return
    try:
        payload = {
            "id": alarm_id,
            "alarm_tag": alarm_tag,
            "assigned_to": assignee,
            "shift": shift,
            "is_active": True,
            "fired_at": fired_at,
            "false_at": None,
            "false_by": None,
            "symbol": symbol,
            "target_price": target_price,
            "created_by": created_by
        }
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload, timeout=8)
        if r.status_code not in (200, 201, 204):
            print(f"[assign] save error: {r.status_code} {r.text[:80]}")
    except Exception as e:
        print(f"[assign] save exc: {e}")

def _sb_false_assignment(alarm_id: str, false_by: str, reason: str = ""):
    """
    وقتی /False زده میشه — assignment رو غیرفعال کن.
    اگه قبلاً false شده بود، یه ردیف history جدید append کن (به جای overwrite).
    """
    if not SUPABASE_KEY: return
    try:
        # خوندن وضعیت فعلی برای گرفتن history قبلی
        r_get = requests.get(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{alarm_id}&select=is_active,false_history",
            headers=_sb_h(), timeout=8)
        prev_history = []
        already_false = False
        if r_get.status_code == 200:
            rows = r_get.json()
            if rows:
                already_false = (rows[0].get("is_active") == False)
                prev_history = rows[0].get("false_history") or []
                if isinstance(prev_history, str):
                    try: prev_history = json.loads(prev_history)
                    except: prev_history = []

        new_entry = {"by": false_by, "at": now_teh(), "reason": reason}
        prev_history.append(new_entry)

        payload = {
            "is_active": False,
            "false_at": now_teh(),
            "false_by": false_by,
            "false_history": prev_history
        }
        if reason:
            payload["false_reason"] = reason
        r = requests.patch(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{alarm_id}",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            json=payload, timeout=8)
        if r.status_code not in (200, 204):
            print(f"[assign] false error: {r.status_code} {r.text[:80]}")
        return already_false  # True اگه قبلاً false شده بود (یعنی این آپدیته)
    except Exception as e:
        print(f"[assign] false exc: {e}")
        return False

def _sb_load_active_assignments():
    """لود همه assignment‌های فعال از Supabase"""
    if not SUPABASE_KEY: return []
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments?is_active=eq.true&select=*",
            headers=_sb_h(), timeout=10)
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"[assign] load exc: {e}")
    return []

def _sb_load_today_assignments():
    """
    لود همه assignment‌هایی که از شنبه‌ی همین هفته (نیمه‌شب تهران) تا الان
    fired_at داشتن — چه الان فعال باشن چه false شده باشن. برای بازسازی
    شمارش عادلانه‌ی هفتگی بعد از ری‌استارت لازمه.
    """
    if not SUPABASE_KEY: return []
    try:
        now = datetime.now(TEHRAN)
        days_since_saturday = (now.weekday() - 5) % 7
        week_start = (now - timedelta(days=days_since_saturday)).replace(hour=0, minute=0, second=0, microsecond=0)
        week_start_str = week_start.strftime("%Y-%m-%d %H:%M:%S")
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments?fired_at=gte.{week_start_str}&select=assigned_to,fired_at",
            headers=_sb_h(), timeout=10)
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"[assign] load week exc: {e}")
    return []

def _rebuild_active_assign_count(rows):
    """بازسازی count از Supabase rows — فقط assigned آلارم‌ها"""
    global _active_assign_count
    _active_assign_count = {}
    for row in rows:
        name = row.get("assigned_to", "")
        if name:
            _active_assign_count[name] = _active_assign_count.get(name, 0) + 1
    print(f"[assign] active counts: {_active_assign_count}")

# ─── انتخاب مسئول (رندوم عادلانه، بدون شیفت) ─────────────────────────

def _pick_assignee(members: list) -> str:
    """
    اولویت اصلی: کسی که همین الان هیچ آلارم فعالی نداره (idle) — یعنی به
    کسی که یه آلارم دستشه، آلارم دوم نمی‌دیم مگر این‌که همه مشغول باشن و
    مجبور باشیم. بین idle ها، کسی که این هفته (از شنبه، نیمه‌شب تهران)
    کمترین آلارم دریافت کرده انتخاب می‌شه — تا هم کسی دوتا نگیره وقتی یکی
    دیگه بی‌کاره، هم در کل هفته شمارش‌ها تقریباً برابر بمونه. اگه یه روز
    کم‌کار بود، بقیه‌ی هفته (وقتی idle باشه) جبران می‌شه.

    اگه همه مشغول باشن (هیچ‌کس idle نیست)، آلارم بی‌صاحب نمی‌مونه —
    بین کسایی که کمترین شمارش هفتگی رو دارن (حتی اگه مشغولن) تقسیم می‌شه.

    استثنا: عضو DEPRIORITIZED_MEMBER تا حد امکان کنار گذاشته می‌شه —
    فقط وقتی انتخاب می‌شه که واقعاً نوبتش شده. تو بازه‌ی
    DEPRIORITIZED_BLOCK_START تا DEPRIORITIZED_BLOCK_END فقط زمانی بهش
    می‌دیم که همه‌ی بقیه‌ی اعضا خودشون یه آلارم فعال داشته باشن (مجبوریم).
    """
    import random
    if not members:
        return ""

    daily_counts = _get_daily_counts(members)
    active_counts = {m: _active_assign_count.get(m, 0) for m in members}

    def _finalize(chosen: str) -> str:
        _active_assign_count[chosen] = _active_assign_count.get(chosen, 0) + 1
        _bump_daily_count(chosen)
        return chosen

    def _choose_among(pool: list) -> str:
        """
        اول بین کسایی که همین الان هیچ آلارم فعالی ندارن (idle) انتخاب کن —
        حتی اگه شمارش هفتگی‌شون از یکی که مشغوله بیشتر باشه. یعنی به کسی که
        همین الان یه آلارم دستشه، یه آلارم دومی نمی‌دیم مگر این‌که واقعاً
        مجبور باشیم (همه مشغول باشن). بین idle ها هم، کسی که شمارش هفتگی‌ش
        کمتره اولویت داره تا عدالت هفتگی هم رعایت بشه.
        """
        idle = [m for m in pool if active_counts[m] == 0]
        if idle:
            min_daily_idle = min(daily_counts[m] for m in idle)
            tied_idle = [m for m in idle if daily_counts[m] == min_daily_idle]
            return random.choice(tied_idle)
        # همه مشغولن — مجبوریم؛ بین کمترین شمارش هفتگی انتخاب کن
        min_daily = min(daily_counts[m] for m in pool)
        tied = [m for m in pool if daily_counts[m] == min_daily]
        return random.choice(tied)

    # اگه اولویت پایین مسعود از پنل خاموش شده باشه، دیگه استثنا قائل نشو —
    # دقیقاً مثل بقیه‌ی اعضا وارد چرخه‌ی عادلانه‌ی خالص می‌شه
    deprioritize_on = _get_deprioritize_masoud()

    others = [m for m in members if m != DEPRIORITIZED_MEMBER]
    has_deprioritized = deprioritize_on and (DEPRIORITIZED_MEMBER in members)

    h_now = datetime.now(TEHRAN).hour
    in_block_hours = DEPRIORITIZED_BLOCK_START <= h_now < DEPRIORITIZED_BLOCK_END

    if has_deprioritized and others:
        min_others_daily = min(daily_counts[m] for m in others)
        min_dep_daily = daily_counts[DEPRIORITIZED_MEMBER]
        # «مجبوریم» یعنی همه‌ی بقیه همین الان یه آلارم فعال دارن — آلارم
        # نباید بی‌صاحب بمونه، پس چاره‌ای جز دادن به یکی نیست
        everyone_else_busy = all(active_counts[m] >= 1 for m in others)

        if in_block_hours:
            forced = everyone_else_busy
            if not forced:
                return _finalize(_choose_among(others))
            return _finalize(_choose_among(members))
        else:
            if everyone_else_busy:
                # مجبوریم بدیم — بین کمترین شمارش روزانه انتخاب کن (شاید مسعود باشه)
                return _finalize(_choose_among(members))
            elif min_others_daily <= min_dep_daily:
                return _finalize(_choose_among(others))
            else:
                return _finalize(_choose_among(members))

    # حالت عادی (بدون عضو دپریوریتایزد در لیست) — عادلانه‌ی روزانه‌ی خالص
    return _finalize(_choose_among(members))

def _get_assignee_for_alarm(alarm_id: str, alarm_tag: str, fired_at: str,
                            symbol: str = "", target_price: float = 0, created_by: str = "") -> tuple:
    """
    مسئول آلارم رو بین کل اعضای تیم به‌صورت رندومِ عادلانه تعیین کن —
    مستقل از ساعت/شیفت. هر آلارم تیمی همیشه یه مسئول می‌گیره.
    اعضایی که از پنل ادمین موقتاً «غیرفعال» (مرخصی/در دسترس نیستن) شدن،
    از این چرخه کنار گذاشته می‌شن — مگر این‌که همه غیرفعال باشن (اون موقع
    برای این‌که آلارم بی‌صاحب نمونه، بین کل اعضا تقسیم می‌شه).
    """
    unavailable = _get_unavailable_members()
    pool = [m for m in TEAM_MEMBERS if m not in unavailable] or TEAM_MEMBERS
    assignee = _pick_assignee(pool)
    threading.Thread(
        target=_sb_save_assignment,
        args=(alarm_id, alarm_tag, assignee, "", fired_at),
        kwargs={"symbol": symbol, "target_price": target_price, "created_by": created_by},
        daemon=True
    ).start()
    return (assignee, "")

# ─── startup: بازسازی state از Supabase ─────────────────────────────

def _sb_restore_on_startup():
    """
    بعد از هر restart، شمارش آلارم‌های فعال هر عضو رو از Supabase بازسازی کن
    تا تقسیم رندومِ عادلانه بدون از دست دادن state ادامه پیدا کنه.
    دیگه هیچ شیفت/handover/scheduler‌ای وجود نداره.
    """
    global _deprioritize_masoud_active, _unavailable_members
    rows = _sb_load_active_assignments()
    _rebuild_active_assign_count(rows)
    _rebuild_daily_assign_count()
    with _deprioritize_masoud_lock:
        _deprioritize_masoud_active = _sb_load_deprioritize_masoud()
    with _unavailable_members_lock:
        _unavailable_members = set(_sb_load_unavailable_members())
    print(f"[assign] startup: {len(rows)} آلارم active از Supabase بازسازی شد — "
          f"اولویت پایین مسعود: {'فعال' if _deprioritize_masoud_active else 'غیرفعال'} — "
          f"غیرفعال‌ها: {sorted(_unavailable_members) or 'هیچ‌کس'}")



def _track_msg(chat_id: str, msg_id: int):
    """id پیام ربات رو ذخیره کن (غیر از fired alerts)"""
    cid = str(chat_id)
    if cid not in _bot_msg_ids:
        _bot_msg_ids[cid] = []
    _bot_msg_ids[cid].append(msg_id)
    # سقف حافظه
    if len(_bot_msg_ids[cid]) > _BOT_MSG_MAX:
        _bot_msg_ids[cid] = _bot_msg_ids[cid][-_BOT_MSG_MAX:]

def delete_chat_history(token: str, chat_id: str):
    """همه پیام‌های track‌شده رو پاک کن — آلارم‌های fire دست‌نخورده می‌مونن"""
    cid = str(chat_id)
    ids = _bot_msg_ids.pop(cid, [])
    deleted = 0
    for mid in ids:
        try:
            requests.post(
                f"https://api.telegram.org/bot{token}/deleteMessage",
                json={"chat_id": cid, "message_id": mid},
                timeout=5, headers=H)
            deleted += 1
        except: pass
    return deleted

def send_tg(token, chat_id, text):
    try:
        r = requests.post(f"https://api.telegram.org/bot{token}/sendMessage", json={"chat_id": str(chat_id), "text": text, "parse_mode": "HTML"}, timeout=10, headers=H)
        mid = r.json().get("result", {}).get("message_id")
        if mid:
            _track_msg(str(chat_id), mid)
        return r.status_code == 200
    except: return False

def broadcast(token, chat_ids, text):
    return [send_tg(token, c, text) for c in chat_ids]

def send_reply_keyboard(token, chat_id, text, rows):
    """ارسال پیام با Reply Keyboard (جای کیبورد موبایل)"""
    try:
        markup = {
            "keyboard": rows,
            "resize_keyboard": True,
            "one_time_keyboard": False,
            "input_field_placeholder": "یه گزینه انتخاب کن..."
        }
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": str(chat_id), "text": text,
                  "parse_mode": "HTML", "reply_markup": markup},
            timeout=10, headers=H)
        mid = r.json().get("result", {}).get("message_id")
        if mid:
            _track_msg(str(chat_id), mid)
        return mid
    except: return None

def edit_reply_keyboard(token, chat_id, message_id, text, rows=None):
    """ادیت پیام با Reply Keyboard یا بدون keyboard — برای flow ثبت آلارم"""
    try:
        if rows is not None:
            markup = {"keyboard": rows, "resize_keyboard": True, "one_time_keyboard": False}
        else:
            markup = {"remove_keyboard": True}
        # editMessageText نمیتونه reply_markup از نوع ReplyKeyboard داشته باشه
        # پس اول پیام رو ادیت میکنیم، بعد اگه keyboard جدید داریم یه پیام کمکی میفرستیم
        requests.post(
            f"https://api.telegram.org/bot{token}/editMessageText",
            json={"chat_id": str(chat_id), "message_id": message_id,
                  "text": text, "parse_mode": "HTML"},
            timeout=10, headers=H)
    except: pass

def _alarm_edit(token, cid, bot_msg_id, text):
    """ادیت پیام جاری flow آلارم (فقط متن، بدون keyboard)"""
    try:
        requests.post(
            f"https://api.telegram.org/bot{token}/editMessageText",
            json={"chat_id": str(cid), "message_id": bot_msg_id,
                  "text": text, "parse_mode": "HTML"},
            timeout=10, headers=H)
    except Exception as e:
        print(f"[alarm_edit] {e}")


    """حذف Reply Keyboard و ارسال پیام"""
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": str(chat_id), "text": text,
                  "parse_mode": "HTML",
                  "reply_markup": {"remove_keyboard": True}},
            timeout=10, headers=H)
        mid = r.json().get("result", {}).get("message_id")
        if mid:
            _track_msg(str(chat_id), mid)
        return mid
    except: return None

def send_tg_keyboard(token, chat_id, text, keyboard, track=True):
    """ارسال پیام با inline keyboard"""
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": str(chat_id), "text": text,
                  "parse_mode": "HTML", "reply_markup": {"inline_keyboard": keyboard}},
            timeout=10, headers=H)
        mid = r.json().get("result", {}).get("message_id")
        if mid and track:
            _track_msg(str(chat_id), mid)
        return mid
    except: return None

def edit_tg_keyboard(token, chat_id, message_id, text, keyboard):
    """ویرایش پیام با inline keyboard"""
    try:
        requests.post(
            f"https://api.telegram.org/bot{token}/editMessageText",
            json={"chat_id": str(chat_id), "message_id": message_id,
                  "text": text, "parse_mode": "HTML",
                  "reply_markup": {"inline_keyboard": keyboard}},
            timeout=10, headers=H)
    except: pass

def answer_callback(token, callback_id, text=""):
    """جواب به callback query"""
    try:
        requests.post(
            f"https://api.telegram.org/bot{token}/answerCallbackQuery",
            json={"callback_query_id": callback_id, "text": text},
            timeout=10, headers=H)
    except: pass

# ── reminder state ─────────────────────────────────────────────
# memory: { cid: { sym: {"interval": int, "active": bool, "tf_sec": int} } }
# Supabase: جدول reminders — هر ردیف یه reminder فعال
_reminders = {}

# ── candle-close helpers ──────────────────────────────────────
_TF_OFFSET = {300: 60, 900: 300, 3600: 900, 14400: 900}  # ثانیه قبل از کلوز
_TF_LABEL  = {300:"M5", 900:"M15", 3600:"H1", 14400:"H4"}
REMINDER_QUICK_SYMBOLS = ["XAUUSD", "EURUSD", "GBPUSD", "BTCUSDT", "ETHUSDT", "USDJPY"]

def _candle_info(tf_sec):
    """
    برمی‌گردونه: (wait_sec, close_teh_str, progress_pct)
    - اگه هنوز وقت هشدار این کندل نرسیده → صبر کن
    - اگه توی پنجره هشدار هستیم (بین alert_time و close) → فوری (۵ ثانیه)
    - اگه کندل بسته → کندل بعدی
    """
    offset = _TF_OFFSET.get(tf_sec, 300)
    now_utc = time.time()
    cur_close = (int(now_utc) // tf_sec + 1) * tf_sec
    alert_t   = cur_close - offset
    wait      = alert_t - now_utc
    if wait <= 0:
        if now_utc < cur_close:
            wait, use_close = 5, cur_close          # توی پنجره → فوری
        else:
            use_close = cur_close + tf_sec          # کندل بسته → بعدی
            wait = (use_close - offset) - now_utc
    else:
        use_close = cur_close
    from datetime import datetime as _dt
    close_teh = _dt.fromtimestamp(use_close, tz=TEHRAN).strftime("%H:%M")
    elapsed   = now_utc - (use_close - tf_sec)
    progress  = min(99, int(elapsed / tf_sec * 100))
    return max(int(wait), 5), close_teh, progress

# ── Supabase reminder helpers ────────────────────────────────
def _sb_save_reminder(cid, sym, interval_sec, tf_sec=0):
    """ذخیره یه reminder در Supabase"""
    if not SUPABASE_KEY: return
    rid = f"{cid}_{sym}"
    try:
        payload = {
            "id": rid, "chat_id": str(cid), "symbol": sym,
            "interval_sec": interval_sec, "tf_sec": tf_sec,
            "created_at": now_teh(), "active": True
        }
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/reminders",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload, timeout=8)
        if r.status_code not in (200,201,204):
            print(f"[reminder] save error: {r.status_code} {r.text[:60]}")
    except Exception as e:
        print(f"[reminder] save exc: {e}")

def _sb_delete_reminder(cid, sym):
    """حذف یه reminder از Supabase"""
    if not SUPABASE_KEY: return
    rid = f"{cid}_{sym}"
    try:
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/reminders?id=eq.{rid}",
            headers=_sb_h(), timeout=8)
    except Exception as e:
        print(f"[reminder] delete exc: {e}")

def _sb_delete_all_reminders(cid):
    """حذف همه reminder‌های یه کاربر از Supabase"""
    if not SUPABASE_KEY: return
    try:
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/reminders?chat_id=eq.{cid}",
            headers=_sb_h(), timeout=8)
    except Exception as e:
        print(f"[reminder] delete_all exc: {e}")

def _sb_load_reminders():
    """لود همه reminder‌های فعال از Supabase — برای startup"""
    if not SUPABASE_KEY: return []
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/reminders?active=eq.true&select=*",
            headers=_sb_h(), timeout=10)
        if r.status_code == 200:
            return r.json()
        print(f"[reminder] load error: {r.status_code}")
    except Exception as e:
        print(f"[reminder] load exc: {e}")
    return []

def _delete_msg_after(token, cid, msg_id, delay=120):
    """پیام رو بعد از delay ثانیه پاک کن"""
    def _do():
        time.sleep(delay)
        try:
            requests.post(
                f"https://api.telegram.org/bot{token}/deleteMessage",
                json={"chat_id": cid, "message_id": msg_id},
                timeout=10, headers=H)
        except: pass
    threading.Thread(target=_do, daemon=True).start()

# =====================================================================
# 📡 سیستم سیگنال
# =====================================================================

SIGNAL_CHANNEL = os.environ.get("SIGNAL_CHANNEL", "")  # مثلاً @mychannel یا chat_id

def _sb_next_signal_seq():
    """شماره سیگنال بعدی — از آخرین seq در جدول +1"""
    if not SUPABASE_KEY: return int(time.time()) % 100000
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/signals?select=seq&order=seq.desc&limit=1",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200:
            data = r.json()
            if data and data[0].get("seq") is not None:
                return int(data[0]["seq"]) + 1
        return 10001  # جدول خالیه
    except:
        return 10001

def _sb_save_signal(sig: dict):
    """ذخیره سیگنال در Supabase — فقط فیلدهای جدول"""
    if not SUPABASE_KEY: return
    # فقط فیلدهایی که در جدول signals وجود دارن
    allowed = {"id","seq","symbol","direction","entry","sl","tp1","tp2","tp3",
               "tf","risk_pips","rr","sent_by","sent_at","channel_msg_id","status","note"}
    clean = {k: v for k, v in sig.items() if k in allowed}
    # مقادیر None رو برای tp2/tp3 به null تبدیل کن
    for f in ("tp2","tp3","channel_msg_id","note"):
        if f not in clean:
            clean[f] = None
    try:
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/signals",
            headers={**_sb_h(), "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=clean, timeout=10)
        if r.status_code not in (200, 201, 204):
            print(f"[signal] save error: {r.status_code} {r.text[:200]}")
        else:
            print(f"[signal] saved OK: {clean.get('id')}")
    except Exception as e:
        print(f"[signal] save exc: {e}")

def _sb_update_signal(sig_id, patch: dict):
    """آپدیت یه فیلد سیگنال"""
    if not SUPABASE_KEY: return
    try:
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/signals?id=eq.{sig_id}",
            headers={**_sb_h(), "Prefer": "return=minimal"},
            json=patch, timeout=8)
    except: pass

def _sb_load_signals(limit=10):
    """آخرین سیگنال‌ها رو بخون"""
    if not SUPABASE_KEY: return []
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/signals?select=*&order=sent_at.desc&limit={limit}",
            headers=_sb_h(), timeout=8)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def _calc_signal(symbol: str, direction: str, entry: float, sl: float, rr: float = 1.5):
    """
    محاسبه TP بر اساس Entry، SL و ریوارد.
    direction: buy_limit / buy_stop / sell_limit / sell_stop
    برمیگردونه: (sl_calc, tp1, risk_pips)
    """
    is_buy = direction.startswith("buy")
    risk = abs(entry - sl)
    mul = get_pip_multiplier(symbol)
    risk_pips = round(risk * mul, 1)
    if is_buy:
        tp1 = round(entry + risk * rr, 5)
    else:
        tp1 = round(entry - risk * rr, 5)
    return sl, tp1, risk_pips

def _sl_from_pips(symbol: str, direction: str, entry: float, pips: float):
    """محاسبه SL از روی پیپ"""
    mul = get_pip_multiplier(symbol)
    dist = pips / mul
    is_buy = direction.startswith("buy")
    sl = round(entry - dist if is_buy else entry + dist, 5)
    return sl

def _fmt_signal_price(p, symbol=""):
    """فرمت عدد برای نمایش در سیگنال"""
    if p is None: return "—"
    v = float(p)
    su = symbol.upper()
    if any(x in su for x in ['BTC','ETH','SOL','BNB']):
        return f"{v:.1f}" if v > 1000 else f"{v:.4f}"
    if "XAU" in su or "XAG" in su: return f"{v:.2f}"
    if "JPY" in su: return f"{v:.3f}"
    return f"{v:.5f}"

def _build_signal_text(sig: dict) -> str:
    """ساخت متن سیگنال — عین فرمت درخواستی، اعداد قابل کپی"""
    sym     = sig.get("symbol","")
    d       = sig.get("direction","")
    entry   = sig.get("entry")
    sl      = sig.get("sl")
    tp1     = sig.get("tp1")
    tp2     = sig.get("tp2")
    tp3     = sig.get("tp3")
    tf      = sig.get("tf","H1")
    sig_id  = sig.get("id","")

    dir_map = {
        "buy_limit":  "✅ Buy limit",
        "buy_stop":   "✅ Buy stop",
        "sell_limit": "🔴 Sell limit",
        "sell_stop":  "🔴 Sell stop",
    }
    dir_txt = dir_map.get(d, d)

    def c(val):
        if val is None: return "<code>-</code>"
        return f"<code>{_fmt_signal_price(val, sym)}</code>"

    tp2_txt = c(tp2) if tp2 else "<code>-</code>"
    tp3_txt = c(tp3) if tp3 else "<code>-</code>"

    return (
        f"#{sig_id}\n"
        f"#{sym}\n"
        f"{dir_txt}\n"
        f"➡️ Entry: {c(entry)}\n"
        f"🛑 SL: {c(sl)}\n"
        f"🎯 TP:\n"
        f"TP1: {c(tp1)}\n"
        f"TP2: {tp2_txt}\n"
        f"TP3: {tp3_txt}\n"
        f"⏱ Timeframe: {tf}"
    )

def _build_signal_preview(sig: dict) -> str:
    """پیش‌نمایش سیگنال — عین چیزی که به کانال میره"""
    return _build_signal_text(sig)

# state ثبت سیگنال در حال ساخت
_pending_signal = {}  # cid → {"step": str, "data": dict, "bot_msg_id": int}

SIGNAL_QUICK_SYMBOLS = ["BTCUSDT", "XAUUSD", "EURUSD", "GBPUSD", "ETHUSDT"]
SIGNAL_DIRECTIONS = [
    ("✅ Buy Limit",  "buy_limit"),
    ("✅ Buy Stop",   "buy_stop"),
    ("🔴 Sell Limit", "sell_limit"),
    ("🔴 Sell Stop",  "sell_stop"),
]
SIGNAL_TF_OPTIONS = ["M5", "M15", "M30", "H1", "H4", "D1"]
SIGNAL_DEFAULT_TF = "H1"
SIGNAL_DEFAULT_RR = 1.5

def _show_signal_preview(token, cid, mid, data):
    """نمایش پیش‌نمایش سیگنال با دکمه‌های ویرایش"""
    tf  = data.get("tf", SIGNAL_DEFAULT_TF)
    note = data.get("note","")
    note_line = f"\n\n📝 <i>{note}</i>" if note else ""
    preview = f"<b>── پیش‌نمایش ──</b>\n\n{_build_signal_text(data)}{note_line}"
    kb = [
        [{"text": f"⏱ TF: {tf}", "callback_data": f"sig_tf:{cid}"},
         {"text": "🎯 TP2/TP3", "callback_data": f"sig_tp:{cid}"}],
        [{"text": "📝 یادداشت", "callback_data": f"sig_note:{cid}"},
         {"text": "🔄 ریوارد", "callback_data": f"sig_recalc:{cid}"}],
        [{"text": "📤 ارسال به گروه", "callback_data": f"sig_send:{cid}:channel"},
         {"text": "💾 ثبت در دیتا", "callback_data": f"sig_send:{cid}:dbonly"}],
        [{"text": "❌ لغو", "callback_data": f"sig_cancel:{cid}"}],
    ]
    edit_tg_keyboard(token, cid, mid, preview, kb)

def _send_reminder(token, cid, sym, tf_sec=0):
    """پیام هشدار کلوز کندل — حذف ۵ دقیقه بعد از ارسال"""
    tf_label = _TF_LABEL.get(tf_sec, "")
    if tf_sec and tf_label:
        _, close_teh, progress = _candle_info(tf_sec)
        offset_min = _TF_OFFSET.get(tf_sec, 300) // 60
        msg = (f"🕯 <b>کلوز کندل {tf_label} — {sym}</b>\n"
               f"━━━━━━━━━━━━━━━━━━\n"
               f"⏰ کلوز تهران: <b>{close_teh}</b>\n"
               f"⏳ مانده تا کلوز: <b>~{offset_min} دقیقه</b>\n"
               f"📊 پیشرفت کندل: <b>{progress}%</b>\n"
               f"━━━━━━━━━━━━━━━━━━\n"
               f"🗑 این پیام ۵ دقیقه دیگه حذف میشه.")
    else:
        msg = f"⚠️ <b>یادآوری:</b> <code>{sym}</code> بررسی بشه!\n\n🗑 این پیام ۵ دقیقه دیگه حذف میشه."
    kb = [[{"text": f"✕ کنسل {sym}", "callback_data": f"cancel_reminder_one:{cid}:{sym}"}]]
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": cid, "text": msg, "parse_mode": "HTML",
                  "reply_markup": {"inline_keyboard": kb}},
            timeout=10, headers=H)
        mid = r.json().get("result", {}).get("message_id")
        if mid:
            _delete_msg_after(token, cid, mid, delay=300)  # ۵ دقیقه
    except: pass

def _schedule_reminder(token, cid, sym, interval_sec, persist=True, tf_sec=0):
    """
    هشدار کلوز کندل — هر چرخه:
      1. محاسبه دقیق چند ثانیه تا هشدار کلوز بعدی
      2. sleep
      3. یه پیام بفرست
      4. برگرد به ۱ (برای کندل بعدی)
    tf_sec: طول تایم‌فریم (M5=300, M15=900, H1=3600, H4=14400)
    interval_sec: همون tf_sec هست (برای سازگاری Supabase)
    """
    tf = tf_sec or interval_sec
    if cid not in _reminders:
        _reminders[cid] = {}
    _reminders[cid][sym] = {"interval": tf, "active": True, "tf_sec": tf}
    entry = _reminders[cid][sym]
    if persist:
        threading.Thread(target=_sb_save_reminder, args=(cid, sym, tf, tf), daemon=True).start()
    def _loop():
        while entry.get("active") and _reminders.get(cid, {}).get(sym, {}).get("active"):
            wait, _, _ = _candle_info(entry["tf_sec"])
            time.sleep(wait)
            if not _reminders.get(cid, {}).get(sym, {}).get("active"):
                break
            _send_reminder(token, cid, sym, tf_sec=entry["tf_sec"])
            # بعد از ارسال پیام، صبر کن تا کلوز کندل رد بشه — جلوی تکرار رو میگیره
            tf = entry["tf_sec"]
            now_utc = time.time()
            cur_close = (int(now_utc) // tf + 1) * tf
            sleep_after = cur_close - now_utc + 5  # ۵ ثانیه بعد از کلوز
            if sleep_after > 0:
                time.sleep(sleep_after)
    threading.Thread(target=_loop, daemon=True).start()

def build_cancel_reminder_msg(cid):
    """لیست هشدارهای فعال با دکمه حذف جداگانه"""
    active = _reminders.get(cid, {})
    if not active:
        return "هیچ هشدار دوره‌ای فعالی نداری.", []
    labels = {300:"۵ دقیقه", 900:"۱۵ دقیقه", 3600:"۱ ساعت", 14400:"۴ ساعت"}
    lines = ["⏰ <b>هشدارهای دوره‌ای فعال:</b>\n"]
    keyboard = []
    for sym, info in active.items():
        tf = info.get("tf_sec", info.get("interval", 0))
        lbl = _TF_LABEL.get(tf, f"{tf//60}m") if tf else "؟"
        lines.append(f"• <b>{sym}</b> — کلوز {lbl}")
        keyboard.append([{"text": f"🗑 حذف {sym}", "callback_data": f"cancel_reminder_one:{cid}:{sym}"}])
    keyboard.append([{"text": "✕ کنسل همه", "callback_data": f"cancel_reminder_all:{cid}"}])
    keyboard.append([{"text": "✓ بستن", "callback_data": "close_myalerts"}])
    return "\n".join(lines), keyboard

def build_myalerts_msg(cid):
    """متن و keyboard لیست آلارم‌های شخصی"""
    alerts = load_alerts().get("alerts", [])
    my = [a for a in alerts if a.get("is_private") and str(a.get("private_cid","")) == cid and a.get("active")]
    if not my:
        return "📭 هیچ آلارم شخصی فعالی نداری.", []
    lines = ["🔒 <b>آلارم‌های شخصی تو:</b>"]
    keyboard = []
    for i, a in enumerate(my, 1):
        sym = a.get("symbol","")
        tgt = a.get("target_price",0)
        cond = "📈 BUY" if a.get("condition") == "below" else "📉 SELL"
        cur2 = a.get("last_price")
        cur_txt = f"<code>{fmt_price(cur2, sym)}</code>" if cur2 else "—"
        cmt = f"\n│  💬 {a['comment']}" if a.get("comment") else ""
        lines.append(
            f"┌─ {i}. <b>{sym}</b>  {cond}\n"
            f"│  🎯 هدف: <code>{tgt}</code>\n"
            f"│  💹 فعلی: {cur_txt}"
            f"{cmt}\n"
            f"└──────────────"
        )
        keyboard.append([{"text": f"🗑 حذف  {sym} @ {tgt}", "callback_data": f"del_confirm:{a['id']}"}])
    keyboard.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
    return "\n".join(lines), keyboard

def _get_token_and_cids():
    data = load_alerts()
    tg = data.get("telegram", {})
    token = BOT_TOKEN_ENV or tg.get("bot_token", "")
    cids = list(tg.get("chat_ids", []))
    leg = tg.get("chat_id", "")
    if leg and leg not in [str(x) for x in cids]:
        cids.append(leg)
    return token, cids, data


FF_NEWS_HOUR = int(os.environ.get("NEWS_HOUR", "7"))   # ساعت ارسال روزانه (تهران)
FF_NEWS_MINUTE = int(os.environ.get("NEWS_MINUTE", "0"))

def fetch_ff_news():
    """
    تقویم اقتصادی ForexFactory رو از RSS می‌گیره.
    فقط رویدادهای USD با impact بالا (⭐⭐⭐) برمی‌گردونه.
    """
    try:
        import xml.etree.ElementTree as ET
        from datetime import timezone
        r = requests.get(
            "https://nfs.faireconomy.media/ff_calendar_thisweek.json",
            headers={**H, "User-Agent": "Mozilla/5.0"},
            timeout=10)
        if r.status_code == 200:
            events = r.json()
        else:
            # fallback: RSS
            r2 = requests.get("https://www.forexfactory.com/ffcal_week_this.xml",
                headers={**H, "User-Agent": "Mozilla/5.0"}, timeout=10)
            if r2.status_code != 200:
                return None, "❌ دریافت داده از ForexFactory ناموفق بود"
            root = ET.fromstring(r2.content)
            events = []
            for ev in root.findall("event"):
                events.append({
                    "title": ev.findtext("title",""),
                    "country": ev.findtext("country",""),
                    "date": ev.findtext("date",""),
                    "time": ev.findtext("time",""),
                    "impact": ev.findtext("impact",""),
                    "forecast": ev.findtext("forecast",""),
                    "previous": ev.findtext("previous",""),
                })

        # فیلتر: فقط رویدادهای high/medium impact — همه ارزها — امروز
        today_teh = datetime.now(TEHRAN).strftime("%Y-%m-%d")
        high_events = []
        for ev in events:
            impact = ev.get("impact","").lower()
            if impact not in ("high","medium","3","2"):
                continue
            ev_date = ev.get("date","")
            try:
                from datetime import datetime as dt
                parsed = dt.strptime(ev_date, "%Y-%m-%dT%H:%M:%S%z")
                ev_date_teh = parsed.astimezone(TEHRAN).strftime("%Y-%m-%d")
                if ev_date_teh != today_teh:
                    continue
                ev["time_teh"] = parsed.astimezone(TEHRAN).strftime("%H:%M")
            except:
                if today_teh not in ev_date:
                    continue
                ev["time_teh"] = ev.get("time","—")
            high_events.append(ev)

        # مرتب‌سازی بر اساس ساعت
        high_events.sort(key=lambda x: x.get("time_teh","99:99"))

        if not high_events:
            return [], "📭 امروز رویداد مهم فارکس نداریم."

        return high_events, None

    except Exception as e:
        return None, f"❌ خطا: {e}"


def format_ff_message(events):
    """پیام تلگرام رو فرمت می‌کنه"""
    today_str = datetime.now(TEHRAN).strftime("%A %d %B %Y")
    lines = [f"📅 <b>تقویم اقتصادی فارکس — امروز</b>\n{today_str}\n"]
    for ev in events:
        impact = ev.get("impact","").lower()
        star = "🔴" if impact in ("high","3") else "🟡"
        time_str = ev.get("time_teh") or ev.get("time","—")
        title = ev.get("title","—")
        forecast = ev.get("forecast","—") or "—"
        previous = ev.get("previous","—") or "—"
        country = ev.get("country","").upper()
        flag_map = {"USD":"🇺🇸","EUR":"🇪🇺","GBP":"🇬🇧","JPY":"🇯🇵","CAD":"🇨🇦",
                    "AUD":"🇦🇺","NZD":"🇳🇿","CHF":"🇨🇭","CNY":"🇨🇳","GER":"🇩🇪"}
        flag = flag_map.get(country, "🌐")
        lines.append(
            f"{star} {flag} <b>{title}</b>\n"
            f"   🕐 {time_str} (تهران)\n"
            f"   پیش‌بینی: <b>{forecast}</b>  |  قبلی: {previous}"
        )
    return "\n\n".join(lines)


def daily_news_scheduler():
    """هر روز سر ساعت NEWS_HOUR تهران اخبار می‌فرسته + ۱۵ دقیقه قبل از هر خبر قرمز هشدار می‌فرسته"""
    sent_today = None
    global _today_red_events, _today_red_events_date, _news_reminder_sent
    # موقع استارت، وضعیت رو از Supabase بازیابی کن (اگه اپ ری‌استارت شده باشه)
    try:
        saved = _sb_load_news_reminder_state()
        if saved.get("date") == datetime.now(TEHRAN).strftime("%Y-%m-%d"):
            _today_red_events = saved.get("events", [])
            _today_red_events_date = datetime.now(TEHRAN).date()
            _news_reminder_sent = set(tuple(k) for k in saved.get("sent", []))
            print(f"[news] وضعیت هشدار از Supabase بازیابی شد — {len(_today_red_events)} خبر قرمز")
    except Exception as e:
        print(f"[news] بازیابی وضعیت ناموفق: {e}")

    while True:
        try:
            now = datetime.now(TEHRAN)
            today = now.date()
            if (now.hour == FF_NEWS_HOUR and now.minute == FF_NEWS_MINUTE
                    and sent_today != today):
                token, cids, _ = _get_token_and_cids()
                if token and cids:
                    events, err = fetch_ff_news()
                    if err and not events:
                        msg = err
                    else:
                        msg = format_ff_message(events) if events else "📭 امروز رویداد مهم فارکس نداریم."
                    broadcast(token, cids, msg)
                    sent_today = today
                    print(f"[news] ارسال شد — {len(events or [])} رویداد")
                    # فقط اخبار قرمز (impact بالا) رو برای هشدار ۱۵ دقیقه‌ای نگه می‌داریم
                    _today_red_events = [ev for ev in (events or [])
                                          if ev.get("impact","").lower() in ("high","3")
                                          and ev.get("country","").upper() == "USD"]
                    _today_red_events_date = today
                    _news_reminder_sent = set()
                    _sb_save_news_reminder_state({
                        "date": today.strftime("%Y-%m-%d"),
                        "events": _today_red_events,
                        "sent": [],
                    })

            _check_news_reminders(now)
        except Exception as e:
            print(f"[news_scheduler] {e}")
        time.sleep(50)


_today_red_events = []
_today_red_events_date = None
_news_reminder_sent = set()

def _check_news_reminders(now):
    """۱۵ دقیقه قبل از هر خبر قرمزِ امروز، یه هشدار به تیم می‌فرسته."""
    global _news_reminder_sent
    if _today_red_events_date != now.date() or not _today_red_events:
        return
    for ev in _today_red_events:
        time_teh = ev.get("time_teh")
        if not time_teh or ":" not in time_teh:
            continue
        key = (now.date().strftime("%Y-%m-%d"), ev.get("title",""), time_teh)
        if key in _news_reminder_sent:
            continue
        try:
            hh, mm = map(int, time_teh.split(":"))
            ev_dt = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
        except Exception:
            continue
        reminder_dt = ev_dt - timedelta(minutes=15)
        if reminder_dt <= now < reminder_dt + timedelta(seconds=60):
            token, cids, _ = _get_token_and_cids()
            if token and cids:
                msg = (
                    "🚨🔴⚠️🚨 هشدار خبر مهم 🚨⚠️🔴🚨\n\n"
                    f"📅 {ev.get('title','—')}\n"
                    f"⏰ ۱۵ دقیقه دیگر ({time_teh})\n\n"
                    "❗️ با توجه به پلن، تریدهای باز داخل حساب باید مدیریت شوند.\n\n"
                    "🔔 لطفاً هرکی این پیام رو می‌بینه به این مورد توجه کنه 🙏"
                )
                broadcast(token, cids, msg)
                print(f"[news_reminder] ارسال شد — {ev.get('title')} ساعت {time_teh}")
            _news_reminder_sent.add(key)
            _sb_save_news_reminder_state({
                "date": now.date().strftime("%Y-%m-%d"),
                "events": _today_red_events,
                "sent": [list(k) for k in _news_reminder_sent],
            })

_pending_name  = {}  # cid → True
_pending_alarm = {}  # cid → {"step": str, "data": dict}
_pending_reminder = {}  # cid → {step, bot_msg_id}
_pending_bulk_false = {}  # cid → {"bot_msg_id": ..., "count": ...} — منتظر متن دلیل فالس‌کردن دسته‌جمعی
_pending_weekly_search = {}  # cid → {"which": str, "msg_id": int}
# آلارم_آیدی → {chat_id: message_id آخرین پیام False/آپدیت}
_false_broadcast_ids: dict = {}

# ── Reply Keyboard rows ──────────────────────────────────────
MAIN_MENU = [
    ["📈 آلارم جدید"],
    ["⭐ آلارم‌های من", "📊 وضعیت"],
    ["⚡ آلارم فوری",   "⏰ هشدار دوره‌ای من"],
    ["📡 سیگنال جدید"],
]
MAIN_MENU_PRIVATE = [
    ["📈 آلارم جدید",  "🔒 آلارم شخصی"],
    ["⭐ آلارم‌های من", "📊 وضعیت"],
    ["⚡ آلارم فوری",   "⏰ هشدار دوره‌ای من"],
    ["📡 سیگنال جدید"],
]
MAIN_MENU_ADMIN = [
    ["📈 آلارم جدید",  "🔒 آلارم شخصی"],
    ["⭐ آلارم‌های من", "📊 وضعیت"],
    ["⚡ آلارم فوری",   "⏰ هشدار دوره‌ای من"],
    ["📡 سیگنال جدید", "⚙️ پنل ادمین"],
]
# منوی محدود برای اعضای جدید تا وقتی که ادمین تاییدشون نکرده
MAIN_MENU_NEW = [
    ["📊 وضعیت"],
    ["📩 درخواست فعال‌سازی آلارم"],
]
DIR_MENU = [["📈 BUY", "📉 SELL"], ["❌ انصراف"]]

def show_main_menu(token, cid, text, is_admin=False):
    if is_admin:
        rows = MAIN_MENU_ADMIN
    elif not _is_approved(cid):
        rows = MAIN_MENU_NEW
    elif _has_private_access(cid):
        rows = MAIN_MENU_PRIVATE
    else:
        rows = MAIN_MENU
    send_reply_keyboard(token, cid, text, rows)

def _get_user_custom_name(cid):
    data = load_alerts()
    for u in data.get("users", []):
        if str(u.get("chat_id","")) == str(cid):
            return u.get("custom_name","") or u.get("username","")
    return ""

def _has_private_access(cid):
    """آیا این کاربر دسترسی آلارم شخصی داره؟ فقط ادمین و کسایی که تایید شدن"""
    if str(cid) == str(YOUR_CHAT_ID):
        return True
    data = load_alerts()
    for u in data.get("users", []):
        if str(u.get("chat_id","")) == str(cid):
            return bool(u.get("private_access", False))
    return False

def _is_approved(cid):
    """آیا این کاربر اجازه‌ی استفاده کامل (ثبت آلارم و بقیه امکانات) داره؟
    کاربرهای قدیمی که فیلد approved ندارن به صورت پیش‌فرض تاییدشده حساب می‌شن،
    فقط اعضای جدید (که این فیلد رو صریحاً False دارن) باید تایید ادمین رو بگیرن."""
    if str(cid) == str(YOUR_CHAT_ID):
        return True
    data = load_alerts()
    for u in data.get("users", []):
        if str(u.get("chat_id","")) == str(cid):
            return bool(u.get("approved", True))
    return False

def _is_name_approved(name: str, pin: str = "") -> bool:
    """نسخه‌ی مبتنی بر اسم _is_approved — برای گیت کردن ثبت آلارم از سایت (وب)،
    جایی که chat_id نداریم و فقط اسمی که کاربر توی سایت انتخاب کرده در دسترسه.
    اسم با custom_name کاربرهای تلگرام مطابقت داده می‌شه.
    برای کاربرهایی که از این به بعد از طریق «درخواست فعال‌سازی» تایید می‌شن، یک کد (web_pin) هم
    باید مطابقت داشته باشه — تا صرفاً تایپ کردن اسمِ یک نفر دیگه کافی برای دور زدن تایید نباشه.
    کاربرهای قدیمی که web_pin ندارن (چون قبل از این قابلیت ثبت‌نام کردن) طبق روال قبل فقط با تطابق اسم عبور می‌کنن."""
    name = (name or "").strip().lower()
    if not name:
        return False
    data = load_alerts()
    for u in data.get("users", []):
        un = (u.get("custom_name") or "").strip().lower()
        if un and un == name:
            required_pin = u.get("web_pin")
            if required_pin:
                return str(pin or "").strip() == str(required_pin)
            return bool(u.get("approved", True))
    # اسمی که به هیچ کاربر تاییدشده‌ی تلگرام تطابق نداشته باشه (یعنی هنوز از تلگرام تاییدیه نگرفته) — اجازه نداره
    return False

def _verified_user_by_name(name: str, pin: str = ""):
    """مثل _is_name_approved ولی به‌جای True/False، خودِ رکورد کاربر رو برمی‌گردونه (یا None اگه
    هویتش تایید نشه). برای جاهایی که بعد از تایید هویت نیاز به فیلدهای دیگه‌ی کاربر
    (private_access, full_access, chat_id) داریم."""
    name_lc = (name or "").strip().lower()
    if not name_lc:
        return None
    data = load_alerts()
    for u in data.get("users", []):
        un = (u.get("custom_name") or "").strip().lower()
        if un and un == name_lc:
            required_pin = u.get("web_pin")
            if required_pin:
                if str(pin or "").strip() != str(required_pin):
                    return None
                return u
            if bool(u.get("approved", True)):
                return u
            return None
    return None


@app.route("/api/me", methods=["GET"])
def get_me():
    """وضعیت یه کاربر (بر اساس اسم+کد) — تا سایت بفهمه کدوم دکمه‌ها (مثلاً آلارم شخصی) رو نشونش بده"""
    name = request.args.get("name", "").strip()
    pin = request.args.get("pin", "").strip()
    u = _verified_user_by_name(name, pin)
    if not u:
        return jsonify({"ok": True, "verified": False, "private_access": False, "full_access": False})
    return jsonify({
        "ok": True,
        "verified": True,
        "private_access": bool(u.get("private_access")),
        "full_access": bool(u.get("full_access")),
    })


def _get_sender_name(msg):
    """اسم فرستنده — اول custom_name، بعد اسم تلگرام"""
    u = msg.get("from", {})
    cid = str(msg.get("chat", {}).get("id", "") or u.get("id", ""))
    if cid:
        users = load_alerts().get("users", [])
        for usr in users:
            if str(usr.get("chat_id", "")) == cid and usr.get("custom_name"):
                return usr["custom_name"]
    fn = u.get("first_name", "")
    ln = u.get("last_name", "")
    un = u.get("username", "")
    return (fn + " " + ln).strip() or ("@" + un if un else "ناشناس")


def _build_myalerts_section(alerts_list, title, start_idx=1):
    """ساخت متن و keyboard برای نمایش لیست آلارم"""
    if not alerts_list:
        return None, []
    lines = [f"{title}"]
    kb = []
    for i, a in enumerate(alerts_list, start_idx):
        sym2  = a.get("symbol","")
        tgt2  = a.get("target_price", 0)
        cond2 = "📈 BUY" if a.get("condition") == "below" else "📉 SELL"
        cur2  = a.get("last_price")
        cur_txt = f"<code>{fmt_price(cur2, sym2)}</code>" if cur2 else "—"
        cmt2  = f"\n💬 {a['comment']}" if a.get("comment") else ""
        block = (
            f"┌─ {i}. <b>{sym2}</b>  {cond2}\n"
            f"│  🎯 هدف: <code>{fmt_price(tgt2, sym2)}</code>\n"
            f"│  💹 فعلی: {cur_txt}"
            f"{cmt2}\n"
            f"└──────────────"
        )
        lines.append(block)
        kb.append([{"text": f"🗑 {i}. {sym2} {cond2} @ {fmt_price(tgt2, sym2)}", "callback_data": f"del_confirm:{a['id']}"}])
    kb.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
    return "\n".join(lines), kb

def poll_telegram():
    last_id = 0
    while True:
        try:
            token, _, _ = _get_token_and_cids()
            if not token:
                time.sleep(30)
                continue
            r = requests.get(
                f"https://api.telegram.org/bot{token}/getUpdates",
                params={"offset": last_id+1, "timeout": 20, "limit": 100},
                timeout=30, headers=H)
            if r.status_code != 200:
                time.sleep(10)
                continue
            _all_updates = r.json().get("result", [])
            for upd in _all_updates:
                last_id = upd["update_id"]
            # هر update در thread جدا — poll بلاک نمیشه
            for upd in _all_updates:
                threading.Thread(target=_do_update, args=(upd, token), daemon=True).start()
            continue

        except Exception as e:
            print(f"[poll] {e}")
        time.sleep(5)

def _do_update(upd, token):
  try:
                # ── callback query (دکمه‌های inline) ─────────────────
                cbq = upd.get("callback_query", {})
                if cbq:
                    cbq_id = cbq.get("id","")
                    cbq_data = cbq.get("data","")
                    cbq_cid = str(cbq.get("from",{}).get("id","") or cbq.get("message",{}).get("chat",{}).get("id",""))
                    cbq_msg_id = cbq.get("message",{}).get("message_id")
                    token_cbq, _, _ = _get_token_and_cids()

                    if cbq_data.startswith("del_confirm:"):
                        # مرحله اول: نمایش تأیید حذف
                        aid = cbq_data.split(":",1)[1]
                        d_conf = load_alerts()
                        a_conf = next((a for a in d_conf["alerts"] if a["id"] == aid), None)
                        answer_callback(token_cbq, cbq_id)
                        if a_conf:
                            sym_c = a_conf.get("symbol","")
                            tgt_c = a_conf.get("target_price", 0)
                            cond_c = "📈 BUY" if a_conf.get("condition") == "below" else "📉 SELL"
                            confirm_text = (
                                f"⚠️ <b>آیا مطمئنی؟</b>\n\n"
                                f"میخوای این آلارم رو حذف کنی:\n"
                                f"<b>{sym_c}</b>  {cond_c}  @ <code>{fmt_price(tgt_c, sym_c)}</code>"
                            )
                            confirm_kb = [
                                [{"text": "✅ بله، حذف کن", "callback_data": f"del_alert:{aid}"}],
                                [{"text": "❌ نه، برگرد", "callback_data": f"del_cancel:{aid}"}],
                            ]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, confirm_text, confirm_kb)
                        else:
                            answer_callback(token_cbq, cbq_id, "⚠️ آلارم پیدا نشد")

                    elif cbq_data.startswith("del_cancel:"):
                        # برگشت به لیست آلارم‌ها بعد از انصراف از حذف
                        answer_callback(token_cbq, cbq_id, "❌ حذف لغو شد")
                        d_can = load_alerts()
                        all_a_can = d_can.get("alerts", [])
                        custom_name_can = _get_user_custom_name(cbq_cid)
                        pub_can  = [a for a in all_a_can if a.get("active") and not a.get("is_private") and a.get("created_by","") == custom_name_can]
                        priv_can = [a for a in all_a_can if a.get("active") and a.get("is_private") and (
                            str(a.get("private_cid","")) == cbq_cid or str(a.get("notify_only","")) == cbq_cid
                        )]
                        combined_can = pub_can + priv_can
                        if pub_can and not priv_can:
                            txt_can, kb_can = _build_myalerts_section(pub_can, f"🌐 <b>آلارم‌های تیمی</b>  ({len(pub_can)} مورد)")
                        elif priv_can and not pub_can:
                            txt_can, kb_can = _build_myalerts_section(priv_can, f"🔒 <b>آلارم‌های شخصی</b>  ({len(priv_can)} مورد)")
                        elif combined_can:
                            txt_can, kb_can = _build_myalerts_section(pub_can, f"📋 <b>همه آلارم‌های من</b>  ({len(combined_can)} مورد)", start_idx=1)
                            priv_txt_can, priv_kb_can = _build_myalerts_section(priv_can, "\n🔒 <b>شخصی</b>", start_idx=len(pub_can)+1)
                            if priv_txt_can:
                                txt_can = txt_can + "\n" + priv_txt_can
                                kb_can = kb_can[:-1] + priv_kb_can
                        else:
                            txt_can, kb_can = "📭 هیچ آلارم فعالی نداری.", [[{"text":"✕ بستن","callback_data":"close_myalerts"}]]
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_can, kb_can)

                    elif cbq_data.startswith("del_alert:"):
                        # مرحله دوم: حذف واقعی بعد از تأیید
                        aid = cbq_data.split(":",1)[1]
                        d = load_alerts()
                        a_del = next((a for a in d["alerts"] if a["id"] == aid), None)
                        before = len(d["alerts"])
                        d["alerts"] = [a for a in d["alerts"] if a["id"] != aid]
                        if len(d["alerts"]) < before:
                            _cache_alerts = d
                            answer_callback(token_cbq, cbq_id, "✅ آلارم حذف شد")
                            threading.Thread(target=_sb_delete_alert, args=(aid,), daemon=True).start()
                        else:
                            answer_callback(token_cbq, cbq_id, "⚠️ آلارم پیدا نشد")
                        # بازسازی لیست در همان پیام — بدون sleep، بدون پیام جدید
                        d2 = load_alerts()
                        all_a2 = d2.get("alerts", [])
                        custom_name2 = _get_user_custom_name(cbq_cid)
                        pub2  = [a for a in all_a2 if a.get("active") and not a.get("is_private") and a.get("created_by","") == custom_name2]
                        priv2 = [a for a in all_a2 if a.get("active") and a.get("is_private") and (
                            str(a.get("private_cid","")) == cbq_cid or str(a.get("notify_only","")) == cbq_cid
                        )]
                        combined2 = pub2 + priv2
                        if not combined2:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ آلارم فعالی نداری.", [])
                        elif pub2 and not priv2:
                            txt2, kb2 = _build_myalerts_section(pub2, f"🌐 <b>آلارم‌های تیمی</b>  ({len(pub2)} مورد)")
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt2, kb2)
                        elif priv2 and not pub2:
                            txt2, kb2 = _build_myalerts_section(priv2, f"🔒 <b>آلارم‌های شخصی</b>  ({len(priv2)} مورد)")
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt2, kb2)
                        else:
                            txt2, kb2 = _build_myalerts_section(pub2, f"📋 <b>همه آلارم‌های من</b>  ({len(combined2)} مورد)", start_idx=1)
                            # اضافه کردن آلارم‌های شخصی به لیست ترکیبی
                            txt2_lines = txt2.split("\n")
                            priv_txt, priv_kb = _build_myalerts_section(priv2, f"\n🔒 <b>شخصی</b>", start_idx=len(pub2)+1)
                            if priv_txt:
                                txt2 = txt2 + "\n" + priv_txt
                                kb2 = kb2[:-1] + priv_kb  # بستن رو از pub حذف کن، priv_kb خودش داره
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt2, kb2)

                    elif cbq_data.startswith("reminder_new:"):
                        # reminder_new:CID — مستقیم از کاربر نماد بگیر
                        answer_callback(token_cbq, cbq_id)
                        _pending_reminder[cbq_cid] = {"step": "rem_symbol", "bot_msg_id": cbq_msg_id}
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            "➕ <b>هشدار جدید</b>\n\nنماد رو بنویس (مثلاً BTCUSDT یا XAUUSD):",
                            [[{"text": "❌ انصراف", "callback_data": "close_myalerts"}]])

                    elif cbq_data.startswith("reminder_sym:"):
                        # reminder_sym:CID:SYM — انتخاب تایم‌فریم
                        parts = cbq_data.split(":", 2)
                        r_cid = parts[1] if len(parts) > 1 else cbq_cid
                        r_sym = parts[2] if len(parts) > 2 else ""
                        answer_callback(token_cbq, cbq_id)
                        if r_sym == "__type__":
                            # کاربر باید خودش تایپ کنه
                            _pending_reminder[cbq_cid] = {"step": "rem_symbol", "bot_msg_id": cbq_msg_id}
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                "✏️ نماد رو بنویس (مثلاً EURUSD):",
                                [[{"text": "❌ انصراف", "callback_data": "close_myalerts"}]])
                        else:
                            kb_tf = [
                                [{"text": "🕯 M5  (۱ دق قبل کلوز)",  "callback_data": f"reminder_go:{r_cid}:{r_sym}:300"}],
                                [{"text": "🕯 M15 (۵ دق قبل کلوز)",  "callback_data": f"reminder_go:{r_cid}:{r_sym}:900"}],
                                [{"text": "🕯 H1  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{r_cid}:{r_sym}:3600"}],
                                [{"text": "🕯 H4  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{r_cid}:{r_sym}:14400"}],
                                [{"text": "✕ برگشت", "callback_data": f"reminder_new:{r_cid}"}],
                            ]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"🕯 تایم‌فریم هشدار برای <b>{r_sym}</b>:", kb_tf)

                    elif cbq_data.startswith("ack_trigger:"):
                        # ack_trigger:alarm_id:expected_user_id — فقط مسئول تریگر می‌تونه بزنه
                        parts_ack = cbq_data.split(":", 2)
                        ack_aid = parts_ack[1] if len(parts_ack) > 1 else ""
                        ack_expected_id = parts_ack[2] if len(parts_ack) > 2 else ""
                        clicker_id = str(cbq.get("from", {}).get("id", ""))
                        if not ack_expected_id or clicker_id != ack_expected_id:
                            answer_callback(token_cbq, cbq_id, "⛔ این دکمه فقط برای مسئول تریگره")
                        else:
                            # جلوگیری از دابل-ack — اگه دابل‌تپ شده یا قبلاً پردازش شده، دومی رو نادیده بگیر
                            with _ack_lock:
                                _already_acked = ack_aid in _ack_done
                                if not _already_acked:
                                    _ack_done.add(ack_aid)
                            if _already_acked:
                                answer_callback(token_cbq, cbq_id, "✅ قبلاً ثبت شده")
                            else:
                                answer_callback(token_cbq, cbq_id, "✅ ثبت شد")
                                ack_name = TEAM_ID_TO_NAME.get(clicker_id, "")
                                ack_time_label = now_pretty()

                                def _do_ack(aid=ack_aid, name=ack_name, tlabel=ack_time_label):
                                    # لایه‌ی دوم ایمنی — حتی اگه سرور بین دو کلیک ری‌استارت شده باشه
                                    # (که _ack_done در حافظه رو ریست می‌کنه)، از خود دیتابیس هم چک کن
                                    if SUPABASE_KEY:
                                        try:
                                            r_chk_ack = requests.get(
                                                f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}&select=ack_at",
                                                headers=_sb_h(), timeout=8)
                                            if r_chk_ack.status_code == 200:
                                                chk_rows_ack = r_chk_ack.json()
                                                if chk_rows_ack and chk_rows_ack[0].get("ack_at"):
                                                    print(f"[ack] {aid} از قبل تو DB ack شده بود — skip")
                                                    return
                                        except Exception as e:
                                            print(f"[ack] pre-check exc: {e}")
                                        try:
                                            requests.patch(
                                                f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}",
                                                headers={**_sb_h(), "Prefer": "return=minimal"},
                                                json={"ack_at": now_teh(), "ack_by": name},
                                                timeout=8)
                                        except Exception as e:
                                            print(f"[ack] save exc: {e}")
                                    # ادیت پیام برای همه‌ی چت‌هایی که این آلارم توشون ارسال شده
                                    tok_ack, _, _ = _get_token_and_cids()
                                    msg_map_ack = _fired_msg_ids.get(aid, {})
                                    base_text = msg_map_ack.get("__text__", "")
                                    if not base_text:
                                        return
                                    ack_line = f"\n✅ <b>{name}</b> قبول کرد در ساعت {tlabel}"
                                    new_text = base_text + ack_line
                                    # هر آلارم یه نماد ثابت داره — از تگ متن استخراج می‌کنیم نه از هر پیام
                                    sym_ack = _extract_symbol_from_tag(msg_map_ack.get("__tag__", ""))
                                    for tc_ack, tm_ack in msg_map_ack.items():
                                        if tc_ack in ("__tag__", "__text__"):
                                            continue
                                        new_kb = {"inline_keyboard": [[
                                            {"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc_ack}:{sym_ack}"}
                                        ]]}
                                        try:
                                            requests.post(
                                                f"https://api.telegram.org/bot{tok_ack}/editMessageText",
                                                json={"chat_id": tc_ack, "message_id": tm_ack,
                                                      "text": new_text, "parse_mode": "HTML",
                                                      "reply_markup": new_kb},
                                                timeout=8, headers=H)
                                        except Exception as e:
                                            print(f"[ack] edit exc: {e}")
                                    # آپدیت متن ذخیره‌شده تا اگه دوباره چیزی edit بشه این تغییر از دست نره
                                    msg_map_ack["__text__"] = new_text
                                    _fired_msg_ids[aid] = msg_map_ack
                                    threading.Thread(target=_sb_save_fired_msgs, args=(aid, msg_map_ack), daemon=True).start()

                                threading.Thread(target=_do_ack, daemon=True).start()

                    elif cbq_data.startswith("set_reminder:"):
                        # set_reminder:cid:SYM — از دکمه کنار الارم
                        parts = cbq_data.split(":", 2)
                        r_cid = parts[1] if len(parts) > 1 else cbq_cid
                        r_sym = parts[2] if len(parts) > 2 else "؟"
                        answer_callback(token_cbq, cbq_id)
                        kb_tf = [
                            [{"text": "🕯 M5  (۱ دق قبل کلوز)",  "callback_data": f"reminder_go:{r_cid}:{r_sym}:300"}],
                            [{"text": "🕯 M15 (۵ دق قبل کلوز)",  "callback_data": f"reminder_go:{r_cid}:{r_sym}:900"}],
                            [{"text": "🕯 H1  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{r_cid}:{r_sym}:3600"}],
                            [{"text": "🕯 H4  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{r_cid}:{r_sym}:14400"}],
                            [{"text": "✕ نه ممنون", "callback_data": "close_myalerts"}],
                        ]
                        send_tg_keyboard(token_cbq, cbq_cid,
                            f"🕯 تایم‌فریم هشدار کلوز برای <b>{r_sym}</b>:", kb_tf)

                    elif cbq_data.startswith("reminder_go:"):
                        parts = cbq_data.split(":")
                        r_cid = parts[1] if len(parts) > 1 else cbq_cid
                        r_sym = parts[2] if len(parts) > 2 else "؟"
                        r_tf  = int(parts[3]) if len(parts) > 3 else 3600
                        tf_label = _TF_LABEL.get(r_tf, f"{r_tf//60}m")
                        offset_min = _TF_OFFSET.get(r_tf, 300) // 60
                        wait_sec, close_teh, progress = _candle_info(r_tf)
                        wait_min = max(1, wait_sec // 60)
                        # کنسل reminder قبلی همین نماد
                        if _reminders.get(r_cid, {}).get(r_sym):
                            _reminders[r_cid][r_sym]["active"] = False
                            del _reminders[r_cid][r_sym]
                            threading.Thread(target=_sb_delete_reminder, args=(r_cid, r_sym), daemon=True).start()
                        _schedule_reminder(token_cbq, r_cid, r_sym, r_tf, tf_sec=r_tf)
                        def _bg_confirm(tok=token_cbq, cid_=cbq_cid, cbid=cbq_id,
                                        sym_=r_sym, tfl=tf_label, wm=wait_min, ct=close_teh,
                                        pr=progress, om=offset_min):
                            answer_callback(tok, cbid, f"✅ هشدار کلوز {tfl} فعال شد")
                            confirm_txt = (
                                f"✅ هشدار کلوز <b>{tfl}</b> برای <code>{sym_}</code> فعال شد.\n"
                                f"━━━━━━━━━━━━━━━━━━\n"
                                f"⏰ کلوز بعدی تهران: <b>{ct}</b>\n"
                                f"📊 پیشرفت کندل: <b>{pr}%</b>\n"
                                f"🔔 هشدار اول: <b>{wm} دقیقه</b> دیگه ({om} دق قبل کلوز)\n"
                                f"━━━━━━━━━━━━━━━━━━\n"
                                f"برای کنسل: /cancel_reminder"
                            )
                            send_tg(tok, cid_, confirm_txt)
                        threading.Thread(target=_bg_confirm, daemon=True).start()

                    elif cbq_data.startswith("cancel_reminder_one:"):
                        parts = cbq_data.split(":", 2)
                        r_cid = parts[1] if len(parts) > 1 else cbq_cid
                        r_sym = parts[2] if len(parts) > 2 else ""
                        if r_sym and _reminders.get(r_cid, {}).get(r_sym):
                            _reminders[r_cid][r_sym]["active"] = False
                            del _reminders[r_cid][r_sym]
                            if not _reminders.get(r_cid):
                                _reminders.pop(r_cid, None)
                            threading.Thread(target=_sb_delete_reminder, args=(r_cid, r_sym), daemon=True).start()
                            answer_callback(token_cbq, cbq_id, f"✅ هشدار {r_sym} کنسل شد")
                        else:
                            answer_callback(token_cbq, cbq_id, "هشداری پیدا نشد")
                        # آپدیت لیست
                        new_text, new_kb = build_cancel_reminder_msg(r_cid)
                        if new_kb:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, new_text, new_kb)
                        else:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "✅ همه هشدارها کنسل شدن.", [])

                    elif cbq_data.startswith("cancel_reminder_all:"):
                        r_cid = cbq_data.split(":", 1)[1] if ":" in cbq_data else cbq_cid
                        if r_cid in _reminders:
                            for info in _reminders[r_cid].values():
                                info["active"] = False
                            del _reminders[r_cid]
                            threading.Thread(target=_sb_delete_all_reminders, args=(r_cid,), daemon=True).start()
                            answer_callback(token_cbq, cbq_id, "✅ همه هشدارها کنسل شد")
                        else:
                            answer_callback(token_cbq, cbq_id, "هشداری فعال نبود")
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "✅ همه هشدارهای دوره‌ای کنسل شدن.", [])

                    elif cbq_data.startswith("myalerts:"):
                        # myalerts:pub:CID / myalerts:priv:CID / myalerts:all:CID
                        parts_mya = cbq_data.split(":", 2)
                        mya_type = parts_mya[1] if len(parts_mya) > 1 else "all"
                        mya_cid  = parts_mya[2] if len(parts_mya) > 2 else cbq_cid
                        answer_callback(token_cbq, cbq_id)
                        d_mya = load_alerts()
                        all_a_mya = d_mya.get("alerts", [])
                        custom_name_mya = _get_user_custom_name(mya_cid)
                        pub_mya  = [a for a in all_a_mya if a.get("active") and not a.get("is_private") and a.get("created_by","") == custom_name_mya]
                        # برای آلارم شخصی: هم private_cid هم notify_only چک کن
                        priv_mya = [a for a in all_a_mya if a.get("active") and a.get("is_private") and (
                            str(a.get("private_cid","")) == mya_cid or str(a.get("notify_only","")) == mya_cid
                        )]
                        if mya_type == "pub":
                            txt_mya, kb_mya = _build_myalerts_section(pub_mya, f"🌐 <b>آلارم‌های تیمی</b>  ({len(pub_mya)} مورد)")
                            if not txt_mya:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ آلارم تیمی فعالی نداری.", [[{"text":"✕ بستن","callback_data":"close_myalerts"}]])
                            else:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_mya, kb_mya)
                        elif mya_type == "priv":
                            txt_mya, kb_mya = _build_myalerts_section(priv_mya, f"🔒 <b>آلارم‌های شخصی</b>  ({len(priv_mya)} مورد)")
                            if not txt_mya:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ آلارم شخصی فعالی نداری.", [[{"text":"✕ بستن","callback_data":"close_myalerts"}]])
                            else:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_mya, kb_mya)
                        else:  # all
                            combined_mya = pub_mya + priv_mya
                            if not combined_mya:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ آلارم فعالی نداری.", [[{"text":"✕ بستن","callback_data":"close_myalerts"}]])
                            elif pub_mya and not priv_mya:
                                txt_mya_all, kb_mya_all = _build_myalerts_section(pub_mya, f"🌐 <b>آلارم‌های تیمی</b>  ({len(pub_mya)} مورد)")
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_mya_all, kb_mya_all)
                            elif priv_mya and not pub_mya:
                                txt_mya_all, kb_mya_all = _build_myalerts_section(priv_mya, f"🔒 <b>آلارم‌های شخصی</b>  ({len(priv_mya)} مورد)")
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_mya_all, kb_mya_all)
                            else:
                                txt_mya_all, kb_mya_all = _build_myalerts_section(pub_mya, f"📋 <b>همه آلارم‌های من</b>  ({len(combined_mya)} مورد)\n\n🌐 <b>تیمی</b>", start_idx=1)
                                priv_txt_all, priv_kb_all = _build_myalerts_section(priv_mya, "\n🔒 <b>شخصی</b>", start_idx=len(pub_mya)+1)
                                if priv_txt_all:
                                    txt_mya_all = txt_mya_all + "\n" + priv_txt_all
                                    kb_mya_all = kb_mya_all[:-1] + priv_kb_all
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, txt_mya_all, kb_mya_all)

                    elif cbq_data == "admin:news" and cbq_cid == YOUR_CHAT_ID:
                        answer_callback(token_cbq, cbq_id, "در حال دریافت اخبار...")
                        def _bg_news(tok=token_cbq, c=cbq_cid):
                            result = fetch_ff_news()
                            msg_text = result[1] if isinstance(result, tuple) else str(result)
                            send_tg(tok, c, msg_text)
                        threading.Thread(target=_bg_news, daemon=True).start()

                    elif cbq_data == "admin:broadcast" and cbq_cid == YOUR_CHAT_ID:
                        answer_callback(token_cbq, cbq_id)
                        _pending_alarm[cbq_cid] = {"step": "broadcast_text", "data": {}}
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=10, headers=H)
                        except: pass
                        send_tg(token_cbq, cbq_cid, "✉️ متن پیام رو بنویس:")

                    elif cbq_data.startswith("req_private:"):
                        # کاربر درخواست فعال‌سازی آلارم شخصی داده
                        req_cid = cbq_data.split(":")[1]
                        answer_callback(token_cbq, cbq_id, "✅ درخواست ارسال شد")
                        # پیام تایید به کاربر
                        send_tg(token_cbq, req_cid,
                            "📩 <b>درخواست شما ثبت شد</b>\n\n"
                            "درخواست فعال‌سازی آلارم شخصی شما ارسال شد و در دست بررسی است.\n"
                            "پس از تایید، یک پیام دریافت خواهید کرد. 🙏")
                        # پیام به ادمین
                        d_req = load_alerts()
                        req_user = next((u for u in d_req.get("users",[]) if str(u.get("chat_id","")) == req_cid), {})
                        req_name = req_user.get("custom_name","") or req_user.get("username","") or req_cid
                        admin_notif = (
                            f"📩 <b>درخواست آلارم شخصی</b>\n\n"
                            f"👤 نام: <b>{req_name}</b>\n"
                            f"🆔 Chat ID: <code>{req_cid}</code>\n"
                            f"⏰ {now_pretty()} (تهران)"
                        )
                        approve_kb = [
                            [{"text": "✅ تایید و فعال‌سازی", "callback_data": f"approve_private:{req_cid}"}],
                            [{"text": "❌ رد درخواست",       "callback_data": f"reject_private:{req_cid}"}],
                        ]
                        send_tg_keyboard(token_cbq, YOUR_CHAT_ID, admin_notif, approve_kb)

                    elif cbq_data.startswith("approve_private:"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            target_cid = cbq_data.split(":")[1]
                            answer_callback(token_cbq, cbq_id, "✅ فعال شد")
                            # آپدیت private_access توی users
                            d_apr = load_alerts()
                            found = False
                            for u in d_apr.get("users", []):
                                if str(u.get("chat_id","")) == target_cid:
                                    u["private_access"] = True
                                    found = True
                                    break
                            if not found:
                                d_apr.setdefault("users",[]).append({"chat_id": target_cid, "username": "", "joined_at": now_pretty(), "custom_name": "", "private_access": True})
                            save_alerts(d_apr)
                            # ویرایش پیام ادمین
                            try:
                                requests.post(f"https://api.telegram.org/bot{token_cbq}/editMessageReplyMarkup",
                                    json={"chat_id": cbq_cid, "message_id": cbq_msg_id, "reply_markup": {"inline_keyboard": []}},
                                    timeout=10, headers=H)
                            except: pass
                            send_tg(token_cbq, cbq_cid, f"✅ آلارم شخصی برای <code>{target_cid}</code> فعال شد.")
                            # پیام به کاربر
                            send_tg(token_cbq, target_cid,
                                "🎉 <b>آلارم شخصی شما فعال شد!</b>\n\n"
                                "یکبار /start بزنید تا منو به‌روز شود. 🔒")

                    elif cbq_data.startswith("reject_private:"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            target_cid = cbq_data.split(":")[1]
                            answer_callback(token_cbq, cbq_id, "❌ رد شد")
                            try:
                                requests.post(f"https://api.telegram.org/bot{token_cbq}/editMessageReplyMarkup",
                                    json={"chat_id": cbq_cid, "message_id": cbq_msg_id, "reply_markup": {"inline_keyboard": []}},
                                    timeout=10, headers=H)
                            except: pass
                            send_tg(token_cbq, cbq_cid, f"❌ درخواست <code>{target_cid}</code> رد شد.")
                            send_tg(token_cbq, target_cid,
                                "❌ <b>درخواست شما رد شد.</b>\n\n"
                                "برای اطلاعات بیشتر با ادمین تماس بگیرید.")

                    elif cbq_data.startswith("approve_signup:"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            target_cid_su = cbq_data.split(":")[1]
                            answer_callback(token_cbq, cbq_id, "✅ فعال شد")
                            web_pin_su = f"{secrets.randbelow(1000000):06d}"
                            d_apr_su = load_alerts()
                            found_su = False
                            for u in d_apr_su.get("users", []):
                                if str(u.get("chat_id","")) == target_cid_su:
                                    u["approved"] = True
                                    u["web_pin"] = web_pin_su
                                    found_su = True
                                    break
                            if not found_su:
                                d_apr_su.setdefault("users",[]).append({"chat_id": target_cid_su, "username": "", "joined_at": now_pretty(), "custom_name": "", "approved": True, "web_pin": web_pin_su})
                            save_alerts(d_apr_su)
                            try:
                                requests.post(f"https://api.telegram.org/bot{token_cbq}/editMessageReplyMarkup",
                                    json={"chat_id": cbq_cid, "message_id": cbq_msg_id, "reply_markup": {"inline_keyboard": []}},
                                    timeout=10, headers=H)
                            except: pass
                            send_tg(token_cbq, cbq_cid, f"✅ عضو <code>{target_cid_su}</code> تایید و فعال شد.")
                            send_tg(token_cbq, target_cid_su,
                                "🎉 <b>درخواست شما تایید شد!</b>\n\n"
                                "یکبار /start بزنید تا منو به‌روز شود. ✅\n\n"
                                f"🔑 کد فعال‌سازی سایت شما: <code>{web_pin_su}</code>\n"
                                "برای ثبت آلارم از طریق سایت، این کد رو یک‌بار توی سایت (دکمه «🔑 کد فعال‌سازی» زیر نام کاربری) وارد کن.")

                    elif cbq_data.startswith("reject_signup:"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            target_cid_su = cbq_data.split(":")[1]
                            answer_callback(token_cbq, cbq_id, "❌ رد شد")
                            try:
                                requests.post(f"https://api.telegram.org/bot{token_cbq}/editMessageReplyMarkup",
                                    json={"chat_id": cbq_cid, "message_id": cbq_msg_id, "reply_markup": {"inline_keyboard": []}},
                                    timeout=10, headers=H)
                            except: pass
                            send_tg(token_cbq, cbq_cid, f"❌ درخواست <code>{target_cid_su}</code> رد شد.")
                            send_tg(token_cbq, target_cid_su,
                                "❌ <b>درخواست شما رد شد.</b>\n\n"
                                "برای اطلاعات بیشتر با ادمین تماس بگیرید.")

                    elif cbq_data.startswith("admin:users"):
                        # لیست کاربران برای ادمین
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            answer_callback(token_cbq, cbq_id)
                            d_usr = load_alerts()
                            all_users = d_usr.get("users", [])
                            if not all_users:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ کاربری ثبت نشده.", [[{"text": "✕ بستن", "callback_data": "close_myalerts"}]])
                            else:
                                lines_u = ["👥 <b>لیست کاربران</b>\n"]
                                kb_u = []
                                for u in all_users:
                                    ucid = str(u.get("chat_id",""))
                                    uname_u = u.get("custom_name","") or u.get("username","") or ucid
                                    priv_icon = "🔒" if u.get("private_access") else "👤"
                                    lines_u.append(f"{priv_icon} <b>{uname_u}</b>  <code>{ucid}</code>")
                                    kb_u.append([{"text": f"🗑 حذف {uname_u}", "callback_data": f"admin:deluser:{ucid}"}])
                                kb_u.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "\n".join(lines_u), kb_u)

                    elif cbq_data.startswith("admin:deluser:"):
                        # مرحله اول: تایید حذف
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            del_cid = cbq_data.split(":")[2]
                            answer_callback(token_cbq, cbq_id)
                            d_confirm = load_alerts()
                            u_confirm = next((u for u in d_confirm.get("users",[]) if str(u.get("chat_id","")) == del_cid), {})
                            uname_confirm = u_confirm.get("custom_name","") or u_confirm.get("username","") or del_cid
                            confirm_kb = [
                                [{"text": "✅ بله، حذف کن", "callback_data": f"admin:confirmdelete:{del_cid}"}],
                                [{"text": "❌ نه، برگشت",   "callback_data": "admin:users"}],
                            ]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"⚠️ <b>تایید حذف کاربر</b>\n\n"
                                f"👤 <b>{uname_confirm}</b>\n"
                                f"🆔 <code>{del_cid}</code>\n\n"
                                f"مطمئنی می‌خوای این کاربر رو حذف کنی؟",
                                confirm_kb)

                    elif cbq_data.startswith("admin:confirmdelete:"):
                        # مرحله دوم: انجام حذف
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            del_cid = cbq_data.split(":")[2]
                            answer_callback(token_cbq, cbq_id, "🗑 حذف شد")
                            d_del = load_alerts()
                            d_del["users"] = [u for u in d_del.get("users",[]) if str(u.get("chat_id","")) != del_cid]
                            d_del["telegram"]["chat_ids"] = [x for x in d_del["telegram"].get("chat_ids",[]) if str(x) != del_cid]
                            save_alerts(d_del)
                            all_users2 = d_del.get("users", [])
                            if not all_users2:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 لیست کاربران خالی شد.", [[{"text": "✕ بستن", "callback_data": "close_myalerts"}]])
                            else:
                                lines_u2 = ["👥 <b>لیست کاربران</b>\n"]
                                kb_u2 = []
                                for u in all_users2:
                                    ucid = str(u.get("chat_id",""))
                                    uname_u2 = u.get("custom_name","") or u.get("username","") or ucid
                                    priv_icon2 = "🔒" if u.get("private_access") else "👤"
                                    lines_u2.append(f"{priv_icon2} <b>{uname_u2}</b>  <code>{ucid}</code>")
                                    kb_u2.append([{"text": f"🗑 حذف {uname_u2}", "callback_data": f"admin:deluser:{ucid}"}])
                                kb_u2.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "\n".join(lines_u2), kb_u2)

                    elif cbq_data == "admin:users":
                        # برگشت به لیست کاربران (از صفحه تایید)
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            answer_callback(token_cbq, cbq_id)
                            d_back = load_alerts()
                            all_users_back = d_back.get("users", [])
                            if not all_users_back:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "📭 هیچ کاربری ثبت نشده.", [[{"text": "✕ بستن", "callback_data": "close_myalerts"}]])
                            else:
                                lines_back = ["👥 <b>لیست کاربران</b>\n"]
                                kb_back = []
                                for u in all_users_back:
                                    ucid = str(u.get("chat_id",""))
                                    uname_back = u.get("custom_name","") or u.get("username","") or ucid
                                    priv_icon_back = "🔒" if u.get("private_access") else "👤"
                                    lines_back.append(f"{priv_icon_back} <b>{uname_back}</b>  <code>{ucid}</code>")
                                    kb_back.append([{"text": f"🗑 حذف {uname_back}", "callback_data": f"admin:deluser:{ucid}"}])
                                kb_back.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "\n".join(lines_back), kb_back)

                    elif cbq_data.startswith("admin_sig:"):
                        # فقط ادمین
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                            return
                        parts_as2 = cbq_data.split(":", 2)
                        as2_action = parts_as2[1]
                        as2_arg    = parts_as2[2] if len(parts_as2) > 2 else ""

                        if as2_action == "list":
                            answer_callback(token_cbq, cbq_id, "⏳ بارگذاری...")
                            page = int(as2_arg) if as2_arg.isdigit() else 1
                            per_page = 5
                            all_sigs = _sb_load_signals(limit=50)
                            total = len(all_sigs)
                            start_i = (page - 1) * per_page
                            page_sigs = all_sigs[start_i: start_i + per_page]
                            if not page_sigs:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    "🗑 <b>مدیریت سیگنال‌ها</b>\n\n📭 سیگنالی وجود ندارد.",
                                    [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                                return
                            lines_asl = [f"🗑 <b>مدیریت سیگنال‌ها</b>  ({total} سیگنال)\n"]
                            kb_asl = []
                            for s in page_sigs:
                                sid2   = s.get("id","")
                                sym2   = s.get("symbol","")
                                dir2   = s.get("direction","")
                                entry2 = s.get("entry")
                                tf2    = s.get("tf","")
                                by2    = s.get("sent_by","")
                                at2    = (s.get("sent_at") or "")[:16]
                                ch2    = s.get("channel_msg_id")
                                origin2 = "📤" if ch2 else "💾"
                                dir_short2 = {"buy_limit":"BL↗","buy_stop":"BS↗",
                                              "sell_limit":"SL↘","sell_stop":"SS↘"}.get(dir2, dir2)
                                lines_asl.append(
                                    f"{origin2} <b>{sid2}</b>  #{sym2}  {dir_short2}\n"
                                    f"   ➡️ <code>{_fmt_signal_price(entry2, sym2)}</code>  "
                                    f"⏱{tf2}  👤{by2}  🕐{at2}"
                                )
                                kb_asl.append([{"text": f"🗑 حذف {sid2} — {sym2} {dir_short2}",
                                                "callback_data": f"admin_sig:confirm:{sid2}:{page}"}])
                            # pagination
                            nav_row = []
                            if page > 1:
                                nav_row.append({"text": "◀️ قبل", "callback_data": f"admin_sig:list:{page-1}"})
                            if start_i + per_page < total:
                                nav_row.append({"text": "بعد ▶️", "callback_data": f"admin_sig:list:{page+1}"})
                            if nav_row:
                                kb_asl.append(nav_row)
                            kb_asl.append([{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}])
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                "\n".join(lines_asl), kb_asl)

                        elif as2_action == "confirm":
                            # as2_arg = "S10001:page"
                            conf_parts = as2_arg.split(":", 1)
                            conf_sid   = conf_parts[0]
                            conf_page  = conf_parts[1] if len(conf_parts) > 1 else "1"
                            answer_callback(token_cbq, cbq_id)
                            # اطلاعات سیگنال برای نمایش در تأیید
                            all_sigs_c = _sb_load_signals(limit=50)
                            sig_c = next((s for s in all_sigs_c if s.get("id") == conf_sid), None)
                            if not sig_c:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    "⚠️ سیگنال پیدا نشد.",
                                    [[{"text": "↩️ برگشت به لیست", "callback_data": f"admin_sig:list:{conf_page}"}]])
                                return
                            sym_c   = sig_c.get("symbol","")
                            dir_c   = sig_c.get("direction","")
                            entry_c = sig_c.get("entry")
                            by_c    = sig_c.get("sent_by","")
                            at_c    = (sig_c.get("sent_at") or "")[:16]
                            ch_c    = sig_c.get("channel_msg_id")
                            ch_warn = "\n\n⚠️ این سیگنال <b>به گروه ارسال شده</b> — حذف از DB فقط رکورد رو پاک میکنه، پیام کانال دست‌نخورده میمونه." if ch_c else ""
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"⚠️ <b>تأیید حذف سیگنال</b>\n\n"
                                f"🆔 <b>{conf_sid}</b>  #{sym_c}  {dir_c}\n"
                                f"➡️ Entry: <code>{_fmt_signal_price(entry_c, sym_c)}</code>\n"
                                f"👤 {by_c}  🕐 {at_c}"
                                f"{ch_warn}\n\n"
                                f"مطمئنی؟",
                                [[{"text": "✅ بله، حذف کن",    "callback_data": f"admin_sig:delete:{conf_sid}:{conf_page}"}],
                                 [{"text": "❌ نه، برگشت",       "callback_data": f"admin_sig:list:{conf_page}"}]])

                        elif as2_action == "delete":
                            del_parts  = as2_arg.split(":", 1)
                            del_sid    = del_parts[0]
                            del_page   = del_parts[1] if len(del_parts) > 1 else "1"
                            answer_callback(token_cbq, cbq_id, "🗑 در حال حذف...")
                            # حذف از Supabase
                            try:
                                r_del = requests.delete(
                                    f"{SUPABASE_URL}/rest/v1/signals?id=eq.{del_sid}",
                                    headers={**_sb_h(), "Prefer": "return=minimal"},
                                    timeout=10)
                                ok_del = r_del.status_code in (200, 204)
                            except Exception as e:
                                print(f"[admin_sig] delete error: {e}")
                                ok_del = False
                            if not ok_del:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    f"❌ خطا در حذف سیگنال <b>{del_sid}</b>.",
                                    [[{"text": "↩️ برگشت", "callback_data": f"admin_sig:list:{del_page}"}]])
                                return
                            # بعد از حذف، لیست آپدیت‌شده رو نشون بده
                            page_back = int(del_page) if del_page.isdigit() else 1
                            per_page2 = 5
                            all_sigs2 = _sb_load_signals(limit=50)
                            total2    = len(all_sigs2)
                            # اگه صفحه الان خالی شد، یه صفحه برگرد
                            start_i2 = (page_back - 1) * per_page2
                            if start_i2 >= total2 and page_back > 1:
                                page_back -= 1
                                start_i2 = (page_back - 1) * per_page2
                            page_sigs2 = all_sigs2[start_i2: start_i2 + per_page2]
                            if not page_sigs2:
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    f"✅ سیگنال <b>{del_sid}</b> حذف شد.\n\n📭 سیگنال دیگری وجود ندارد.",
                                    [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                                return
                            lines_af = [f"✅ <b>{del_sid}</b> حذف شد.\n\n🗑 <b>مدیریت سیگنال‌ها</b>  ({total2} سیگنال)\n"]
                            kb_af = []
                            for s in page_sigs2:
                                sid_f   = s.get("id",""); sym_f2  = s.get("symbol","")
                                dir_f   = s.get("direction",""); entry_f = s.get("entry")
                                tf_f    = s.get("tf",""); by_f    = s.get("sent_by","")
                                at_f    = (s.get("sent_at") or "")[:16]
                                ch_f    = s.get("channel_msg_id")
                                orig_f  = "📤" if ch_f else "💾"
                                ds_f    = {"buy_limit":"BL↗","buy_stop":"BS↗",
                                           "sell_limit":"SL↘","sell_stop":"SS↘"}.get(dir_f, dir_f)
                                lines_af.append(
                                    f"{orig_f} <b>{sid_f}</b>  #{sym_f2}  {ds_f}\n"
                                    f"   ➡️ <code>{_fmt_signal_price(entry_f, sym_f2)}</code>  "
                                    f"⏱{tf_f}  👤{by_f}  🕐{at_f}"
                                )
                                kb_af.append([{"text": f"🗑 حذف {sid_f} — {sym_f2} {ds_f}",
                                               "callback_data": f"admin_sig:confirm:{sid_f}:{page_back}"}])
                            nav_f = []
                            if page_back > 1:
                                nav_f.append({"text": "◀️ قبل", "callback_data": f"admin_sig:list:{page_back-1}"})
                            if start_i2 + per_page2 < total2:
                                nav_f.append({"text": "بعد ▶️", "callback_data": f"admin_sig:list:{page_back+1}"})
                            if nav_f: kb_af.append(nav_f)
                            kb_af.append([{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}])
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "\n".join(lines_af), kb_af)

                        elif as2_action == "back":
                            answer_callback(token_cbq, cbq_id)
                            admin_kb_b = [
                                [{"text": "📰 اخبار فارکس",      "callback_data": "admin:news"}],
                                [{"text": "✉️ پیام به گروه",     "callback_data": "admin:broadcast"}],
                                [{"text": "👥 لیست کاربران",      "callback_data": "admin:users"}],
                                [{"text": "🗑 مدیریت سیگنال‌ها", "callback_data": "admin_sig:list:1"}],
                                [{"text": "📋 مسئولین آلارم",        "callback_data": "admin:shift:1"}],
                                [{"text": "❌ فالس همه آلارم‌های فعال", "callback_data": "admin:bulkfalse:0"}],
                                [{"text": "✕ بستن",               "callback_data": "close_myalerts"}],
                            ]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                "⚙️ <b>پنل ادمین</b>\n\nیه گزینه انتخاب کن:", admin_kb_b)

                    # ─── مدیریت مسئولین آلارم (جابجایی دستی) ──────────────────
                    elif cbq_data.startswith("admin:shift"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            parts_sh = cbq_data.split(":")
                            sh_action = parts_sh[2] if len(parts_sh) > 2 else "1"

                            # ── لیست آلارم‌های فعال (صفحه‌بندی ۵تایی) ──
                            if sh_action.isdigit():
                                answer_callback(token_cbq, cbq_id, "⏳ بارگذاری...")
                                sh_page = int(sh_action)
                                sh_rows = _sb_load_active_assignments()
                                PER_PAGE_SH = 5
                                total_sh = len(sh_rows)
                                if total_sh == 0:
                                    edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                        "📭 <b>هیچ آلارم فعالی نیست.</b>",
                                        [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                                else:
                                    start_sh = (sh_page - 1) * PER_PAGE_SH
                                    page_rows = sh_rows[start_sh: start_sh + PER_PAGE_SH]
                                    lines_sh = [f"📋 <b>آلارم‌های فعال</b>  ({total_sh} عدد)\n"]
                                    kb_sh = []
                                    for row_sh in page_rows:
                                        tag_sh   = row_sh.get("alarm_tag", "")
                                        asn_sh   = row_sh.get("assigned_to") or "—"
                                        sym_sh   = row_sh.get("symbol", "")
                                        lines_sh.append(
                                            f"• {tag_sh}  <code>{sym_sh}</code>\n"
                                            f"  👤 {asn_sh}"
                                        )
                                        aid_sh = row_sh.get("id","")
                                        kb_sh.append([{"text": f"🔀 {tag_sh} ({asn_sh})",
                                                        "callback_data": f"admin:shift:assign:{aid_sh}:{sh_page}"}])
                                    # ناوبری صفحه
                                    nav_sh = []
                                    if sh_page > 1:
                                        nav_sh.append({"text": "◀️ قبل", "callback_data": f"admin:shift:{sh_page-1}"})
                                    if start_sh + PER_PAGE_SH < total_sh:
                                        nav_sh.append({"text": "بعد ▶️", "callback_data": f"admin:shift:{sh_page+1}"})
                                    if nav_sh:
                                        kb_sh.append(nav_sh)
                                    kb_sh.append([{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}])
                                    edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                        "\n".join(lines_sh), kb_sh)

                            # ── انتخاب آلارم برای reassign — نشون دادن لیست اعضا ──
                            elif sh_action == "assign" and len(parts_sh) >= 5:
                                answer_callback(token_cbq, cbq_id)
                                aid_asgn  = parts_sh[3]
                                page_asgn = parts_sh[4]
                                kb_asgn = []
                                for m in TEAM_MEMBERS:
                                    kb_asgn.append([{"text": f"👤 {m}",
                                                      "callback_data": f"admin:shift:do:{aid_asgn}:{m}:{page_asgn}"}])
                                kb_asgn.append([{"text": "↩️ برگشت به لیست",
                                                  "callback_data": f"admin:shift:{page_asgn}"}])
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    f"👤 <b>انتخاب مسئول جدید</b>\n\n<code>{aid_asgn}</code>\n\nکدوم نفر؟",
                                    kb_asgn)

                            # ── تایید reassign — ذخیره در Supabase ──
                            elif sh_action == "do" and len(parts_sh) >= 6:
                                answer_callback(token_cbq, cbq_id, "✅ در حال جابجایی...")
                                aid_do    = parts_sh[3]
                                new_asn   = parts_sh[4]
                                page_do   = parts_sh[5]
                                def _do_reassign(aid=aid_do, asn=new_asn, tok=token_cbq,
                                                  c=cbq_cid, mid=cbq_msg_id, pg=page_do):
                                    # خوندن اطلاعات آلارم
                                    try:
                                        r_get = requests.get(
                                            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}&select=*",
                                            headers=_sb_h(), timeout=8)
                                        row_d = r_get.json()[0] if r_get.status_code == 200 and r_get.json() else {}
                                    except:
                                        row_d = {}
                                    old_asn = row_d.get("assigned_to","")
                                    tag_d   = row_d.get("alarm_tag","")
                                    already_acked = bool(row_d.get("ack_at"))
                                    # آپدیت شمارش حافظه
                                    if old_asn and old_asn in _active_assign_count:
                                        _active_assign_count[old_asn] = max(0, _active_assign_count[old_asn] - 1)
                                    _active_assign_count[asn] = _active_assign_count.get(asn, 0) + 1
                                    # ذخیره در Supabase
                                    requests.patch(
                                        f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}",
                                        headers={**_sb_h(), "Prefer": "return=minimal"},
                                        json={"assigned_to": asn},
                                        timeout=8)
                                    # ارسال reply به گروه
                                    tg_tok, cids_d, _ = _get_token_and_cids()
                                    msg_map_d = _fired_msg_ids.get(aid, {})
                                    reply_d = (f"🔀 <b>جابجایی دستی</b>\n\n"
                                               f"{tag_d}\n"
                                               f"👤 مسئول جدید: <b>{asn}</b>"
                                               + (f"\n↩️ قبلی: {old_asn}" if old_asn else ""))
                                    for tc_d, tm_d in msg_map_d.items():
                                        if tc_d in ("__tag__","__text__"): continue
                                        try:
                                            requests.post(
                                                f"https://api.telegram.org/bot{tg_tok}/sendMessage",
                                                json={"chat_id": tc_d, "text": reply_d,
                                                      "parse_mode": "HTML",
                                                      "reply_to_message_id": tm_d},
                                                timeout=8, headers=H)
                                        except: pass
                                    # آپدیت دکمه «دیدم» روی پیام اصلی هر گیرنده —
                                    # فقط اگه هنوز کسی ack نکرده باشه، منتقلش کن به مسئول جدید
                                    if not already_acked:
                                        new_ack_id = TEAM_MEMBER_IDS.get(asn, "")
                                        sym_d = _extract_symbol_from_tag(tag_d)
                                        for tc_d, tm_d in msg_map_d.items():
                                            if tc_d in ("__tag__","__text__"): continue
                                            kb_d = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc_d}:{sym_d}"}]]
                                            if new_ack_id and str(tc_d) == str(new_ack_id):
                                                kb_d.append([{"text": "✅ دیدم", "callback_data": f"ack_trigger:{aid}:{new_ack_id}"}])
                                            try:
                                                requests.post(
                                                    f"https://api.telegram.org/bot{tg_tok}/editMessageReplyMarkup",
                                                    json={"chat_id": tc_d, "message_id": tm_d,
                                                          "reply_markup": {"inline_keyboard": kb_d}},
                                                    timeout=8, headers=H)
                                            except: pass
                                    edit_tg_keyboard(tok, c, mid,
                                        f"✅ <b>جابجایی انجام شد</b>\n\n{tag_d}\n👤 {asn}",
                                        [[{"text": "📋 برگشت به لیست", "callback_data": f"admin:shift:{pg}"}],
                                         [{"text": "↩️ پنل ادمین",    "callback_data": "admin_sig:back"}]])
                                threading.Thread(target=_do_reassign, daemon=True).start()

                    # ─── فالس دسته‌جمعی همه‌ی آلارم‌های فعال ──────────────────
                    elif cbq_data.startswith("admin:bulkfalse"):
                        if cbq_cid != YOUR_CHAT_ID:
                            answer_callback(token_cbq, cbq_id, "⛔ فقط ادمین")
                        else:
                            bf_parts = cbq_data.split(":")
                            bf_step = bf_parts[2] if len(bf_parts) > 2 else "0"

                            # مرحله ۰: نمایش تعداد و درخواست تایید
                            if bf_step == "0":
                                answer_callback(token_cbq, cbq_id, "⏳ شمارش...")
                                bf_rows = _sb_load_active_assignments()
                                bf_count = len(bf_rows)
                                if bf_count == 0:
                                    edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                        "📭 <b>هیچ آلارم فعالی نیست.</b>",
                                        [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                                else:
                                    lines_bf = [f"⚠️ <b>فالس دسته‌جمعی</b>\n\n{bf_count} آلارم فعال پیدا شد:\n"]
                                    for row_bf in bf_rows[:20]:
                                        lines_bf.append(f"• {row_bf.get('alarm_tag','')}  👤 {row_bf.get('assigned_to','—')}")
                                    if bf_count > 20:
                                        lines_bf.append(f"... و {bf_count-20} مورد دیگه")
                                    lines_bf.append(f"\nهمه‌ی این {bf_count} آلارم false می‌شن. مطمئنی؟")
                                    edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                        "\n".join(lines_bf),
                                        [[{"text": f"✅ بله، همه رو false کن ({bf_count})", "callback_data": "admin:bulkfalse:reason"}],
                                         [{"text": "↩️ انصراف", "callback_data": "admin_sig:back"}]])

                            # مرحله ۱: درخواست متن دلیل
                            elif bf_step == "reason":
                                answer_callback(token_cbq, cbq_id)
                                bf_rows2 = _sb_load_active_assignments()
                                _pending_bulk_false[cbq_cid] = {"bot_msg_id": cbq_msg_id, "count": len(bf_rows2)}
                                edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                    f"📝 <b>دلیل فالس کردن {len(bf_rows2)} آلارم رو بنویس</b>\n\n"
                                    f"(مثلاً: آخر هفته / تعطیلی / تست)\n\n"
                                    f"یا دکمه‌ی زیر رو بزن تا بدون دلیل ثبت بشه.",
                                    [[{"text": "⏭ بدون دلیل، فقط false کن", "callback_data": "admin:bulkfalse:go:__no_reason__"}],
                                     [{"text": "❌ انصراف", "callback_data": "admin_sig:back"}]])

                            # مرحله ۲: اجرای نهایی (یا از متن تایپ‌شده، یا از دکمه‌ی بدون دلیل)
                            elif bf_step == "go":
                                bf_reason = bf_parts[3] if len(bf_parts) > 3 else ""
                                if bf_reason == "__no_reason__":
                                    bf_reason = ""
                                answer_callback(token_cbq, cbq_id, "⏳ در حال فالس کردن همه...")
                                _pending_bulk_false.pop(cbq_cid, None)

                                def _do_bulk_false(reason=bf_reason, tok=token_cbq, c=cbq_cid, mid=cbq_msg_id):
                                    rows_bf = _sb_load_active_assignments()
                                    admin_name_bf = _get_user_custom_name(c) or "ادمین"
                                    done_count = 0
                                    for row_bf in rows_bf:
                                        aid_bf = row_bf.get("id")
                                        tag_bf = row_bf.get("alarm_tag", "")
                                        # همون تابع false تکی رو صدا بزن — یکسان با /False دستی
                                        _sb_false_assignment(aid_bf, admin_name_bf, reason)
                                        done_count += 1
                                        # ریپلای به همه‌ی گیرنده‌های این آلارم
                                        msg_map_bf = _fired_msg_ids.get(aid_bf, {})
                                        reason_line_bf = f"\n📝 علت: {reason}" if reason else ""
                                        reply_bf = (f"❌ <b>فالس (دسته‌جمعی توسط ادمین)</b>\n\n"
                                                    f"{tag_bf}{reason_line_bf}")
                                        for tc_bf, tm_bf in msg_map_bf.items():
                                            if tc_bf in ("__tag__", "__text__"): continue
                                            try:
                                                requests.post(
                                                    f"https://api.telegram.org/bot{tok}/sendMessage",
                                                    json={"chat_id": tc_bf, "text": reply_bf,
                                                          "parse_mode": "HTML", "reply_to_message_id": tm_bf},
                                                    timeout=8, headers=H)
                                            except: pass
                                    # بعد از فالس کردن همه، شمارش عادلانه رو یک‌بار نهایی بازسازی کن
                                    _rebuild_active_assign_count(_sb_load_active_assignments())
                                    edit_tg_keyboard(tok, c, mid,
                                        f"✅ <b>{done_count} آلارم false شدن</b>" + (f"\n📝 علت: {reason}" if reason else ""),
                                        [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                                threading.Thread(target=_do_bulk_false, daemon=True).start()

                    elif cbq_data.startswith("clean_chat:"):
                        answer_callback(token_cbq, cbq_id, "🧹 در حال پاک‌سازی...")
                        target_cid = cbq_data.split(":", 1)[1]
                        if cbq_msg_id:
                            _track_msg(cbq_cid, cbq_msg_id)
                        def _do_clean(tok=token_cbq, c=cbq_cid, tc=target_cid):
                            cnt = delete_chat_history(tok, tc)
                            send_tg_keyboard(tok, c,
                                f"✅ <b>{cnt} پیام پاک شد.</b>",
                                [[{"text": "✕ بستن", "callback_data": "close_myalerts"}]])
                        threading.Thread(target=_do_clean, daemon=True).start()

                    elif cbq_data.startswith("edit_name:"):
                        # شروع flow ویرایش اسم از طریق استاتوس
                        en_cid = cbq_data.split(":", 1)[1]
                        answer_callback(token_cbq, cbq_id)
                        d_en = load_alerts()
                        cur_name_en = next(
                            (u.get("custom_name","") for u in d_en.get("users",[]) if str(u.get("chat_id","")) == en_cid),
                            "")
                        cur_info_en = f"\nاسم فعلی: <b>{cur_name_en}</b>" if cur_name_en else ""
                        edit_tg_keyboard(token_cbq, en_cid, cbq_msg_id,
                            f"✏️ <b>ویرایش اسم</b>{cur_info_en}\n\nاسم جدیدت رو بنویس:",
                            [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{en_cid}"}]])
                        _pending_alarm[en_cid] = {"step": "edit_name_input", "data": {}, "bot_msg_id": cbq_msg_id}

                    elif cbq_data == "close_myalerts":
                        answer_callback(token_cbq, cbq_id, "بسته شد")
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=10, headers=H)
                        except: pass

                    elif cbq_data.startswith("flow_cancel:"):
                        # لغو هر flow در حال اجرا (آلارم عادی یا SOS)
                        f_cid = cbq_data.split(":", 1)[1]
                        _pending_alarm.pop(f_cid, None)
                        _pending_signal.pop(f_cid, None)
                        answer_callback(token_cbq, cbq_id, "لغو شد")
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            "❌ <b>عملیات لغو شد.</b>", [])

                    # ── signal callbacks ──────────────────────────────
                    elif cbq_data.startswith("sig_cancel:"):
                        sc_cid = cbq_data.split(":",1)[1]
                        _pending_signal.pop(sc_cid, None)
                        answer_callback(token_cbq, cbq_id, "لغو شد")
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            "❌ <b>ساخت سیگنال لغو شد.</b>", [])

                    elif cbq_data.startswith("sig_sym:"):
                        # انتخاب نماد از shortcut
                        parts_ss = cbq_data.split(":", 2)
                        ss_cid = parts_ss[1]; ss_sym = parts_ss[2]
                        ps = _pending_signal.get(ss_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        ps["data"]["symbol"] = ss_sym
                        ps["step"] = "sig_direction"
                        dir_kb = [[{"text": lbl, "callback_data": f"sig_dir:{ss_cid}:{val}"} for lbl,val in SIGNAL_DIRECTIONS[:2]],
                                  [{"text": lbl, "callback_data": f"sig_dir:{ss_cid}:{val}"} for lbl,val in SIGNAL_DIRECTIONS[2:]],
                                  [{"text": "❌ انصراف", "callback_data": f"sig_cancel:{ss_cid}"}]]
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"📡 <b>{ss_sym}</b>\n\nنوع سفارش:", dir_kb)

                    elif cbq_data.startswith("sig_dir:"):
                        # انتخاب جهت
                        parts_sd2 = cbq_data.split(":", 2)
                        sd2_cid = parts_sd2[1]; sd2_dir = parts_sd2[2]
                        ps = _pending_signal.get(sd2_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        dir_lbl_map = {"buy_limit":"✅ Buy Limit","buy_stop":"✅ Buy Stop",
                                       "sell_limit":"🔴 Sell Limit","sell_stop":"🔴 Sell Stop"}
                        ps["data"]["direction"] = sd2_dir
                        ps["data"]["dir_lbl"] = dir_lbl_map.get(sd2_dir, sd2_dir)
                        ps["step"] = "sig_sl_mode"
                        sym_sd2 = ps["data"].get("symbol","")
                        # انتخاب نوع SL
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"📡 <b>{sym_sd2}</b>  {ps['data']['dir_lbl']}\n\nاستاپ رو چطور میدی؟",
                            [[{"text": "🔢 عدد مستقیم", "callback_data": f"sig_slmode:{sd2_cid}:price"},
                              {"text": "📏 پیپ", "callback_data": f"sig_slmode:{sd2_cid}:pip"}],
                             [{"text": "❌ انصراف", "callback_data": f"sig_cancel:{sd2_cid}"}]])

                    elif cbq_data.startswith("sig_slmode:"):
                        parts_sm = cbq_data.split(":", 2)
                        sm_cid = parts_sm[1]; sm_mode = parts_sm[2]
                        ps = _pending_signal.get(sm_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        ps["data"]["sl_mode"] = sm_mode
                        ps["step"] = "sig_entry_sl"
                        sym_sm = ps["data"].get("symbol","")
                        dir_lbl_sm = ps["data"].get("dir_lbl","")
                        mode_hint = "قیمت SL" if sm_mode == "price" else "پیپ SL"
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"📡 <b>{sym_sm}</b>  {dir_lbl_sm}\n\n"
                            f"بنویس:  <code>Entry  {mode_hint}</code>\n"
                            f"مثال:   <code>73370  {'72550' if sm_mode == 'price' else '82'}</code>",
                            [[{"text": "❌ انصراف", "callback_data": f"sig_cancel:{sm_cid}"}]])

                    elif cbq_data.startswith("sig_tf:"):
                        # انتخاب تایم‌فریم
                        stf_cid = cbq_data.split(":",1)[1]
                        ps = _pending_signal.get(stf_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        cur_tf = ps["data"].get("tf", SIGNAL_DEFAULT_TF)
                        tf_kb = [[{"text": f"{'✅ ' if tf==cur_tf else ''}{tf}",
                                   "callback_data": f"sig_settf:{stf_cid}:{tf}"}
                                  for tf in SIGNAL_TF_OPTIONS[:3]],
                                 [{"text": f"{'✅ ' if tf==cur_tf else ''}{tf}",
                                   "callback_data": f"sig_settf:{stf_cid}:{tf}"}
                                  for tf in SIGNAL_TF_OPTIONS[3:]],
                                 [{"text": "↩️ بازگشت", "callback_data": f"sig_back:{stf_cid}"}]]
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"⏱ تایم‌فریم رو انتخاب کن (فعلی: <b>{cur_tf}</b>):", tf_kb)

                    elif cbq_data.startswith("sig_settf:"):
                        parts_stf = cbq_data.split(":", 2)
                        stf_cid2 = parts_stf[1]; stf_val = parts_stf[2]
                        ps = _pending_signal.get(stf_cid2)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id, f"✅ {stf_val}")
                        ps["data"]["tf"] = stf_val
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token_cbq, cbq_cid, cbq_msg_id, ps["data"])

                    elif cbq_data.startswith("sig_tp:"):
                        stp_cid = cbq_data.split(":",1)[1]
                        ps = _pending_signal.get(stp_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        ps["data"]["_editing_tp"] = "tp2"
                        ps["step"] = "sig_tp_edit"
                        cur_tp2 = ps["data"].get("tp2")
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"🎯 <b>TP2</b> رو بنویس (اختیاری):\n"
                            + (f"فعلی: <code>{_fmt_signal_price(cur_tp2, ps['data'].get('symbol',''))}</code>\n" if cur_tp2 else "")
                            + "برای حذف بنویس: <code>0</code>",
                            [[{"text": "⏭ رد کن", "callback_data": f"sig_skip_tp:{stp_cid}:tp2"},
                              {"text": "↩️ بازگشت", "callback_data": f"sig_back:{stp_cid}"}]])

                    elif cbq_data.startswith("sig_skip_tp:"):
                        parts_skp = cbq_data.split(":", 2)
                        skp_cid = parts_skp[1]; skp_which = parts_skp[2]
                        ps = _pending_signal.get(skp_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        # اگه tp2 رد شد، tp3 رو هم رد کن
                        ps["data"]["tp2"] = None
                        ps["data"]["tp3"] = None
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token_cbq, cbq_cid, cbq_msg_id, ps["data"])

                    elif cbq_data.startswith("sig_note:"):
                        sn_cid = cbq_data.split(":",1)[1]
                        ps = _pending_signal.get(sn_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        ps["step"] = "sig_note"
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            "📝 یادداشت بنویس (اختیاری):",
                            [[{"text": "⏭ رد کن", "callback_data": f"sig_back:{sn_cid}"},
                              {"text": "❌ انصراف", "callback_data": f"sig_cancel:{sn_cid}"}]])

                    elif cbq_data.startswith("sig_recalc:"):
                        # محاسبه مجدد با RR متفاوت
                        src_cid = cbq_data.split(":",1)[1]
                        ps = _pending_signal.get(src_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        rr_kb = [[{"text": f"{'✅ ' if ps['data'].get('rr')==v else ''}{v}R",
                                   "callback_data": f"sig_setrr:{src_cid}:{v}"}
                                  for v in [1.0, 1.5, 2.0]],
                                 [{"text": f"{'✅ ' if ps['data'].get('rr')==v else ''}{v}R",
                                   "callback_data": f"sig_setrr:{src_cid}:{v}"}
                                  for v in [2.5, 3.0]],
                                 [{"text": "↩️ بازگشت", "callback_data": f"sig_back:{src_cid}"}]]
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"🔄 ریوارد رو انتخاب کن (فعلی: <b>{ps['data'].get('rr', SIGNAL_DEFAULT_RR)}R</b>):", rr_kb)

                    elif cbq_data.startswith("sig_setrr:"):
                        parts_rr = cbq_data.split(":", 2)
                        rr_cid = parts_rr[1]; rr_val = float(parts_rr[2])
                        ps = _pending_signal.get(rr_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id, f"✅ ریوارد {rr_val}R")
                        d_rr = ps["data"]
                        # محاسبه مجدد
                        _, tp1_new, _ = _calc_signal(d_rr["symbol"], d_rr["direction"],
                                                      d_rr["entry"], d_rr["sl"], rr_val)
                        d_rr["rr"] = rr_val
                        d_rr["tp1"] = tp1_new
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token_cbq, cbq_cid, cbq_msg_id, d_rr)

                    elif cbq_data.startswith("sig_back:"):
                        sb_cid = cbq_data.split(":",1)[1]
                        ps = _pending_signal.get(sb_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id)
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token_cbq, cbq_cid, cbq_msg_id, ps["data"])

                    elif cbq_data.startswith("signals_view:"):
                        # نمایش لیست سیگنال‌ها
                        parts_sv = cbq_data.split(":", 2)
                        sv_cid  = parts_sv[1]
                        sv_mode = parts_sv[2]  # mine | all
                        answer_callback(token_cbq, cbq_id, "⏳ در حال بارگذاری...")
                        sigs = _sb_load_signals(limit=20)
                        if sv_mode == "mine":
                            my_name = _get_user_custom_name(sv_cid) or sv_cid
                            sigs = [s for s in sigs if s.get("sent_by","") == my_name]
                            title = f"📡 <b>سیگنال‌های من</b>"
                        else:
                            title = f"📊 <b>همه سیگنال‌ها</b>"
                        if not sigs:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"{title}\n\n📭 سیگنالی یافت نشد.",
                                [[{"text": "↩️ بازگشت", "callback_data": f"signals_close:{sv_cid}"}]])
                            return
                        lines = [title, ""]
                        for s in sigs:
                            sym_sv   = s.get("symbol","")
                            sid_sv   = s.get("id","")
                            dir_sv   = s.get("direction","")
                            entry_sv = s.get("entry")
                            sl_sv    = s.get("sl")
                            tp1_sv   = s.get("tp1")
                            tf_sv    = s.get("tf","")
                            sent_by_sv = s.get("sent_by","")
                            sent_at_sv = s.get("sent_at","")[:16] if s.get("sent_at") else ""
                            ch_mid_sv  = s.get("channel_msg_id")
                            # آیکون ارسال به کانال یا فقط DB
                            origin = "📤" if ch_mid_sv else "💾"
                            dir_short = {"buy_limit":"BL↗","buy_stop":"BS↗",
                                         "sell_limit":"SL↘","sell_stop":"SS↘"}.get(dir_sv, dir_sv)
                            lines.append(
                                f"{origin} <b>{sid_sv}</b>  #{sym_sv}  <i>{dir_short}</i>\n"
                                f"   ➡️ <code>{_fmt_signal_price(entry_sv,sym_sv)}</code>  "
                                f"🛑 <code>{_fmt_signal_price(sl_sv,sym_sv)}</code>  "
                                f"🎯 <code>{_fmt_signal_price(tp1_sv,sym_sv)}</code>\n"
                                f"   ⏱ {tf_sv}  👤 {sent_by_sv}  🕐 {sent_at_sv}"
                            )
                            lines.append("──────────────────")
                        legend = "\n📤 = ارسال به گروه   💾 = فقط ثبت دیتا"
                        kb_sv = [
                            [{"text": "📡 سیگنال‌های من", "callback_data": f"signals_view:{sv_cid}:mine"},
                             {"text": "📊 همه سیگنال‌ها", "callback_data": f"signals_view:{sv_cid}:all"}],
                            [{"text": "✕ بستن", "callback_data": f"signals_close:{sv_cid}"}]
                        ]
                        full_text = "\n".join(lines) + legend
                        # تلگرام max 4096 کاراکتر
                        if len(full_text) > 4000:
                            full_text = full_text[:3980] + "\n\n<i>... (برای دیدن بیشتر فیلتر کن)</i>"
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, full_text, kb_sv)

                    elif cbq_data.startswith("signals_close:"):
                        answer_callback(token_cbq, cbq_id, "بسته شد")
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=8, headers=H)
                        except: pass

                    elif cbq_data.startswith("trigger_list:"):
                        answer_callback(token_cbq, cbq_id, "⏳ در حال بارگذاری...")
                        tl_cid = cbq_data.split(":", 1)[1]
                        rows_tl_all = _sb_load_active_assignments()
                        # فقط آلارم‌های تیمی — فقط is_private رو برای همون IDهای لازم چک کن (نه کل جدول)
                        tl_ids = sorted({str(r.get("id","")) for r in rows_tl_all if r.get("id")})
                        private_ids = set()
                        if SUPABASE_KEY and tl_ids:
                            try:
                                r_tl = requests.get(
                                    f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(tl_ids)})&select=id,is_private",
                                    headers=_sb_h(), timeout=10)
                                if r_tl.status_code == 200:
                                    private_ids = {str(x.get("id","")) for x in r_tl.json() if x.get("is_private")}
                            except Exception as e:
                                print(f"[trigger_list] targeted fetch error: {e}")
                        rows_tl = [r for r in rows_tl_all if str(r.get("id","")) not in private_ids]
                        my_name_tl = _get_user_custom_name(tl_cid) or ""
                        if not rows_tl:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                "🎯 <b>لیست تریگر</b>\n\n📭 هیچ آلارم فعالی در لیست تریگر نیست.",
                                [[{"text": "↩️ بازگشت", "callback_data": f"trigger_list_close:{tl_cid}"}]])
                        else:
                            # گروه‌بندی بر اساس مسئول
                            by_member = {}
                            unassigned = []
                            for row_tl in rows_tl:
                                m = row_tl.get("assigned_to", "")
                                tag = row_tl.get("alarm_tag", "—")
                                shift_tl = row_tl.get("shift", "")
                                if m:
                                    by_member.setdefault(m, []).append((tag, shift_tl))
                                else:
                                    unassigned.append((tag, shift_tl))
                            lines_tl = ["🎯 <b>لیست تریگر فعال</b>\n"]
                            for member, items in sorted(by_member.items()):
                                marker = " 👈" if member == my_name_tl else ""
                                lines_tl.append(f"👤 <b>{member}</b>{marker}")
                                for tag_tl, sh_tl in items:
                                    lines_tl.append(f"   • {tag_tl}")
                                lines_tl.append("")
                            if unassigned:
                                lines_tl.append("⏳ <b>بدون مسئول</b>")
                                for tag_tl, sh_tl in unassigned:
                                    lines_tl.append(f"   • {tag_tl}")
                            full_tl = "\n".join(lines_tl)
                            if len(full_tl) > 4000:
                                full_tl = full_tl[:3980] + "\n\n<i>...</i>"
                            kb_tl = [[{"text": "🔄 بروزرسانی", "callback_data": f"trigger_list:{tl_cid}"},
                                      {"text": "✕ بستن",       "callback_data": f"trigger_list_close:{tl_cid}"}]]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, full_tl, kb_tl)

                    elif cbq_data.startswith("trigger_list_close:"):
                        answer_callback(token_cbq, cbq_id, "بسته شد")
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=8, headers=H)
                        except: pass

                    elif cbq_data.startswith("weekly_menu:"):
                        answer_callback(token_cbq, cbq_id, "")
                        wm_cid = cbq_data.split(":", 1)[1]
                        kb_menu = [
                            [{"text": "📅 این هفته", "callback_data": f"weekly_report:{wm_cid}:this:0"},
                             {"text": "📅 هفته قبل", "callback_data": f"weekly_report:{wm_cid}:last:0"}],
                        ]
                        if APP_BASE_URL:
                            kb_menu.append([{"text": "🌐 نسخه وب", "url": f"{APP_BASE_URL}/report/weekly"}])
                        kb_menu.append([{"text": "✕ بستن", "callback_data": f"weekly_report_close:{wm_cid}"}])
                        send_tg_keyboard(token_cbq, cbq_cid, "📊 <b>آنالیز هفتگی تیم</b>\n\nیک گزینه رو انتخاب کن:", kb_menu)

                    elif cbq_data.startswith("weekly_report:"):
                        answer_callback(token_cbq, cbq_id, "⏳ در حال بارگذاری...")
                        wr_parts = cbq_data.split(":")
                        cbq_cid_wr = wr_parts[1]
                        wr_which   = wr_parts[2] if len(wr_parts) > 2 else "this"
                        wr_page    = int(wr_parts[3]) if len(wr_parts) > 3 else 0
                        PER_PAGE   = 5
                        now_dt_wr = datetime.now(TEHRAN)
                        days_since_sat = (now_dt_wr.weekday() - 5) % 7
                        this_week_start = (now_dt_wr - timedelta(days=days_since_sat)).replace(
                            hour=0, minute=0, second=0, microsecond=0)
                        if wr_which == "last":
                            week_start = this_week_start - timedelta(days=7)
                            week_end   = this_week_start
                        else:
                            week_start = this_week_start
                            week_end   = None
                        week_start_str = week_start.strftime("%Y-%m-%dT%H:%M:%S")
                        week_label = f"{week_start.strftime('%d/%m')} — {(week_end - timedelta(days=1)).strftime('%d/%m') if week_end else 'الان'}"
                        rows_wr = []
                        if SUPABASE_KEY:
                            try:
                                url_wr = (f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                                          f"?fired_at=gte.{week_start_str}&select=*&order=fired_at.asc")
                                if week_end:
                                    url_wr += f"&fired_at=lt.{week_end.strftime('%Y-%m-%dT%H:%M:%S')}"
                                r_wr = requests.get(url_wr, headers=_sb_h(), timeout=10)
                                if r_wr.status_code == 200:
                                    rows_wr = r_wr.json()
                            except Exception as e:
                                print(f"[weekly] load exc: {e}")
                        # لود فقط alertهایی که تو این بازه‌ی هفته هستن (نه کل جدول)
                        wr_ids = sorted({str(r.get("id","")) for r in rows_wr if r.get("id")})
                        alerts_by_id_wr = {}
                        if SUPABASE_KEY and wr_ids:
                            try:
                                r_wr2 = requests.get(
                                    f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(wr_ids)})&select=*",
                                    headers=_sb_h(), timeout=10)
                                if r_wr2.status_code == 200:
                                    for row_wr in r_wr2.json():
                                        alerts_by_id_wr[str(row_wr.get("id",""))] = row_wr
                            except Exception as e:
                                print(f"[weekly] targeted alerts fetch error: {e}")
                        all_alerts_wr = list(alerts_by_id_wr.values())
                        private_ids_wr = {str(a["id"]) for a in all_alerts_wr if a.get("is_private")}
                        rows_wr = [r for r in rows_wr if str(r.get("id","")) not in private_ids_wr]
                        # صفحه‌بندی
                        total = len(rows_wr)
                        total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
                        wr_page = max(0, min(wr_page, total_pages - 1))
                        page_rows = rows_wr[wr_page * PER_PAGE:(wr_page + 1) * PER_PAGE]

                        # فقط صفحه‌بندی و بستن
                        kb_wr_nav = []
                        if total_pages > 1:
                            page_row = []
                            if wr_page > 0:
                                page_row.append({"text": "‹ قبلی", "callback_data": f"weekly_report:{cbq_cid_wr}:{wr_which}:{wr_page-1}"})
                            page_row.append({"text": f"{wr_page+1}/{total_pages}", "callback_data": "noop"})
                            if wr_page < total_pages - 1:
                                page_row.append({"text": "بعدی ›", "callback_data": f"weekly_report:{cbq_cid_wr}:{wr_which}:{wr_page+1}"})
                            kb_wr_nav.append(page_row)
                        kb_wr_nav.append([{"text": "🔍 جستجو", "callback_data": f"weekly_search:{cbq_cid_wr}:{wr_which}"},
                                          {"text": "✕ بستن", "callback_data": f"weekly_report_close:{cbq_cid_wr}"}])

                        if not rows_wr:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"📋 <b>گزارش هفتگی تیم</b>\n<i>{week_label}</i>\n\n📭 هیچ آلارم تیمی ثبت نشده.",
                                kb_wr_nav)
                        else:
                            lines_wr = [
                                f"📋 <b>گزارش هفتگی تیم</b>",
                                f"<i>{week_label} — {total} آلارم | صفحه {wr_page+1}/{total_pages}</i>",
                                ""
                            ]
                            for row_wr in page_rows:
                                aid_wr       = str(row_wr.get("id", ""))
                                tag_wr       = row_wr.get("alarm_tag", "—")
                                assignee_wr  = row_wr.get("assigned_to", "") or "—"
                                fired_wr     = row_wr.get("fired_at", "")[:16]
                                false_by_wr  = row_wr.get("false_by", "") or ""
                                false_at_wr  = row_wr.get("false_at", "")[:16] if row_wr.get("false_at") else ""
                                false_rsn_wr = row_wr.get("false_reason", "") or ""
                                is_active_wr = row_wr.get("is_active", True)
                                alert_wr   = alerts_by_id_wr.get(aid_wr, {})
                                sym_wr     = alert_wr.get("symbol", "") or row_wr.get("symbol", "") or ""
                                tgt_raw    = alert_wr.get("target_price", 0) or row_wr.get("target_price", 0) or 0
                                target_wr  = fmt_price(float(tgt_raw), sym_wr) if tgt_raw else "—"
                                creator_wr = alert_wr.get("created_by", "") or row_wr.get("created_by", "") or "—"
                                created_wr = str(alert_wr.get("created_at", ""))[:16]
                                cond_wr    = alert_wr.get("condition", "") or row_wr.get("condition", "")
                                dir_wr     = "📈 ناحیه سل" if cond_wr == "above" else ("📉 ناحیه بای" if cond_wr == "below" else "")
                                lines_wr.append(f"🔖 <b>{tag_wr}</b>  |  #{sym_wr}" + (f"  |  {dir_wr}" if dir_wr else ""))
                                if created_wr:
                                    lines_wr.append(f"📅 ثبت: {created_wr}")
                                lines_wr.append(f"🎯 هدف: <code>{target_wr}</code>")
                                lines_wr.append(f"👤 سازنده: {creator_wr}")
                                lines_wr.append(f"⏰ فایر شد: {fired_wr}")
                                lines_wr.append(f"🙋 مسئول: {assignee_wr}")
                                if is_active_wr:
                                    lines_wr.append(f"✅ وضعیت: فعال")
                                else:
                                    hist_wr = row_wr.get("false_history") or []
                                    if isinstance(hist_wr, str):
                                        try: hist_wr = json.loads(hist_wr)
                                        except: hist_wr = []
                                    if hist_wr:
                                        lines_wr.append("❌ <b>تاریخچه False:</b>")
                                        for i_h, h_wr in enumerate(hist_wr, 1):
                                            h_at  = str(h_wr.get("at",""))[:16]
                                            h_by  = h_wr.get("by","")
                                            h_rsn = h_wr.get("reason","")
                                            h_line = f"  {i_h}. {h_by}  |  {h_at}"
                                            if h_rsn:
                                                h_line += f"\n     📝 {h_rsn}"
                                            lines_wr.append(h_line)
                                    else:
                                        lines_wr.append(f"❌ وضعیت: False — {false_by_wr}  |  {false_at_wr}")
                                        if false_rsn_wr:
                                            lines_wr.append(f"📝 علت: {false_rsn_wr}")
                                lines_wr.append("──────────────")
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, "\n".join(lines_wr), kb_wr_nav)

                    elif cbq_data.startswith("weekly_report_close:"):
                        answer_callback(token_cbq, cbq_id, "بسته شد")
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=8, headers=H)
                        except: pass

                    elif cbq_data.startswith("weekly_search:"):
                        answer_callback(token_cbq, cbq_id, "")
                        ws_parts = cbq_data.split(":")
                        ws_cid   = ws_parts[1]
                        ws_which = ws_parts[2] if len(ws_parts) > 2 else "this"
                        _pending_weekly_search[ws_cid] = {"which": ws_which, "msg_id": cbq_msg_id}
                        send_tg(token_cbq, ws_cid, "🔍 اسم مسئول، نماد یا تگ رو بفرست (مثلاً: مسعود):")

                    elif cbq_data.startswith("resend_active:"):
                        answer_callback(token_cbq, cbq_id, "⏳ در حال ارسال...")
                        ra_cid = cbq_data.split(":", 1)[1]
                        rows_ra = _sb_load_active_assignments()
                        # فقط تیمی
                        all_alerts_ra = load_alerts().get("alarms", [])
                        private_ids_ra = {str(a["id"]) for a in all_alerts_ra if a.get("is_private")}
                        rows_ra = [r for r in rows_ra if str(r.get("id","")) not in private_ids_ra]
                        if not rows_ra:
                            send_tg(token_cbq, ra_cid, "📭 هیچ آلارم فعالی در لیست تریگر نیست.")
                        else:
                            send_tg(token_cbq, ra_cid,
                                f"🔔 <b>{len(rows_ra)} آلارم فعال</b> — ریپلای زیر:")
                            for row_ra in rows_ra:
                                aid_ra = str(row_ra.get("id", ""))
                                tag_ra = row_ra.get("alarm_tag", "—")
                                assignee_ra = row_ra.get("assigned_to", "") or "⏳ منتظر تقسیم"
                                # پیدا کردن message_id اصلی برای این کاربر
                                cid_map_ra = _fired_msg_ids.get(aid_ra, {})
                                orig_mid = cid_map_ra.get(ra_cid)
                                orig_text = cid_map_ra.get("__text__", "")
                                if orig_mid:
                                    # ریپلای روی همون پیام اصلی
                                    try:
                                        requests.post(
                                            f"https://api.telegram.org/bot{token_cbq}/sendMessage",
                                            json={"chat_id": ra_cid,
                                                  "text": f"🔔 <b>{tag_ra}</b>\n👤 مسئول: {assignee_ra}",
                                                  "parse_mode": "HTML",
                                                  "reply_to_message_id": orig_mid},
                                            timeout=8, headers=H)
                                    except: pass
                                else:
                                    # پیام اصلی پیدا نشد — متن خلاصه بفرست
                                    send_tg(token_cbq, ra_cid,
                                        f"🔔 <b>{tag_ra}</b>\n👤 مسئول: {assignee_ra}")

                    elif cbq_data.startswith("today_alarms:"):
                        answer_callback(token_cbq, cbq_id, "⏳ در حال بارگذاری...")
                        ta_parts = cbq_data.split(":")
                        ta_cid   = ta_parts[1]
                        ta_mode  = ta_parts[2] if len(ta_parts) > 2 else "active"
                        # ابتدای هفته — شنبه تهران
                        now_dt_ta = datetime.now(TEHRAN)
                        days_since_sat = (now_dt_ta.weekday() - 5) % 7
                        week_start_ta = (now_dt_ta - timedelta(days=days_since_sat)).replace(
                            hour=0, minute=0, second=0, microsecond=0)
                        week_start_str_ta = week_start_ta.strftime("%Y-%m-%dT%H:%M:%S")
                        rows_ta = []
                        if SUPABASE_KEY:
                            try:
                                if ta_mode == "active":
                                    url_ta = (
                                        f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                                        f"?fired_at=gte.{week_start_str_ta}&is_active=eq.true"
                                        f"&select=*&order=fired_at.asc"
                                    )
                                else:
                                    url_ta = (
                                        f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                                        f"?fired_at=gte.{week_start_str_ta}"
                                        f"&select=*&order=fired_at.asc"
                                    )
                                r_ta = requests.get(url_ta, headers=_sb_h(), timeout=10)
                                if r_ta.status_code == 200:
                                    rows_ta = r_ta.json()
                            except Exception as e:
                                print(f"[weekly_list] load exc: {e}")
                        # فقط تیمی
                        all_alerts_ta = load_alerts().get("alarms", [])
                        private_ids_ta = {str(a["id"]) for a in all_alerts_ta if a.get("is_private")}
                        rows_ta = [r for r in rows_ta if str(r.get("id","")) not in private_ids_ta]
                        mode_label = "فعال" if ta_mode == "active" else "همه"
                        week_label = week_start_ta.strftime("%d/%m")
                        if not rows_ta:
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                                f"📋 <b>هفتگی ({mode_label}) از {week_label}</b>\n\n📭 آلارمی پیدا نشد.",
                                [[{"text": "✅ فعال", "callback_data": f"today_alarms:{ta_cid}:active"},
                                  {"text": "📊 همه",  "callback_data": f"today_alarms:{ta_cid}:all"}],
                                 [{"text": "✕ بستن", "callback_data": f"today_alarms_close:{ta_cid}"}]])
                        else:
                            lines_ta = [f"📋 <b>هفتگی ({mode_label}) از {week_label} — {len(rows_ta)} آلارم</b>", ""]
                            alerts_by_id_ta = {str(a["id"]): a for a in all_alerts_ta}
                            for row_ta in rows_ta:
                                aid_ta      = str(row_ta.get("id",""))
                                tag_ta      = row_ta.get("alarm_tag", "—")
                                assignee_ta = row_ta.get("assigned_to", "") or "⏳ منتظر تقسیم"
                                fired_ta    = row_ta.get("fired_at", "")[:16]
                                is_act_ta   = row_ta.get("is_active", True)
                                alert_ta    = alerts_by_id_ta.get(aid_ta, {})
                                sym_ta      = alert_ta.get("symbol", "")
                                target_ta   = alert_ta.get("target_price", "") or alert_ta.get("price", "")
                                status_icon = "✅" if is_act_ta else "❌"
                                lines_ta.append(f"{status_icon} <b>{tag_ta}</b>  {sym_ta}")
                                if target_ta:
                                    lines_ta.append(f"   🎯 هدف: <code>{target_ta}</code>")
                                lines_ta.append(f"   ⏰ {fired_ta}  |  👤 {assignee_ta}")
                                lines_ta.append("")
                            full_ta = "\n".join(lines_ta)
                            if len(full_ta) > 4000:
                                full_ta = full_ta[:3980] + "\n<i>...</i>"
                            kb_ta = [
                                [{"text": "✅ فعال", "callback_data": f"today_alarms:{ta_cid}:active"},
                                 {"text": "📊 همه",  "callback_data": f"today_alarms:{ta_cid}:all"}],
                                [{"text": "🔄 بروزرسانی", "callback_data": f"today_alarms:{ta_cid}:{ta_mode}"},
                                 {"text": "✕ بستن",        "callback_data": f"today_alarms_close:{ta_cid}"}]
                            ]
                            edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, full_ta, kb_ta)

                    elif cbq_data.startswith("today_alarms_close:"):
                        answer_callback(token_cbq, cbq_id, "بسته شد")
                        try:
                            requests.post(
                                f"https://api.telegram.org/bot{token_cbq}/deleteMessage",
                                json={"chat_id": cbq_cid, "message_id": cbq_msg_id},
                                timeout=8, headers=H)
                        except: pass

                    elif cbq_data.startswith("sig_send:"):
                        parts_snd = cbq_data.split(":", 2)
                        send_cid  = parts_snd[1]
                        send_mode = parts_snd[2] if len(parts_snd) > 2 else "channel"
                        ps = _pending_signal.get(send_cid)
                        if not ps:
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد")
                            return
                        answer_callback(token_cbq, cbq_id, "⏳ در حال ثبت...")
                        d_send = ps["data"]
                        seq = _sb_next_signal_seq()
                        sig_id = f"S{seq:05d}"
                        d_send["id"] = sig_id
                        d_send["seq"] = seq
                        d_send["sent_by"] = _get_user_custom_name(send_cid) or send_cid
                        d_send["sent_at"] = now_teh()
                        d_send["status"]  = "active"
                        d_send["channel_msg_id"] = None
                        # ارسال به کانال فقط در حالت channel
                        channel_mid = None
                        if send_mode == "channel" and SIGNAL_CHANNEL:
                            try:
                                r_ch = requests.post(
                                    f"https://api.telegram.org/bot{token_cbq}/sendMessage",
                                    json={"chat_id": SIGNAL_CHANNEL, "text": _build_signal_text(d_send),
                                          "parse_mode": "HTML"},
                                    timeout=10, headers=H)
                                if r_ch.status_code == 200:
                                    channel_mid = r_ch.json().get("result",{}).get("message_id")
                                    d_send["channel_msg_id"] = channel_mid
                            except Exception as e:
                                print(f"[signal] channel send error: {e}")
                        # همیشه در Supabase ذخیره میشه
                        threading.Thread(target=_sb_save_signal, args=(d_send,), daemon=True).start()
                        del _pending_signal[send_cid]
                        sig_text_conf = _build_signal_text(d_send)
                        if send_mode == "channel":
                            status_line = f"📤 ارسال شد به گروه" if channel_mid else "⚠️ کانال تنظیم نشده — فقط ذخیره شد"
                        else:
                            status_line = "💾 فقط در دیتابیس ثبت شد"
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"✅ <b>سیگنال {sig_id} ثبت شد</b>  {status_line}\n\n{sig_text_conf}", [])

                    elif cbq_data.startswith("alarm_dir:"):
                        # alarm_dir:cid:buy|sell
                        parts_ad = cbq_data.split(":", 2)
                        ad_cid = parts_ad[1] if len(parts_ad) > 1 else cbq_cid
                        ad_raw = parts_ad[2] if len(parts_ad) > 2 else "sell"
                        pend_ad = _pending_alarm.get(ad_cid)
                        if not pend_ad or pend_ad.get("step") != "alarm_dir":
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد — دوباره شروع کن")
                            return
                        answer_callback(token_cbq, cbq_id)
                        dw_ad = pend_ad["data"]
                        if ad_raw == "buy":
                            dw_ad["condition"] = "below"
                            dir_lbl_ad = "📈 BUY"
                        else:
                            dw_ad["condition"] = "above"
                            dir_lbl_ad = "📉 SELL"
                        pend_ad["step"] = "alarm_price"
                        ptype_lbl_ad = "🔒 شخصی" if dw_ad.get("ptype") == "private" else "🌐 تیمی"
                        edit_tg_keyboard(token_cbq, ad_cid, cbq_msg_id,
                            f"🔔 <b>{dw_ad['symbol']}</b>  {dir_lbl_ad}  ({ptype_lbl_ad})\n\nقیمت هدف رو بنویس:\nمثال: <code>1.08500</code>  یا  <code>2350</code>",
                            [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{ad_cid}"}]])

                    elif cbq_data.startswith("alarm_submit:"):
                        # «ثبت بدون یادداشت» — بعدش هنوز باید مدت اعتبار انتخاب بشه
                        as_cid = cbq_data.split(":", 1)[1]
                        pend_as = _pending_alarm.get(as_cid)
                        if not pend_as or pend_as.get("step") != "alarm_comment":
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد — دوباره شروع کن")
                            return
                        answer_callback(token_cbq, cbq_id, "")
                        pend_as["data"]["comment"] = ""
                        pend_as["step"] = "alarm_expiry"
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"🔔 <b>{pend_as['data'].get('symbol','')}</b>\n\n⏳ این آلارم چند روز معتبر باشه؟ (اگه تا اون موقع فایر نشه، خودکار حذف می‌شه)",
                            [
                                [{"text": "۱ روز", "callback_data": f"alarm_expiry:{as_cid}:1"},
                                 {"text": "۳ روز", "callback_data": f"alarm_expiry:{as_cid}:3"},
                                 {"text": "۷ روز", "callback_data": f"alarm_expiry:{as_cid}:7"}],
                                [{"text": "♾ بدون انقضا", "callback_data": f"alarm_expiry:{as_cid}:0"}],
                                [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{as_cid}"}]
                            ])

                    elif cbq_data.startswith("alarm_expiry:"):
                        # مرحله‌ی آخر — انتخاب مدت اعتبار، و ثبت نهایی آلارم
                        parts_ae = cbq_data.split(":", 2)
                        ae_cid = parts_ae[1] if len(parts_ae) > 1 else cbq_cid
                        ae_days = int(parts_ae[2]) if len(parts_ae) > 2 and parts_ae[2].isdigit() else 0
                        pend_ae = _pending_alarm.get(ae_cid)
                        if not pend_ae or pend_ae.get("step") != "alarm_expiry":
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد — دوباره شروع کن")
                            return
                        answer_callback(token_cbq, cbq_id, "⏳ در حال ثبت...")
                        dw_ae = pend_ae["data"]
                        ae_uname = cbq.get("from",{}).get("username","") or cbq.get("from",{}).get("first_name","")
                        is_private_ae = dw_ae.get("ptype","public") == "private"
                        sender_ae = _get_user_custom_name(ae_cid) or ae_uname
                        sym_ae = dw_ae["symbol"]
                        atype_ae = dw_ae["atype"]
                        comment_ae = dw_ae.get("comment","")
                        expires_at_ae = None
                        if ae_days > 0:
                            expires_at_ae = (datetime.now(TEHRAN) + timedelta(days=ae_days)).strftime("%Y-%m-%d %H:%M:%S")
                        new_alert_ae = {
                            "id": str(int(time.time()*1000)),
                            "symbol": sym_ae, "type": atype_ae,
                            "target_price": dw_ae["target_price"], "condition": dw_ae["condition"],
                            "comment": comment_ae, "created_by": sender_ae,
                            "active": True, "last_price": None, "last_checked": None,
                            "created_at": now_teh(),
                            "expires_at": expires_at_ae,
                            "is_private": is_private_ae,
                            "private_cid": ae_cid if is_private_ae else None,
                            "notify_only": ae_cid if is_private_ae else (YOUR_CHAT_ID if not BROADCAST_MODE else None)
                        }
                        d_ae = load_alerts()
                        d_ae["alerts"].append(new_alert_ae)
                        _sb_upsert_alert(new_alert_ae)
                        _cache_alerts = d_ae
                        del _pending_alarm[ae_cid]
                        dir_lbl_ae = "📈 BUY" if new_alert_ae["condition"] == "below" else "📉 SELL"
                        priv_lbl_ae = "  🔒 شخصی" if is_private_ae else "  🌐 تیمی"
                        expiry_lbl_ae = f"\n⏳ اعتبار: {ae_days} روز (تا {expires_at_ae[:10]})" if expires_at_ae else "\n♾ بدون انقضا"
                        confirm_ae = (
                            f"✅ <b>آلارم ثبت شد!</b>\n\n"
                            f"💰 <b>{sym_ae}</b>  {dir_lbl_ae}{priv_lbl_ae}\n"
                            f"🎯 هدف: <code>{fmt_price(new_alert_ae['target_price'], sym_ae)}</code>\n"
                            + (f"💬 {comment_ae}\n" if comment_ae else "")
                            + expiry_lbl_ae
                            + f"\n\n⏰ {now_teh()} (تهران)"
                        )
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id, confirm_ae, [])
                        def _bg_ae(alert=new_alert_ae, s=sym_ae, t=atype_ae):
                            try:
                                cur = get_price(s, t)
                                if cur:
                                    alert["last_price"] = cur
                                    alert["last_checked"] = now_teh()
                                    _sb_upsert_alert(alert)
                            except: pass
                        threading.Thread(target=_bg_ae, daemon=True).start()

                    elif cbq_data.startswith("sos_dir:"):
                        # sos_dir:cid:buy|sell
                        parts_sd = cbq_data.split(":", 2)
                        sd_cid = parts_sd[1] if len(parts_sd) > 1 else cbq_cid
                        sd_raw = parts_sd[2] if len(parts_sd) > 2 else "sell"
                        pend = _pending_alarm.get(sd_cid)
                        if not pend or pend.get("step") != "sos_dir":
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد — دوباره شروع کن")
                            return
                        answer_callback(token_cbq, cbq_id)
                        dw_sd = pend["data"]
                        if sd_raw == "buy":
                            dw_sd["condition"] = "below"
                            dw_sd["dir_lbl"] = "📈 BUY"
                        else:
                            dw_sd["condition"] = "above"
                            dw_sd["dir_lbl"] = "📉 SELL"
                        pend["step"] = "sos_comment"
                        sym_sd = dw_sd.get("symbol", "")
                        dir_lbl_sd = dw_sd["dir_lbl"]
                        edit_tg_keyboard(token_cbq, sd_cid, cbq_msg_id,
                            f"⚡ <b>{sym_sd}</b>  {dir_lbl_sd}\n\nیادداشت اختیاری بنویس\nیا دکمه «بدون یادداشت» رو بزن:",
                            [
                                [{"text": "✅ ارسال بدون یادداشت", "callback_data": f"sos_nocomment:{sd_cid}"}],
                                [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{sd_cid}"}]
                            ])

                    elif cbq_data.startswith("sos_nocomment:"):
                        # کاربر بدون یادداشت ارسال کرد
                        sc_cid = cbq_data.split(":", 1)[1]
                        pend_sc = _pending_alarm.get(sc_cid)
                        if not pend_sc or pend_sc.get("step") != "sos_comment":
                            answer_callback(token_cbq, cbq_id, "⚠️ جلسه منقضی شد — دوباره شروع کن")
                            return
                        answer_callback(token_cbq, cbq_id, "⏳ در حال ارسال...")
                        dw_sc = pend_sc["data"]
                        sym_sc = dw_sc["symbol"]
                        condition_sc = dw_sc["condition"]
                        dir_lbl_sc = dw_sc.get("dir_lbl", "📈 BUY" if condition_sc == "below" else "📉 SELL")
                        atype_sc = "forex" if any(x in sym_sc for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                        sender_sc = _get_user_custom_name(sc_cid) or cbq_cid
                        alarm_num_tag_sc = _make_alarm_tag(sym_sc)
                        sender_tag_sc = "#" + re.sub(r"[^\w]","_", sender_sc).strip("_")
                        arrow_sc = "📈 ناحیه سل" if condition_sc == "above" else "📉 ناحیه بای"
                        try: cur_sc = get_price(sym_sc, atype_sc)
                        except: cur_sc = None
                        out_sc = (
                            f"🚨 <b>آلارم فوری!</b>\n"
                            f"━━━━━━━━━━━━━━━━━━\n"
                            f"💰 <b>#{sym_sc}</b>  {arrow_sc}\n"
                            f"🔖 {alarm_num_tag_sc}\n"
                            f"👤 {sender_tag_sc}\n"
                            f"📊 قیمت: <b>{fmt_price(cur_sc, sym_sc) if cur_sc else '—'}</b>\n"
                            f"━━━━━━━━━━━━━━━━━━\n"
                            f"⏰ {now_pretty()} (تهران)"
                        )
                        _, all_cids_sc, _ = _get_token_and_cids()
                        targets_sc = all_cids_sc if BROADCAST_MODE else [YOUR_CHAT_ID]
                        del _pending_alarm[sc_cid]
                        edit_tg_keyboard(token_cbq, cbq_cid, cbq_msg_id,
                            f"✅ <b>آلارم فوری ارسال شد!</b>\n\n"
                            f"💰 <b>{sym_sc}</b>  {dir_lbl_sc}\n"
                            f"⏰ {now_pretty()} (تهران)", [])
                        sos_aid = f"sos_{sym_sc}_{int(time.time())}"
                        def _bg_sos(tok=token_cbq, tgts=targets_sc, msg=out_sc, s=sym_sc, aid=sos_aid,
                                    atag=alarm_num_tag_sc, sndr=sender_sc, cond=condition_sc, atp=atype_sc, cur=cur_sc):
                            sos_cid_to_mid = {}
                            if s.upper() not in TEMP_MUTED_SYMBOLS:
                                for tc_sc in tgts:
                                    kb_sc = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc_sc}:{s}"}]]
                                    mid_sc = send_tg_keyboard(tok, str(tc_sc), msg, kb_sc, track=False)
                                    if mid_sc:
                                        sos_cid_to_mid[str(tc_sc)] = mid_sc
                            if sos_cid_to_mid:
                                sos_cid_to_mid["__tag__"] = atag
                                sos_cid_to_mid["__text__"] = msg
                                _fired_msg_ids[aid] = sos_cid_to_mid
                                threading.Thread(target=_sb_save_fired_msgs, args=(aid, sos_cid_to_mid), daemon=True).start()
                            # ذخیره توی archive
                            sos_arch_entry = {"id": aid, "symbol": s, "type": atp,
                                "condition": cond, "comment": "", "created_by": sndr,
                                "active": False, "fired_at": now_teh(), "fired_price": cur,
                                "instant": True, "created_at": now_teh(), "tag": atag}
                            d_arc = load_alerts()
                            d_arc.setdefault("archive", []).append(sos_arch_entry)
                            save_alerts(d_arc)
                            threading.Thread(target=_sb_upsert_alert, args=(sos_arch_entry,), daemon=True).start()
                        threading.Thread(target=_bg_sos, daemon=True).start()

                msg = upd.get("message", {})
                raw_txt = msg.get("text", "") or ""
                # normalize: /cmd@botname → /cmd
                txt = raw_txt.split("@")[0] if raw_txt.startswith("/") else raw_txt
                ch = msg.get("chat", {})
                cid = str(ch.get("id", ""))
                uname = ch.get("username", "") or ch.get("first_name", "")

                # track پیام کاربر برای پاک‌سازی
                user_msg_id = msg.get("message_id")
                if cid and user_msg_id:
                    _track_msg(cid, user_msg_id)

                # ── /start ──────────────────────────────────────────
                if txt.startswith("/start") and cid:
                    data = load_alerts()
                    users = data.get("users", [])
                    existing_user = next((u for u in users if str(u.get("chat_id","")) == cid), None)
                    existing_name = existing_user.get("custom_name","").strip() if existing_user else ""
                    # ثبت کاربر اگه جدیده
                    if not existing_user:
                        users.append({"chat_id": cid, "username": uname, "joined_at": now_teh(), "custom_name": "", "private_access": False, "approved": False})
                        data["users"] = users
                        ids = data.get("telegram", {}).get("chat_ids", [])
                        if cid not in [str(x) for x in ids]:
                            ids.append(cid)
                        data["telegram"]["chat_ids"] = ids
                        save_alerts(data)
                    is_adm = (cid == YOUR_CHAT_ID)
                    if existing_name:
                        # کاربر قدیمی که هنوز کد فعال‌سازی سایت (web_pin) نداره —
                        # الان که وارد شده، یه بار براش می‌سازیم تا دیگه کسی نتونه با تایپ کردن اسمش تو سایت جاش بزنه
                        if existing_user and not existing_user.get("web_pin"):
                            new_pin_old = f"{secrets.randbelow(1000000):06d}"
                            for u in users:
                                if str(u.get("chat_id","")) == cid:
                                    u["web_pin"] = new_pin_old
                                    break
                            data["users"] = users
                            save_alerts(data)
                            send_tg(token, cid,
                                f"🔑 <b>کد فعال‌سازی سایت شما:</b> <code>{new_pin_old}</code>\n\n"
                                "برای جلوگیری از سوءاستفاده‌ی افراد دیگه از اسمت، این کد رو یک‌بار توی سایت "
                                "(دکمه «🔑 کد فعال‌سازی» زیر نام کاربری) وارد کن.")
                        # کاربر قبلاً اسم داده — مستقیم به منو برو
                        show_main_menu(token, cid,
                            f"👋 خوش برگشتی <b>{existing_name}</b>!\n\nاز منوی زیر انتخاب کن 👇",
                            is_adm)
                    else:
                        # اولین بار — اسم بخواه
                        _pending_name[cid] = True
                        send_tg(token, cid,
                            f"👋 سلام <b>{uname}</b>!\n\n"
                            f"لطفاً <b>اسمی که در سایت استفاده می‌کنی</b> رو بنویس:\n"
                            f"(آلارم‌های شخصیت با همین اسم شناسایی میشن)")

                # ── دریافت کوئری جستجوی گزارش هفتگی ──────────────────
                elif cid in _pending_weekly_search and not txt.startswith("/"):
                    ws_info = _pending_weekly_search.pop(cid)
                    ws_which = ws_info["which"]
                    query = txt.strip().lower()

                    now_dt_ws = datetime.now(TEHRAN)
                    days_since_sat_ws = (now_dt_ws.weekday() - 5) % 7
                    this_week_start_ws = (now_dt_ws - timedelta(days=days_since_sat_ws)).replace(
                        hour=0, minute=0, second=0, microsecond=0)
                    if ws_which == "last":
                        week_start_ws = this_week_start_ws - timedelta(days=7)
                        week_end_ws   = this_week_start_ws
                    else:
                        week_start_ws = this_week_start_ws
                        week_end_ws   = None
                    week_start_str_ws = week_start_ws.strftime("%Y-%m-%dT%H:%M:%S")
                    week_label_ws = f"{week_start_ws.strftime('%d/%m')} — {(week_end_ws - timedelta(days=1)).strftime('%d/%m') if week_end_ws else 'الان'}"

                    rows_ws = []
                    if SUPABASE_KEY:
                        try:
                            url_ws = (f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                                      f"?fired_at=gte.{week_start_str_ws}&select=*&order=fired_at.asc")
                            if week_end_ws:
                                url_ws += f"&fired_at=lt.{week_end_ws.strftime('%Y-%m-%dT%H:%M:%S')}"
                            r_ws = requests.get(url_ws, headers=_sb_h(), timeout=10)
                            if r_ws.status_code == 200:
                                rows_ws = r_ws.json()
                        except Exception as e:
                            print(f"[weekly_search] load exc: {e}")

                    # لود فقط alertهایی که تو این بازه‌ی هفته هستن (نه کل جدول)
                    ws_ids = sorted({str(r.get("id","")) for r in rows_ws if r.get("id")})
                    alerts_by_id_ws = {}
                    if SUPABASE_KEY and ws_ids:
                        try:
                            r_ws2 = requests.get(
                                f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(ws_ids)})&select=*",
                                headers=_sb_h(), timeout=10)
                            if r_ws2.status_code == 200:
                                for row_ws in r_ws2.json():
                                    alerts_by_id_ws[str(row_ws.get("id",""))] = row_ws
                        except Exception as e:
                            print(f"[weekly_search] targeted alerts fetch error: {e}")
                    all_alerts_ws = list(alerts_by_id_ws.values())
                    private_ids_ws = {str(a["id"]) for a in all_alerts_ws if a.get("is_private")}
                    rows_ws = [r for r in rows_ws if str(r.get("id","")) not in private_ids_ws]

                    # فیلتر بر اساس کوئری — مسئول، نماد، تگ، سازنده
                    filtered = []
                    for row_ws in rows_ws:
                        aid_ws  = str(row_ws.get("id",""))
                        alert_ws = alerts_by_id_ws.get(aid_ws, {})
                        sym_ws  = (alert_ws.get("symbol","") or row_ws.get("symbol","") or "").lower()
                        tag_ws  = (row_ws.get("alarm_tag","") or "").lower()
                        assignee_ws = (row_ws.get("assigned_to","") or "").lower()
                        creator_ws  = (alert_ws.get("created_by","") or row_ws.get("created_by","") or "").lower()
                        if query in sym_ws or query in tag_ws or query in assignee_ws or query in creator_ws:
                            filtered.append(row_ws)

                    if not filtered:
                        send_tg(token, cid, f"🔍 نتیجه‌ای برای «<b>{txt.strip()}</b>» پیدا نشد.")
                    else:
                        lines_ws = [f"🔍 <b>نتایج جستجو: {txt.strip()}</b>",
                                     f"<i>{week_label_ws} — {len(filtered)} نتیجه</i>", ""]
                        for row_ws in filtered[:10]:
                            aid_ws       = str(row_ws.get("id",""))
                            tag_ws       = row_ws.get("alarm_tag","—")
                            assignee_ws  = row_ws.get("assigned_to","") or "—"
                            fired_ws     = row_ws.get("fired_at","")[:16]
                            is_active_ws = row_ws.get("is_active", True)
                            false_by_ws  = row_ws.get("false_by","") or ""
                            false_at_ws  = row_ws.get("false_at","")[:16] if row_ws.get("false_at") else ""
                            false_rsn_ws = row_ws.get("false_reason","") or ""
                            alert_ws     = alerts_by_id_ws.get(aid_ws, {})
                            sym_ws       = alert_ws.get("symbol","") or row_ws.get("symbol","") or ""
                            tgt_raw_ws   = alert_ws.get("target_price",0) or row_ws.get("target_price",0) or 0
                            target_ws    = fmt_price(float(tgt_raw_ws), sym_ws) if tgt_raw_ws else "—"
                            creator_ws   = alert_ws.get("created_by","") or row_ws.get("created_by","") or "—"
                            created_ws   = str(alert_ws.get("created_at",""))[:16]
                            cond_ws      = alert_ws.get("condition","") or row_ws.get("condition","")
                            dir_ws       = "📈 ناحیه سل" if cond_ws == "above" else ("📉 ناحیه بای" if cond_ws == "below" else "")
                            lines_ws.append(f"🔖 <b>{tag_ws}</b>  |  #{sym_ws}" + (f"  |  {dir_ws}" if dir_ws else ""))
                            if created_ws:
                                lines_ws.append(f"📅 ثبت: {created_ws}")
                            lines_ws.append(f"🎯 هدف: <code>{target_ws}</code>")
                            lines_ws.append(f"👤 سازنده: {creator_ws}")
                            lines_ws.append(f"⏰ فایر شد: {fired_ws}")
                            lines_ws.append(f"🙋 مسئول: {assignee_ws}")
                            if is_active_ws:
                                lines_ws.append(f"✅ وضعیت: فعال")
                            else:
                                hist_ws = row_ws.get("false_history") or []
                                if isinstance(hist_ws, str):
                                    try: hist_ws = json.loads(hist_ws)
                                    except: hist_ws = []
                                if hist_ws:
                                    lines_ws.append("❌ <b>تاریخچه False:</b>")
                                    for i_hw, h_ws in enumerate(hist_ws, 1):
                                        h_at_w  = str(h_ws.get("at",""))[:16]
                                        h_by_w  = h_ws.get("by","")
                                        h_rsn_w = h_ws.get("reason","")
                                        h_line_w = f"  {i_hw}. {h_by_w}  |  {h_at_w}"
                                        if h_rsn_w:
                                            h_line_w += f"\n     📝 {h_rsn_w}"
                                        lines_ws.append(h_line_w)
                                else:
                                    lines_ws.append(f"❌ وضعیت: False — {false_by_ws}  |  {false_at_ws}")
                                    if false_rsn_ws:
                                        lines_ws.append(f"📝 علت: {false_rsn_ws}")
                            lines_ws.append("──────────────")
                        if len(filtered) > 10:
                            lines_ws.append(f"<i>... و {len(filtered)-10} مورد دیگر</i>")
                        full_ws = "\n".join(lines_ws)
                        if len(full_ws) > 4000:
                            full_ws = full_ws[:3980] + "\n\n<i>...</i>"
                        send_tg(token, cid, full_ws)

                # ── دریافت اسم custom بعد از /start یا /setname ──────
                elif cid in _pending_name and not txt.startswith("/"):
                    custom_name = txt.strip()
                    if len(custom_name) < 2:
                        send_tg(token, cid, "⚠️ اسم باید حداقل ۲ حرف باشه. دوباره بنویس:")
                    else:
                        data = load_alerts()
                        users = data.get("users", [])
                        found = False
                        for usr in users:
                            if str(usr.get("chat_id", "")) == cid:
                                usr["custom_name"] = custom_name
                                found = True
                                break
                        if not found:
                            users.append({"chat_id": cid, "username": uname, "joined_at": now_teh(), "custom_name": custom_name, "approved": False})
                            data["users"] = users
                            ids = data.get("telegram", {}).get("chat_ids", [])
                            if cid not in [str(x) for x in ids]:
                                ids.append(cid)
                            data["telegram"]["chat_ids"] = ids
                        data["users"] = users
                        save_alerts(data)
                        del _pending_name[cid]
                        is_adm = (cid == YOUR_CHAT_ID)
                        show_main_menu(token, cid,
                            f"✅ خوش اومدی <b>{custom_name}</b>!\n\n"
                            f"آلارم‌هات با اسم <b>{custom_name}</b> شناسایی میشن.\n"
                            f"از منوی زیر انتخاب کن 👇",
                            is_adm)

                # ── /setname — تغییر اسم ────────────────────────────
                elif txt.startswith("/setname"):
                    _pending_name[cid] = True
                    data = load_alerts()
                    users = data.get("users", [])
                    cur = next((u.get("custom_name","") for u in users if str(u.get("chat_id",""))==cid), "")
                    cur_info = f"\nاسم فعلی: <b>{cur}</b>" if cur else ""
                    send_tg(token, cid, f"✏️ اسم جدیدت رو بنویس:{cur_info}")

                # ── /del — حذف پیام fired از همه چت‌ها ────────────
                # فرمت‌ها: ریپلای روی پیام + /del
                #          یا /del XAUUSD7
                elif txt.strip().startswith("/del"):
                    del_parts = txt.strip().split(maxsplit=1)
                    del_tag = del_parts[1].upper().lstrip("#") if len(del_parts) > 1 else None
                    replied = msg.get("reply_to_message", {})
                    replied_mid = replied.get("message_id")

                    target_aid = None

                    if del_tag:
                        # جستجو با هشتگ — مثلاً /del XAUUSD7
                        tag_search = f"#{del_tag}"
                        for aid, cid_map in _fired_msg_ids.items():
                            if cid_map.get("__tag__") == tag_search:
                                target_aid = aid
                                break
                        if not target_aid:
                            send_tg(token, cid, f"⚠️ آلارم <b>{tag_search}</b> پیدا نشد یا قبلاً پاک شده.")
                    elif replied_mid:
                        # جستجو با ریپلای
                        for aid, cid_map in _fired_msg_ids.items():
                            if str(cid_map.get(cid)) == str(replied_mid):
                                target_aid = aid
                                break
                        if not target_aid:
                            send_tg(token, cid, "⚠️ این پیام توی لیست آلارم‌های ذخیره‌شده نیست یا قبلاً پاک شده.")
                    else:
                        send_tg(token, cid, "⚠️ روی پیام آلارم ریپلای بزن یا بنویس /del XAUUSD7")

                    if target_aid:
                        cid_map = _fired_msg_ids.pop(target_aid, {})
                        deleted_count = 0
                        for tc, tm in cid_map.items():
                            if tc in ("__tag__", "__text__"): continue
                            try:
                                r_del = requests.post(
                                    f"https://api.telegram.org/bot{token}/deleteMessage",
                                    json={"chat_id": tc, "message_id": tm},
                                    timeout=5, headers=H)
                                if r_del.status_code == 200:
                                    deleted_count += 1
                            except: pass
                        # حذف پیام fired از Supabase
                        threading.Thread(target=_sb_delete_fired_msgs, args=(target_aid,), daemon=True).start()
                        # حذف خود آلارم از Supabase — تا دیگه fire نشه
                        threading.Thread(target=_sb_delete_alert, args=(target_aid,), daemon=True).start()
                        _cache_alerts = None
                        send_tg(token, cid, f"🗑 پیام آلارم از <b>{deleted_count}</b> چت پاک شد.")

                # ── /false — آلارم منقضی/ترید شده، از لیست تریگر خارج بشه ──
                elif txt.strip().lower() in ("/false",) or txt.strip().lower().startswith("/false"):
                    # کامنت بعد از /False — مثلاً: /False شکسته شد ناحیه
                    # یا /False XAUUSD5 برای جستجو با هشتگ
                    false_parts = txt.strip().split(maxsplit=1)
                    false_arg = false_parts[1].strip() if len(false_parts) > 1 else ""
                    false_reason = ""
                    false_tag_search = None
                    # چک کن اول arg یه هشتگ/تگ نماد هست؟ (مثلاً XAUUSD5 یا #XAUUSD5)
                    if false_arg:
                        candidate = false_arg.upper().lstrip("#").split()[0]
                        # اگه شبیه تگ نماد باشه (حروف + عدد، بدون فاصله)
                        import re as _re2
                        if _re2.match(r'^[A-Z0-9]+$', candidate) and any(c.isdigit() for c in candidate):
                            false_tag_search = f"#{candidate}"
                            # بقیه arg رو reason بدون
                            rest = false_arg[len(candidate):].strip().lstrip("#").strip()
                            false_reason = rest
                        else:
                            false_reason = false_arg
                    false_tag = None
                    replied = msg.get("reply_to_message", {})
                    replied_mid = replied.get("message_id")
                    target_aid_false = None
                    # ۱. جستجو با هشتگ (اگه داده شده)
                    if false_tag_search:
                        for aid_f, cid_map_f in _fired_msg_ids.items():
                            if cid_map_f.get("__tag__") == false_tag_search:
                                target_aid_false = aid_f
                                false_tag = cid_map_f.get("__tag__", "")
                                break
                        if not target_aid_false:
                            send_tg(token, cid, f"⚠️ آلارم <b>{false_tag_search}</b> پیدا نشد یا قبلاً پاک شده.\n\nشاید سرور restart شده — مستقیم روی پیام آلارم ریپلای بزن.")
                    # ۲. جستجو با ریپلای
                    elif replied_mid:
                        for aid_f, cid_map_f in _fired_msg_ids.items():
                            if str(cid_map_f.get(cid)) == str(replied_mid):
                                target_aid_false = aid_f
                                false_tag = cid_map_f.get("__tag__", "")
                                break
                    if target_aid_false:
                        sender_name_false = _get_user_custom_name(cid) or uname

                        # جلوگیری از race condition — اگه الان داره false میشه، ignore کن
                        with _false_in_progress_lock:
                            _fp_busy = target_aid_false in _false_in_progress
                            if not _fp_busy:
                                _false_in_progress.add(target_aid_false)

                        if _fp_busy:
                            send_tg(token, cid, "⏳ این آلارم الان داره پردازش میشه، چند ثانیه صبر کن.")
                        else:
                            try:
                                # چک کن این آلارم قبلاً false شده یا نه (sync — قبل از ارسال پیام)
                                already_false = False
                                if SUPABASE_KEY:
                                    try:
                                        r_chk = requests.get(
                                            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{target_aid_false}&select=is_active",
                                            headers=_sb_h(), timeout=8)
                                        if r_chk.status_code == 200:
                                            chk_rows = r_chk.json()
                                            if chk_rows:
                                                already_false = (chk_rows[0].get("is_active") == False)
                                    except Exception as e:
                                        print(f"[false] check exc: {e}")

                                def _false_then_rebuild(aid=target_aid_false, sender=sender_name_false, reason=false_reason):
                                    # اول false رو تو Supabase ذخیره کن، بعد شمارش رو بازسازی کن —
                                    # اگه این دوتا موازی اجرا بشن، ممکنه rebuild قبل از ثبت is_active=False
                                    # از Supabase بخونه و شمارش طرف false‌کننده درست کم نشه (race condition)
                                    _sb_false_assignment(aid, sender, reason)
                                    _rebuild_active_assign_count(_sb_load_active_assignments())
                                threading.Thread(target=_false_then_rebuild, daemon=True).start()
                                tag_txt = f" <b>{false_tag}</b>" if false_tag else ""
                                reason_line = f"\n📝 علت: {false_reason}" if false_reason else ""
                                now_label = now_pretty()

                                false_cid_map = _fired_msg_ids.get(target_aid_false, {})
                                prev_broadcast_map = _false_broadcast_ids.get(target_aid_false, {})
                                if not prev_broadcast_map and already_false and SUPABASE_KEY:
                                    try:
                                        r_bm = requests.get(
                                            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{target_aid_false}&select=false_broadcast_map",
                                            headers=_sb_h(), timeout=8)
                                        if r_bm.status_code == 200:
                                            bm_rows = r_bm.json()
                                            if bm_rows and bm_rows[0].get("false_broadcast_map"):
                                                prev_broadcast_map = bm_rows[0]["false_broadcast_map"]
                                                if isinstance(prev_broadcast_map, str):
                                                    prev_broadcast_map = json.loads(prev_broadcast_map)
                                    except Exception as e:
                                        print(f"[false] load broadcast_map exc: {e}")
                                new_broadcast_map = {}

                                if already_false and prev_broadcast_map:
                                    # ── آپدیت — ریپلای روی پیام False قبلی، متن قبلی حذف نمیشه
                                    update_msg = (
                                        f"🔄 <b>آپدیت</b> — {now_label}\n"
                                        f"👤 توسط: <b>{sender_name_false}</b>"
                                        f"{reason_line}"
                                    )
                                    for tc, tm in false_cid_map.items():
                                        if tc in ("__tag__", "__text__"): continue
                                        reply_target = prev_broadcast_map.get(tc, tm)
                                        try:
                                            r_send = requests.post(
                                                f"https://api.telegram.org/bot{token}/sendMessage",
                                                json={"chat_id": tc, "text": update_msg,
                                                      "parse_mode": "HTML", "reply_to_message_id": reply_target},
                                                timeout=8, headers=H)
                                            if r_send.status_code == 200:
                                                new_mid = r_send.json().get("result", {}).get("message_id")
                                                if new_mid:
                                                    new_broadcast_map[tc] = new_mid
                                        except: pass
                                else:
                                    # ── اولین بار false میشه
                                    false_broadcast = (
                                        f"❌ آلارم{tag_txt} از لیست تریگر خارج شد\n"
                                        f"👤 توسط: <b>{sender_name_false}</b>"
                                        f"{reason_line}"
                                    )
                                    for tc, tm in false_cid_map.items():
                                        if tc in ("__tag__", "__text__"): continue
                                        try:
                                            r_send = requests.post(
                                                f"https://api.telegram.org/bot{token}/sendMessage",
                                                json={"chat_id": tc, "text": false_broadcast,
                                                      "parse_mode": "HTML", "reply_to_message_id": tm},
                                                timeout=8, headers=H)
                                            if r_send.status_code == 200:
                                                new_mid = r_send.json().get("result", {}).get("message_id")
                                                if new_mid:
                                                    new_broadcast_map[tc] = new_mid
                                        except: pass

                                if new_broadcast_map:
                                    _false_broadcast_ids[target_aid_false] = new_broadcast_map
                                    threading.Thread(
                                        target=lambda aid=target_aid_false, m=new_broadcast_map: requests.patch(
                                            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}",
                                            headers={**_sb_h(), "Prefer": "return=minimal"},
                                            json={"false_broadcast_map": m}, timeout=8) if SUPABASE_KEY else None,
                                        daemon=True).start()
                            finally:
                                # همیشه lock رو آزاد کن
                                with _false_in_progress_lock:
                                    _false_in_progress.discard(target_aid_false)
                    else:
                        send_tg(token, cid, "⚠️ روی پیام آلارم ریپلای بزن و /False بنویس.\n\n💡 اگه آلارم قبل از restart سرور آمده، با /False XAUUSD5 (هشتگ آلارم) امتحان کن.")

                # ── /check — ثبت در ژورنال روی پیام ─────────────
                elif txt.strip().startswith("/check"):
                    check_parts = txt.strip().split(maxsplit=1)
                    check_note = check_parts[1].strip() if len(check_parts) > 1 else ""
                    replied = msg.get("reply_to_message", {})
                    replied_mid = replied.get("message_id")
                    replied_text = replied.get("text") or replied.get("caption") or ""
                    if not replied_mid:
                        send_tg(token, cid, "⚠️ باید روی پیام آلارم ریپلای بزنی و /check بنویسی.")
                    else:
                        target_aid = None
                        for aid, cid_map in _fired_msg_ids.items():
                            if str(cid_map.get(cid)) == str(replied_mid):
                                target_aid = aid
                                break
                        if not target_aid:
                            send_tg(token, cid, "⚠️ این پیام توی لیست آلارم‌های ذخیره‌شده نیست یا قبلاً ثبت شده.")
                        else:
                            cid_map = _fired_msg_ids.get(target_aid, {})
                            orig_text = cid_map.get("__text__") or replied_text
                            note_line = f"\n🗒 {check_note}" if check_note else ""
                            journal_line = f"\n──────────────\n📋 بررسی شد{note_line}\n🕐 {now_pretty()}"
                            new_text = orig_text + journal_line
                            edited_count = 0
                            for tc, tm in cid_map.items():
                                if tc in ("__tag__", "__text__"): continue
                                try:
                                    r_edit = requests.post(
                                        f"https://api.telegram.org/bot{token}/editMessageText",
                                        json={"chat_id": tc, "message_id": tm,
                                              "text": new_text, "parse_mode": "HTML"},
                                        timeout=8, headers=H)
                                    if r_edit.status_code == 200:
                                        edited_count += 1
                                except: pass
                            if edited_count:
                                send_tg(token, cid, f"✅ در <b>{edited_count}</b> چت ثبت شد.")
                            else:
                                send_tg(token, cid, "⚠️ نشد ویرایش کرد.")
                elif txt == "📩 درخواست فعال‌سازی آلارم":
                    if _is_approved(cid):
                        show_main_menu(token, cid, "✅ شما قبلاً تایید شدید و به همه امکانات دسترسی دارید.", cid == YOUR_CHAT_ID)
                    else:
                        send_tg(token, cid,
                            "📩 <b>درخواست شما ثبت شد</b>\n\n"
                            "درخواست فعال‌سازی شما برای ادمین ارسال شد و در دست بررسی است.\n"
                            "پس از تایید، یک پیام دریافت خواهید کرد. 🙏")
                        d_su = load_alerts()
                        su_user = next((u for u in d_su.get("users",[]) if str(u.get("chat_id","")) == cid), {})
                        su_name = su_user.get("custom_name","") or su_user.get("username","") or cid
                        admin_notif_su = (
                            f"📩 <b>درخواست فعال‌سازی عضو جدید</b>\n\n"
                            f"👤 نام: <b>{su_name}</b>\n"
                            f"🆔 Chat ID: <code>{cid}</code>\n"
                            f"⏰ {now_pretty()} (تهران)"
                        )
                        approve_kb_su = [
                            [{"text": "✅ تایید و فعال‌سازی", "callback_data": f"approve_signup:{cid}"}],
                            [{"text": "❌ رد درخواست",       "callback_data": f"reject_signup:{cid}"}],
                        ]
                        send_tg_keyboard(token, YOUR_CHAT_ID, admin_notif_su, approve_kb_su)

                elif txt in ("📊 وضعیت",) or (txt.startswith("/status") and txt not in ("/statuspage",)):
                    d2 = load_alerts()
                    all_active2 = [a for a in d2.get("alerts",[]) if a.get("active")]
                    team_active2 = [a for a in all_active2 if not a.get("is_private")]
                    private_active2 = [a for a in all_active2 if a.get("is_private")]
                    my_rem = _reminders.get(cid, {})
                    is_open = is_forex_market_open()
                    is_adm = (cid == YOUR_CHAT_ID)
                    has_priv = _has_private_access(cid)
                    status_text = (
                        f"📊 <b>وضعیت سیستم</b>\n\n"
                        f"{'🟢' if is_open else '🔴'} فارکس: {'باز' if is_open else 'بسته'}\n"
                        f"📈 آلارم فعال (کل): <b>{len(all_active2)}</b>\n"
                        f"🌐 تیمی: <b>{len(team_active2)}</b> | 🔒 شخصی: <b>{len(private_active2)}</b>\n"
                        f"⏰ هشدار دوره‌ای من: <b>{len(my_rem)}</b>\n"
                        f"🔒 آلارم شخصی: {'✅ فعال' if has_priv else '❌ غیرفعال'}\n"
                        f"⏱ {now_pretty()} (تهران)"
                    )
                    status_kb = []
                    if not has_priv:
                        status_kb.append([{"text": "📩 درخواست فعال‌سازی آلارم شخصی", "callback_data": f"req_private:{cid}"}])
                    status_kb.append([{"text": "✏️ ویرایش اسم", "callback_data": f"edit_name:{cid}"}])
                    status_kb.append([{"text": "📡 سیگنال‌های من", "callback_data": f"signals_view:{cid}:mine"},
                                      {"text": "📊 همه سیگنال‌ها", "callback_data": f"signals_view:{cid}:all"}])
                    status_kb.append([{"text": "🎯 لیست تریگر", "callback_data": f"trigger_list:{cid}"}])
                    status_kb.append([{"text": "📊 آنالیز هفتگی", "callback_data": f"weekly_menu:{cid}"}])
                    status_kb.append([{"text": "🔔 نمایش آلارم‌های فعال", "callback_data": f"resend_active:{cid}"}])
                    send_tg_keyboard(token, cid, status_text, status_kb)

                elif txt == "⭐ آلارم‌های من":
                    is_adm = (cid == YOUR_CHAT_ID)
                    has_priv = _has_private_access(cid)
                    btns = [{"text": "🌐 آلارم‌های تیمی", "callback_data": f"myalerts:pub:{cid}"}]
                    if has_priv:
                        btns.append({"text": "🔒 آلارم‌های شخصی", "callback_data": f"myalerts:priv:{cid}"})
                    btns = [[b] if isinstance(b, dict) else b for b in btns]
                    btns.append([{"text": "📋 همه آلارم‌های من", "callback_data": f"myalerts:all:{cid}"}])
                    btns.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
                    send_tg_keyboard(token, cid, "⭐ <b>آلارم‌های من</b>\n\nکدوم رو می‌خوای ببینی؟", btns)

                elif txt == "⏰ هشدار دوره‌ای من":
                    text_msg2, kb2 = build_cancel_reminder_msg(cid)
                    is_adm = (cid == YOUR_CHAT_ID)
                    # دکمه ثبت هشدار جدید همیشه نشون داده میشه
                    kb2.append([{"text": "➕ هشدار جدید", "callback_data": f"reminder_new:{cid}"}])
                    send_tg_keyboard(token, cid, text_msg2 if kb2 else "هیچ هشدار فعالی نداری.\n\nمیخوای هشدار جدید بذاری؟", kb2)

                elif txt == "📡 سیگنال جدید":
                    # مرحله ۱ — نماد
                    quick_btns = [[{"text": s, "callback_data": f"sig_sym:{cid}:{s}"} for s in SIGNAL_QUICK_SYMBOLS[:3]],
                                  [{"text": s, "callback_data": f"sig_sym:{cid}:{s}"} for s in SIGNAL_QUICK_SYMBOLS[3:]],
                                  [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]]
                    mid_sig = send_tg_keyboard(token, cid,
                        "📡 <b>سیگنال جدید</b>\n\nنماد رو انتخاب کن یا بنویس:",
                        quick_btns)
                    _pending_signal[cid] = {"step": "sig_symbol", "data": {}, "bot_msg_id": mid_sig}

                elif txt == "⚙️ پنل ادمین" and cid == YOUR_CHAT_ID:
                    admin_kb = [
                        [{"text": "📰 اخبار فارکس",       "callback_data": "admin:news"}],
                        [{"text": "✉️ پیام به گروه",      "callback_data": "admin:broadcast"}],
                        [{"text": "👥 لیست کاربران",       "callback_data": "admin:users"}],
                        [{"text": "🗑 مدیریت سیگنال‌ها",  "callback_data": "admin_sig:list:1"}],
                        [{"text": "📋 مسئولین آلارم",         "callback_data": "admin:shift:1"}],
                        [{"text": "❌ فالس همه آلارم‌های فعال", "callback_data": "admin:bulkfalse:0"}],
                        [{"text": "✕ بستن",                "callback_data": "close_myalerts"}],
                    ]
                    send_tg_keyboard(token, cid,
                        "⚙️ <b>پنل ادمین</b>\n\nیه گزینه انتخاب کن:", admin_kb)


                elif txt == "❌ انصراف":
                    pend_cancel = _pending_alarm.pop(cid, None)
                    if pend_cancel and pend_cancel.get("bot_msg_id"):
                        edit_tg_keyboard(token, cid, pend_cancel["bot_msg_id"],
                            "❌ <b>عملیات لغو شد.</b>", [])

                elif txt == "📈 آلارم جدید" and (cid == YOUR_CHAT_ID or (BROADCAST_MODE and _is_approved(cid))):
                    kb_new = [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]]
                    mid_new = send_tg_keyboard(token, cid,
                        "🔔 <b>آلارم جدید</b>\n\nاسم نماد رو بنویس:\n<code>EURUSD</code>  <code>XAUUSD</code>  <code>BTC</code>",
                        kb_new)
                    _pending_alarm[cid] = {"step": "alarm_symbol", "data": {"ptype": "public"}, "bot_msg_id": mid_new}

                elif txt == "🔒 آلارم شخصی":
                    if not _has_private_access(cid):
                        is_adm = (cid == YOUR_CHAT_ID)
                        show_main_menu(token, cid, "⚠️ این قابلیت برای شما فعال نیست.\nاز بخش 📊 وضعیت می‌توانید درخواست دسترسی بدهید.", is_adm)
                    else:
                        kb_priv = [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]]
                        mid_priv = send_tg_keyboard(token, cid,
                            "🔒 <b>آلارم شخصی</b>\n\nاسم نماد رو بنویس:\n<code>EURUSD</code>  <code>XAUUSD</code>  <code>BTC</code>",
                            kb_priv)
                        _pending_alarm[cid] = {"step": "alarm_symbol", "data": {"ptype": "private"}, "bot_msg_id": mid_priv}


                elif txt == "⚡ آلارم فوری" and (cid == YOUR_CHAT_ID or (BROADCAST_MODE and _is_approved(cid))):
                    kb_sos = [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]]
                    mid_sos = send_tg_keyboard(token, cid,
                        "⚡ <b>آلارم فوری</b>\n\nاسم نماد رو بنویس:\n<code>EURUSD</code>  <code>XAUUSD</code>  <code>BTC</code>",
                        kb_sos)
                    _pending_alarm[cid] = {"step": "sos_symbol", "data": {}, "bot_msg_id": mid_sos}

                elif cid in _pending_bulk_false and not txt.startswith("/") and cid == YOUR_CHAT_ID:
                    bf_info = _pending_bulk_false.pop(cid)
                    bf_mid = bf_info.get("bot_msg_id")
                    bf_reason_txt = txt.strip()
                    send_tg(token, cid, f"⏳ در حال false کردن {bf_info.get('count',0)} آلارم...")

                    def _do_bulk_false_txt(reason=bf_reason_txt, tok=token, c=cid, mid=bf_mid):
                        rows_bft = _sb_load_active_assignments()
                        admin_name_bft = _get_user_custom_name(c) or "ادمین"
                        done_count_t = 0
                        for row_bft in rows_bft:
                            aid_bft = row_bft.get("id")
                            tag_bft = row_bft.get("alarm_tag", "")
                            _sb_false_assignment(aid_bft, admin_name_bft, reason)
                            done_count_t += 1
                            msg_map_bft = _fired_msg_ids.get(aid_bft, {})
                            reason_line_bft = f"\n📝 علت: {reason}" if reason else ""
                            reply_bft = (f"❌ <b>فالس (دسته‌جمعی توسط ادمین)</b>\n\n"
                                         f"{tag_bft}{reason_line_bft}")
                            for tc_bft, tm_bft in msg_map_bft.items():
                                if tc_bft in ("__tag__", "__text__"): continue
                                try:
                                    requests.post(
                                        f"https://api.telegram.org/bot{tok}/sendMessage",
                                        json={"chat_id": tc_bft, "text": reply_bft,
                                              "parse_mode": "HTML", "reply_to_message_id": tm_bft},
                                        timeout=8, headers=H)
                                except: pass
                        _rebuild_active_assign_count(_sb_load_active_assignments())
                        edit_tg_keyboard(tok, c, mid,
                            f"✅ <b>{done_count_t} آلارم false شدن</b>\n📝 علت: {reason}",
                            [[{"text": "↩️ پنل ادمین", "callback_data": "admin_sig:back"}]])
                    threading.Thread(target=_do_bulk_false_txt, daemon=True).start()

                elif cid in _pending_reminder and not txt.startswith("/"):
                    pr_step = _pending_reminder[cid].get("step")
                    pr_mid  = _pending_reminder[cid].get("bot_msg_id")
                    if pr_step == "rem_symbol":
                        r_sym = txt.upper().replace("/","").strip()
                        if len(r_sym) < 2:
                            edit_tg_keyboard(token, cid, pr_mid,
                                "❌ نماد نامعتبر. دوباره بنویس:", [[{"text":"❌ انصراف","callback_data":"close_myalerts"}]])
                        else:
                            del _pending_reminder[cid]
                            kb_tf = [
                                [{"text": "🕯 M5  (۱ دق قبل کلوز)",  "callback_data": f"reminder_go:{cid}:{r_sym}:300"}],
                                [{"text": "🕯 M15 (۵ دق قبل کلوز)",  "callback_data": f"reminder_go:{cid}:{r_sym}:900"}],
                                [{"text": "🕯 H1  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{cid}:{r_sym}:3600"}],
                                [{"text": "🕯 H4  (۱۵ دق قبل کلوز)", "callback_data": f"reminder_go:{cid}:{r_sym}:14400"}],
                                [{"text": "✕ انصراف", "callback_data": "close_myalerts"}],
                            ]
                            edit_tg_keyboard(token, cid, pr_mid,
                                f"🕯 تایم‌فریم هشدار برای <b>{r_sym}</b>:", kb_tf)

                elif cid in _pending_alarm and not txt.startswith("/"):
                    step = _pending_alarm[cid]["step"]
                    dw   = _pending_alarm[cid]["data"]
                    bot_msg_id = _pending_alarm[cid].get("bot_msg_id")
                    is_adm = (cid == YOUR_CHAT_ID)

                    # ── لغو در هر مرحله ─────────────────────────────────
                    if txt in ("↩️ برگشت", "❌ انصراف"):
                        del _pending_alarm[cid]
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id, "❌ <b>عملیات لغو شد.</b>", [])
                        # پیام کاربر رو ادیت نکن، فقط state رو پاک کردیم

                    elif step == "alarm_symbol":
                        sym_w = txt.upper().replace("/","")
                        if len(sym_w) < 2:
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    "🔔 <b>آلارم جدید</b>\n\n❌ نماد نامعتبر.\nاسم نماد رو بنویس:\n<code>EURUSD</code>  <code>XAUUSD</code>  <code>BTC</code>",
                                    [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]])
                        else:
                            dw["symbol"] = sym_w
                            dw["atype"]  = "forex" if any(x in sym_w for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                            _pending_alarm[cid]["step"] = "alarm_dir"
                            ptype_lbl = "🔒 شخصی" if dw.get("ptype") == "private" else "🌐 تیمی"
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    f"🔔 <b>{sym_w}</b>  ({ptype_lbl})\n\nجهت معامله رو انتخاب کن:",
                                    [
                                        [{"text": "📈 BUY", "callback_data": f"alarm_dir:{cid}:buy"},
                                         {"text": "📉 SELL", "callback_data": f"alarm_dir:{cid}:sell"}],
                                        [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                    ])

                    elif step == "alarm_dir":
                        # این step از inline callback میاد (alarm_dir:cid:buy|sell)
                        # اگه کاربر text فرستاد، remind کن
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id,
                                f"🔔 <b>{dw.get('symbol','')}</b>\n\nلطفاً از دکمه‌های زیر انتخاب کن:",
                                [
                                    [{"text": "📈 BUY", "callback_data": f"alarm_dir:{cid}:buy"},
                                     {"text": "📉 SELL", "callback_data": f"alarm_dir:{cid}:sell"}],
                                    [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                ])

                    elif step == "alarm_price":
                        try:
                            dw["target_price"] = float(txt.replace(",",""))
                            _pending_alarm[cid]["step"] = "alarm_comment"
                            dir_lbl2 = "📈 BUY" if dw["condition"] == "below" else "📉 SELL"
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    f"🔔 <b>{dw['symbol']}</b>  {dir_lbl2}  @  <code>{fmt_price(dw['target_price'], dw['symbol'])}</code>\n\n"
                                    f"یادداشت بنویس یا بدون یادداشت ثبت کن:",
                                    [
                                        [{"text": "✅ ثبت بدون یادداشت", "callback_data": f"alarm_submit:{cid}"}],
                                        [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                    ])
                        except ValueError:
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    f"🔔 <b>{dw.get('symbol','')}</b>\n\n❌ عدد نامعتبر. مثال: <code>1.08500</code> یا <code>2350</code>\n\nدوباره قیمت هدف رو بنویس:",
                                    [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]])

                    elif step == "alarm_comment":
                        comment_w = "" if txt in ("✅ ثبت بدون یادداشت", "✅ ثبت") else txt
                        dw["comment"] = comment_w
                        _pending_alarm[cid]["step"] = "alarm_expiry"
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id,
                                f"🔔 <b>{dw['symbol']}</b>\n\n⏳ این آلارم چند روز معتبر باشه؟ (اگه تا اون موقع فایر نشه، خودکار حذف می‌شه)",
                                [
                                    [{"text": "۱ روز", "callback_data": f"alarm_expiry:{cid}:1"},
                                     {"text": "۳ روز", "callback_data": f"alarm_expiry:{cid}:3"},
                                     {"text": "۷ روز", "callback_data": f"alarm_expiry:{cid}:7"}],
                                    [{"text": "♾ بدون انقضا", "callback_data": f"alarm_expiry:{cid}:0"}],
                                    [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                ])

                    elif step == "alarm_expiry":
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id,
                                f"🔔 <b>{dw.get('symbol','')}</b>\n\nلطفاً از دکمه‌های زیر مدت اعتبار رو انتخاب کن:",
                                [
                                    [{"text": "۱ روز", "callback_data": f"alarm_expiry:{cid}:1"},
                                     {"text": "۳ روز", "callback_data": f"alarm_expiry:{cid}:3"},
                                     {"text": "۷ روز", "callback_data": f"alarm_expiry:{cid}:7"}],
                                    [{"text": "♾ بدون انقضا", "callback_data": f"alarm_expiry:{cid}:0"}],
                                    [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                ])

                    elif step == "edit_name_input":
                        new_name = txt.strip()
                        if len(new_name) < 2:
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    "✏️ <b>ویرایش اسم</b>\n\n❌ اسم باید حداقل ۲ حرف باشه.\nدوباره بنویس:",
                                    [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]])
                        else:
                            d_en2 = load_alerts()
                            found_en = False
                            for u in d_en2.get("users", []):
                                if str(u.get("chat_id","")) == cid:
                                    u["custom_name"] = new_name
                                    found_en = True
                                    break
                            if not found_en:
                                d_en2.setdefault("users",[]).append({
                                    "chat_id": cid, "username": uname,
                                    "joined_at": now_teh(), "custom_name": new_name
                                })
                            save_alerts(d_en2)
                            del _pending_alarm[cid]
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    f"✅ <b>اسم با موفقیت ذخیره شد!</b>\n\nاسم جدید: <b>{new_name}</b>", [])

                    elif step == "sos_symbol":
                        sym_w2 = txt.upper().replace("/","")
                        if len(sym_w2) < 2:
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    "⚡ <b>آلارم فوری</b>\n\n❌ نماد نامعتبر.\nمثال: <code>EURUSD</code>  <code>XAUUSD</code>  <code>BTC</code>",
                                    [[{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]])
                        else:
                            dw["symbol"] = sym_w2
                            _pending_alarm[cid]["step"] = "sos_dir"
                            if bot_msg_id:
                                edit_tg_keyboard(token, cid, bot_msg_id,
                                    f"⚡ <b>آلارم فوری</b>  ─  <b>{sym_w2}</b>\n\nجهت معامله رو انتخاب کن:",
                                    [
                                        [{"text": "📈 BUY", "callback_data": f"sos_dir:{cid}:buy"},
                                         {"text": "📉 SELL", "callback_data": f"sos_dir:{cid}:sell"}],
                                        [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                    ])


                    elif step == "sos_dir":
                        # این step دیگه از text نمیاد — از callback میاد (sos_dir:cid:buy/sell)
                        # اگه کاربر text فرستاد، remind کن که از دکمه استفاده کنه
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id,
                                f"⚡ <b>آلارم فوری</b>  ─  <b>{dw.get('symbol','')}</b>\n\nلطفاً از دکمه‌های زیر انتخاب کن:",
                                [
                                    [{"text": "📈 BUY", "callback_data": f"sos_dir:{cid}:buy"},
                                     {"text": "📉 SELL", "callback_data": f"sos_dir:{cid}:sell"}],
                                    [{"text": "❌ انصراف", "callback_data": f"flow_cancel:{cid}"}]
                                ])

                    elif step == "sos_comment":
                        comment_s = "" if txt in ("✅ ارسال بدون یادداشت", "ارسال") else txt
                        sym_s = dw["symbol"]
                        condition_s = dw["condition"]
                        atype_s = "forex" if any(x in sym_s for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                        sender_s = _get_user_custom_name(cid) or uname
                        alarm_num_tag_s = _make_alarm_tag(sym_s)
                        sender_tag_s = "#" + re.sub(r"[^\w]","_", sender_s).strip("_")
                        arrow_s = "📈 ناحیه سل" if condition_s == "above" else "📉 ناحیه بای"
                        dir_lbl_s = dw.get("dir_lbl", "📈 BUY" if condition_s == "below" else "📉 SELL")
                        try: cur_s = get_price(sym_s, atype_s)
                        except: cur_s = None
                        out_s = (
                            f"🚨 <b>آلارم فوری!</b>\n"
                            f"━━━━━━━━━━━━━━━━━━\n"
                            f"💰 <b>#{sym_s}</b>  {arrow_s}\n"
                            f"🔖 {alarm_num_tag_s}\n"
                            f"👤 {sender_tag_s}\n"
                            f"📊 قیمت: <b>{fmt_price(cur_s, sym_s) if cur_s else '—'}</b>\n"
                            + (f"💬 {comment_s}\n" if comment_s else "")
                            + f"━━━━━━━━━━━━━━━━━━\n"
                            f"⏰ {now_pretty()} (تهران)"
                        )
                        _, all_cids2, _ = _get_token_and_cids()
                        targets2 = all_cids2 if BROADCAST_MODE else [YOUR_CHAT_ID]
                        del _pending_alarm[cid]
                        # ادیت پیام اصلی به تأییدیه
                        if bot_msg_id:
                            edit_tg_keyboard(token, cid, bot_msg_id,
                                f"✅ <b>آلارم فوری ارسال شد!</b>\n\n"
                                f"💰 <b>{sym_s}</b>  {dir_lbl_s}\n"
                                + (f"💬 {comment_s}\n" if comment_s else "")
                                + f"⏰ {now_pretty()} (تهران)", [])
                        # broadcast به بقیه
                        if sym_s.upper() not in TEMP_MUTED_SYMBOLS:
                            for tc2 in targets2:
                                kb2 = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc2}:{sym_s}"}]]
                                send_tg_keyboard(token, str(tc2), out_s, kb2, track=False)

                        # ذخیره توی آرشیو (این مسیر قبلاً اصلاً ذخیره نمی‌شد)
                        sos_arch_entry_s = {"id": str(int(time.time()*1000)), "symbol": sym_s, "type": atype_s,
                            "condition": condition_s, "comment": comment_s, "created_by": sender_s,
                            "active": False, "fired_at": now_teh(), "fired_price": cur_s,
                            "instant": True, "created_at": now_teh(), "tag": alarm_num_tag_s}
                        d_arc_s = load_alerts()
                        d_arc_s.setdefault("archive", []).append(sos_arch_entry_s)
                        save_alerts(d_arc_s)
                        threading.Thread(target=_sb_upsert_alert, args=(sos_arch_entry_s,), daemon=True).start()

                    elif step == "broadcast_text":
                        _, all_cids3, _ = _get_token_and_cids()
                        ok3 = sum(1 for tc3 in all_cids3 if send_tg(token, tc3, txt))
                        del _pending_alarm[cid]
                        show_main_menu(token, cid, f"✅ پیام به {ok3} نفر ارسال شد.", is_adm)

                # ── signal pending text steps ─────────────────────────
                elif cid in _pending_signal and not txt.startswith("/"):
                    ps = _pending_signal[cid]
                    ps_step = ps["step"]
                    ps_data = ps["data"]
                    ps_mid  = ps.get("bot_msg_id")

                    if txt in ("❌ انصراف",):
                        del _pending_signal[cid]
                        if ps_mid:
                            edit_tg_keyboard(token, cid, ps_mid, "❌ <b>ساخت سیگنال لغو شد.</b>", [])

                    elif ps_step == "sig_symbol":
                        # کاربر نماد تایپ کرد
                        sym_s = txt.upper().replace("/","").strip()
                        if len(sym_s) < 2:
                            edit_tg_keyboard(token, cid, ps_mid,
                                "📡 <b>سیگنال جدید</b>\n\n❌ نماد نامعتبر. دوباره بنویس:",
                                [[{"text": s, "callback_data": f"sig_sym:{cid}:{s}"} for s in SIGNAL_QUICK_SYMBOLS[:3]],
                                 [{"text": s, "callback_data": f"sig_sym:{cid}:{s}"} for s in SIGNAL_QUICK_SYMBOLS[3:]],
                                 [{"text": "❌ انصراف", "callback_data": f"sig_cancel:{cid}"}]])
                        else:
                            ps_data["symbol"] = sym_s
                            ps["step"] = "sig_direction"
                            dir_kb = [[{"text": lbl, "callback_data": f"sig_dir:{cid}:{val}"} for lbl,val in SIGNAL_DIRECTIONS[:2]],
                                      [{"text": lbl, "callback_data": f"sig_dir:{cid}:{val}"} for lbl,val in SIGNAL_DIRECTIONS[2:]],
                                      [{"text": "❌ انصراف", "callback_data": f"sig_cancel:{cid}"}]]
                            edit_tg_keyboard(token, cid, ps_mid,
                                f"📡 <b>{sym_s}</b>\n\nنوع سفارش:", dir_kb)

                    elif ps_step == "sig_entry_sl":
                        # کاربر Entry + SL رو نوشته: "73370 72550"
                        parts_es = txt.strip().split()
                        if len(parts_es) < 2:
                            edit_tg_keyboard(token, cid, ps_mid,
                                f"📡 <b>{ps_data.get('symbol')}</b>  {ps_data.get('dir_lbl','')}\n\n"
                                "❌ دو عدد بنویس: <code>Entry  SL</code>\nمثال: <code>73370 72550</code>",
                                [[{"text": "❌ انصراف", "callback_data": f"sig_cancel:{cid}"}]])
                            return
                        try:
                            entry_v = float(parts_es[0].replace(",",""))
                            sl_raw  = float(parts_es[1].replace(",",""))
                        except:
                            edit_tg_keyboard(token, cid, ps_mid,
                                f"📡 <b>{ps_data.get('symbol')}</b>\n\n❌ عدد نامعتبر. دوباره بنویس:",
                                [[{"text": "❌ انصراف", "callback_data": f"sig_cancel:{cid}"}]])
                            return
                        # تشخیص پیپ vs قیمت — از inline button قبلاً تعیین شده
                        sl_mode = ps_data.get("sl_mode", "price")
                        sym_v   = ps_data["symbol"]
                        direction_v = ps_data["direction"]
                        if sl_mode == "pip":
                            sl_v = _sl_from_pips(sym_v, direction_v, entry_v, sl_raw)
                        else:
                            sl_v = sl_raw
                        sl_final, tp1, risk_pips = _calc_signal(sym_v, direction_v, entry_v, sl_v, SIGNAL_DEFAULT_RR)
                        ps_data.update({"entry": entry_v, "sl": sl_final, "tp1": tp1,
                                        "tp2": None, "tp3": None, "risk_pips": risk_pips,
                                        "tf": SIGNAL_DEFAULT_TF, "rr": SIGNAL_DEFAULT_RR})
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token, cid, ps_mid, ps_data)

                    elif ps_step == "sig_tp_edit":
                        # کاربر TP2 یا TP3 تایپ کرد
                        which = ps_data.get("_editing_tp", "tp2")
                        try:
                            val = float(txt.replace(",",""))
                            ps_data[which] = val
                        except:
                            pass
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token, cid, ps_mid, ps_data)

                    elif ps_step == "sig_note":
                        ps_data["note"] = txt.strip()
                        ps["step"] = "sig_preview"
                        _show_signal_preview(token, cid, ps_mid, ps_data)

                # ── /sos ─────────────────────────────────────────────
                elif txt.startswith("/sos") and (cid == YOUR_CHAT_ID or (BROADCAST_MODE and _is_approved(cid))):
                    parts = txt.split(maxsplit=3)
                    if len(parts) < 2:
                        send_tg(token, cid,
                            "⚠️ فرمت:\n<code>/sos SYMBOL [buy|sell] [کامنت]</code>\n"
                            "مثال:\n<code>/sos GBPUSD sell</code>")
                    else:
                        sym = parts[1].upper().replace("/", "")
                        raw_dir = parts[2].lower() if len(parts) > 2 else "sell"
                        comment = parts[3] if len(parts) > 3 else ""
                        condition = "above" if raw_dir in ("sell","s","سل","above") else "below"
                        atype = "forex" if any(x in sym for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                        sender_name = _get_sender_name(msg)
                        cur = None
                        try: cur = get_price(sym, atype)
                        except: pass
                        arrow = "📈 ناحیه سل" if condition == "above" else "📉 ناحیه بای"
                        cmt = f"\n💬 <i>{comment}</i>" if comment else ""
                        price_text = fmt_price(cur, sym) if cur else "—"
                        alarm_num_tag = _make_alarm_tag(sym)
                        sender_hashtag = "#" + re.sub(r'[^\w]', '_', sender_name).strip('_')
                        out_msg = (
                            f"🚨 <b>آلارم فوری!</b>\n\n"
                            f"💰 <b>#{sym}</b> — {arrow}\n"
                            f"🔖 {alarm_num_tag}\n"
                            f"👤 {sender_hashtag}\n\n"
                            f"📊 قیمت لحظه‌ای: <b>{price_text}</b>"
                            f"{cmt}\n\n⏰ {now_pretty()} (تهران)"
                        )
                        _, all_cids, _ = _get_token_and_cids()
                        targets = all_cids if BROADCAST_MODE else [YOUR_CHAT_ID]
                        sos_aid_txt = f"sos_{sym}_{int(time.time())}"
                        sos_cid_to_mid_txt = {}
                        if sym.upper() not in TEMP_MUTED_SYMBOLS:
                            for tc in targets:
                                mid_sos_txt = send_tg_keyboard(token, tc, out_msg,
                                    [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc}:{sym}"}]],
                                    track=False)
                                if mid_sos_txt:
                                    sos_cid_to_mid_txt[str(tc)] = mid_sos_txt
                        if sos_cid_to_mid_txt:
                            sos_cid_to_mid_txt["__tag__"] = alarm_num_tag
                            sos_cid_to_mid_txt["__text__"] = out_msg
                            _fired_msg_ids[sos_aid_txt] = sos_cid_to_mid_txt
                            threading.Thread(target=_sb_save_fired_msgs, args=(sos_aid_txt, sos_cid_to_mid_txt), daemon=True).start()
                        d = load_alerts()
                        arch = d.get("archive", [])
                        new_sos_entry = {"id": str(int(time.time()*1000)), "symbol": sym, "type": atype,
                            "condition": condition, "comment": comment, "created_by": sender_name,
                            "active": False, "fired_at": now_teh(), "fired_price": cur,
                            "instant": True, "created_at": now_teh(), "tag": alarm_num_tag}
                        arch.append(new_sos_entry)
                        d["archive"] = arch
                        save_alerts(d)
                        threading.Thread(target=_sb_upsert_alert, args=(new_sos_entry,), daemon=True).start()

                # ── /alarm ───────────────────────────────────────────
                elif txt.startswith("/alarm") and (cid == YOUR_CHAT_ID or (BROADCAST_MODE and _is_approved(cid))):
                    parts = txt.split(maxsplit=4)
                    if len(parts) < 4:
                        send_tg(token, cid,
                            "⚠️ فرمت:\n<code>/alarm SYMBOL buy|sell PRICE [کامنت]</code>\n\n"
                            "مثال‌ها:\n"
                            "<code>/alarm eurusd sell 1.12345 ناحیه سل</code>\n"
                            "<code>/alarm xauusd sell 2350 مقاومت مهم</code>")
                    else:
                        sym = parts[1].upper().replace("/", "")
                        raw_dir = parts[2].lower()
                        raw_price = parts[3]
                        comment = parts[4] if len(parts) > 4 else ""
                        condition = "above" if raw_dir in ("sell","s","سل","above") else "below"
                        atype = "forex" if any(x in sym for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                        sender_name = _get_sender_name(msg)
                        tgt_f = None
                        try:
                            tgt_f = float(raw_price)
                        except ValueError:
                            send_tg(token, cid, f"❌ قیمت نامعتبر: <code>{raw_price}</code>")
                        if tgt_f is not None:
                            dir_arrow = "📈 BUY" if condition == "below" else "📉 SELL"
                            new_alert = {
                                "id": str(int(time.time()*1000)),
                                "symbol": sym, "type": atype,
                                "target_price": tgt_f, "condition": condition,
                                "comment": comment, "created_by": sender_name,
                                "active": True, "last_price": None,
                                "last_checked": None,
                                "created_at": now_teh(),
                                "notify_only": YOUR_CHAT_ID if not BROADCAST_MODE else None
                            }
                            d = load_alerts()
                            d["alerts"].append(new_alert)
                            _sb_upsert_alert(new_alert)
                            _cache_alerts = d
                            _is_adm_alarm = (cid == YOUR_CHAT_ID)
                            confirm_alarm_txt = (
                                f"✅ آلارم ثبت شد\n\n"
                                f"<b>{sym}</b>  {dir_arrow}  @  <code>{fmt_price(tgt_f, sym)}</code>"
                                + (f"\n💬 {comment}" if comment else "")
                            )
                            show_main_menu(token, cid, confirm_alarm_txt, _is_adm_alarm)
                            def _bg_price(alert=new_alert, s=sym, t=atype, tok=token, c=cid):
                                try:
                                    cur = get_price(s, t)
                                    if cur:
                                        alert["last_price"] = cur
                                        alert["last_checked"] = now_teh()
                                        _sb_upsert_alert(alert)
                                except: pass
                            threading.Thread(target=_bg_price, daemon=True).start()

                # ── /mealarm — آلارم شخصی (فعلاً فقط ادمین) ───
                elif txt.startswith("/mealarm"):
                    if cid != YOUR_CHAT_ID:
                        send_tg(token, cid, "⚠️ این قابلیت فعلاً در دسترس نیست.")
                    else:
                        parts = txt.split(maxsplit=4)
                        if len(parts) < 4:
                            send_tg(token, cid,
                                "⚠️ فرمت:\n<code>/mealarm SYMBOL buy|sell PRICE [کامنت]</code>\n\n"
                                "مثال:\n"
                                "<code>/mealarm xauusd sell 2350 ناحیه شخصی</code>\n\n"
                                "این آلارم فقط برای شما ثبت میشه و بقیه نمیبینن.")
                        else:
                            sym = parts[1].upper().replace("/", "")
                            raw_dir = parts[2].lower()
                            raw_price = parts[3]
                            comment = parts[4] if len(parts) > 4 else ""
                            condition = "above" if raw_dir in ("sell","s","سل","above") else "below"
                            atype = "forex" if any(x in sym for x in ["EUR","GBP","JPY","XAU","XAG","CHF","CAD","AUD","NZD"]) else "crypto"
                            sender_name = _get_sender_name(msg)
                            tgt_f = None
                            try:
                                tgt_f = float(raw_price)
                            except ValueError:
                                send_tg(token, cid, f"❌ قیمت نامعتبر: <code>{raw_price}</code>")
                            if tgt_f is not None:
                                arrow = "سل 📈" if condition == "above" else "بای 📉"
                                new_alert = {
                                    "id": str(int(time.time()*1000)),
                                    "symbol": sym, "type": atype,
                                    "target_price": tgt_f, "condition": condition,
                                    "comment": comment, "created_by": sender_name,
                                    "active": True, "last_price": None,
                                    "last_checked": None,
                                    "created_at": now_teh(),
                                    "notify_only": cid,
                                    "private_cid": cid,
                                    "is_private": True
                                }
                                d = load_alerts()
                                d["alerts"].append(new_alert)
                                _sb_upsert_alert(new_alert)
                                _cache_alerts = d
                                _is_adm_me = (cid == YOUR_CHAT_ID)
                                _me_dir = "📈 BUY" if condition == "below" else "📉 SELL"
                                _me_confirm = (
                                    f"✅ آلارم شخصی ثبت شد 🔒\n\n"
                                    f"<b>{sym}</b>  {_me_dir}  @  <code>{fmt_price(tgt_f, sym)}</code>"
                                    + (f"\n💬 {comment}" if comment else "")
                                )
                                show_main_menu(token, cid, _me_confirm, _is_adm_me)
                                def _bg_price_me(alert=new_alert, s=sym, t=atype):
                                    try:
                                        cur = get_price(s, t)
                                        if cur:
                                            alert["last_price"] = cur
                                            alert["last_checked"] = now_teh()
                                            _sb_upsert_alert(alert)
                                    except: pass
                                threading.Thread(target=_bg_price_me, daemon=True).start()
                # ── /news ────────────────────────────────────────────
                elif txt.startswith("/cancel_reminder"):
                    text_msg, keyboard = build_cancel_reminder_msg(cid)
                    if keyboard:
                        send_tg_keyboard(token, cid, text_msg, keyboard)
                    else:
                        send_tg(token, cid, text_msg)

                elif txt.startswith("/myalerts"):
                    btns2 = [[{"text": "🌐 آلارم\u200cهای تیمی", "callback_data": f"myalerts:pub:{cid}"}]]
                    if _has_private_access(cid):
                        btns2.append([{"text": "🔒 آلارم\u200cهای شخصی", "callback_data": f"myalerts:priv:{cid}"}])
                    btns2.append([{"text": "📋 همه آلارم\u200cهای من", "callback_data": f"myalerts:all:{cid}"}])
                    btns2.append([{"text": "✕ بستن", "callback_data": "close_myalerts"}])
                    send_tg_keyboard(token, cid, "⭐ <b>آلارم\u200cهای من</b>\n\nکدوم رو می\u200cخوای ببینی؟", btns2)

                elif txt.startswith("/news") and cid == YOUR_CHAT_ID:
                    send_tg(token, cid, "⏳ در حال دریافت تقویم اقتصادی...")
                    events, err = fetch_ff_news()
                    if err and not events:
                        send_tg(token, cid, err)
                    else:
                        msg = format_ff_message(events) if events else "📭 امروز رویداد مهم فارکس نداریم."
                        send_tg(token, cid, msg)

                # ── /text ────────────────────────────────────────────
                elif txt.startswith("/text") and cid == YOUR_CHAT_ID:
                    body_text = txt[5:].strip()
                    if not body_text:
                        send_tg(token, cid, "\u26a0\ufe0f \u0641\u0631\u0645\u062a:\n<code>/text \u0645\u062a\u0646 \u067e\u06cc\u0627\u0645\u062a \u0627\u06cc\u0646\u062c\u0627</code>")
                    else:
                        _, all_cids, _ = _get_token_and_cids()
                        if not all_cids:
                            send_tg(token, cid, "\u274c \u0647\u06cc\u0686 \u06a9\u0627\u0631\u0628\u0631\u06cc \u062b\u0628\u062a \u0646\u0634\u062f\u0647")
                        else:
                            ok_count = 0
                            for tc in all_cids:
                                r = send_tg(token, tc, body_text)
                                if r: ok_count += 1
                            send_tg(token, cid, f"\u2705 \u067e\u06cc\u0627\u0645 \u0628\u0647 {ok_count} \u0646\u0641\u0631 \u0627\u0631\u0633\u0627\u0644 \u0634\u062f")

  except Exception as e:
    print(f"[do_update] {e}")

notified = set()
_deleted_ids: set = set()  # آلارم‌هایی که پاک شدن — دیگه fire نشن
_loop_count = 0
# حلقه‌ی live price هر ۱ دقیقه یک بار اجرا می‌شود.
CHECK_INTERVAL_SECONDS = 60
# ⏸️ موقت: نمادهایی که آلارمشون فایر می‌شه ولی پیام تلگرام نمی‌فرستیم (فقط طلا، فعلاً)
TEMP_MUTED_SYMBOLS = {"XAUUSD"}
# refresh آلارم‌های فعال از Supabase هر ۵ دور = حدود ۱۰ دقیقه یک بار.
ACTIVE_REFRESH_EVERY_LOOPS = 5

_price_fetch_executor = ThreadPoolExecutor(max_workers=8, thread_name_prefix="price_fetch")
PRICE_FETCH_TIMEOUT = 15  # ثانیه — اگه یه API بیشتر از این طول کشید، skip میشه

def _fetch_with_timeout(fn, *args, timeout=PRICE_FETCH_TIMEOUT):
    """یه تابع price رو با timeout مستقل اجرا کن"""
    future = _price_fetch_executor.submit(fn, *args)
    try:
        return future.result(timeout=timeout)
    except FuturesTimeoutError:
        print(f"[check] ⚠️ price fetch timeout ({timeout}s): {args}")
        future.cancel()
        return None
    except Exception as e:
        print(f"[check] price fetch error: {e}")
        return None

def _load_fired_backup_ids() -> set:
    """لود آیدی‌های fired از فایل بک‌آپ محلی — لایه‌ی دفاعی دوم مستقل از Supabase"""
    if os.path.exists(FIRED_BACKUP_FILE):
        try:
            with open(FIRED_BACKUP_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            return set()
    return set()

def _sb_load_active_only():
    """
    فقط ردیف‌های فعال (status=active) + ردیف کانفیگ رو از Supabase بخون —
    بدون آرشیو تاریخی. حلقه‌ی زنده‌ی چک قیمت فقط به آلارم‌های فعال نیاز داره؛
    آرشیو جدا و فقط وقتی کاربر واقعاً صفحه‌ی بایگانی رو تو سایت باز می‌کنه
    خونده می‌شه (endpoint /api/archive از قبل مستقیم و بدون کش این کارو می‌کنه).
    """
    if not SUPABASE_KEY: return None
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?or=(id.eq.__config__,status.eq.active)&select=id,status,active,symbol,type,condition,target_price,created_by,created_at,comment,is_private,private_cid,notify_only,last_price,last_checked,fired_at,fired_price,expires_at,telegram_token,chat_ids,users&limit=500",
            headers=_sb_h(), timeout=10)
        if r.status_code != 200:
            print(f"[alerts] load active-only failed: {r.status_code} {r.text[:80]}")
            return None
        rows = r.json()
        if not rows:
            return None
        config_row = next((x for x in rows if x["id"] == "__config__"), None)
        tg = {"bot_token": "", "chat_ids": []}
        users = []
        if config_row:
            tg["bot_token"] = config_row.get("telegram_token","") or ""
            raw_cids = config_row.get("chat_ids") or []
            tg["chat_ids"] = raw_cids if isinstance(raw_cids, list) else json.loads(raw_cids)
            raw_users = config_row.get("users") or []
            users = raw_users if isinstance(raw_users, list) else json.loads(raw_users)
        alerts = []
        for row in rows:
            if row["id"] == "__config__": continue
            if row.get("status") != "active" or not row.get("active"):
                continue
            alerts.append({
                "id":           row["id"],
                "symbol":       row.get("symbol",""),
                "type":         row.get("type","forex"),
                "condition":    row.get("condition","above"),
                "target_price": row.get("target_price",0),
                "created_by":   row.get("created_by",""),
                "created_at":   row.get("created_at",""),
                "comment":      row.get("comment",""),
                "is_private":   row.get("is_private", False),
                "private_cid":  row.get("private_cid") or row.get("notify_only"),
                "notify_only":  row.get("notify_only"),
                "active":       row.get("active", True),
                "last_price":   row.get("last_price"),
                "last_checked": row.get("last_checked"),
                "fired_at":     row.get("fired_at"),
                "fired_price":  row.get("fired_price"),
                "expires_at":   row.get("expires_at"),
            })
        data = {"alerts": alerts, "archive": [], "telegram": tg,
                "users": users, "errors": [], "last_update": now_teh()}
        print(f"[alerts] Loaded ACTIVE-ONLY from Supabase — {len(alerts)} active (بدون آرشیو)")
        return data
    except Exception as e:
        print(f"[alerts] load active-only error: {e}")
        return None


def check_alerts():
    global _loop_count
    while True:
        try:
            _loop_count += 1
            global _cache_alerts
            with _alerts_cache_lock:
                # حلقه‌ی چک قیمت هر ۲ دقیقه اجرا می‌شود. برای کم کردن Egress،
                # خودِ Supabase را فقط هر ۵ دور (حدود ۱۰ دقیقه) برای آلارم‌های فعال refresh می‌کنیم.
                # بین این refreshها از cache حافظه استفاده می‌شود. ثبت/حذف آلارم از
                # endpointهای خودش cache را به‌روز می‌کند، بنابراین نیاز به fetch مداوم نیست.
                if _loop_count % ACTIVE_REFRESH_EVERY_LOOPS == 1:
                    fresh = _sb_load_active_only()
                    if fresh is not None:
                        _cache_alerts = fresh
                    # اگه fresh None بود (خطای موقت شبکه/Supabase)، کش قبلی رو دست
                    # نخورده نگه می‌داریم — نال کردنش باعث می‌شد هر load_alerts()
                    # بعدی (از هر جای برنامه) مجبور به یه fetch کامل و سنگین بشه.
                token, cids, data = _get_token_and_cids()
            _fired_backup_ids = _load_fired_backup_ids()
            all_active_raw = [a for a in data.get("alerts", [])
                      if a.get("active") and str(a["id"]) not in _deleted_ids
                      and not a.get("fired_at")
                      and str(a["id"]) not in _fired_backup_ids]

            # ── چک انقضا — از همون لیستی که همین الان تو حافظه لود شده استفاده می‌کنه،
            # هیچ خوندن اضافه‌ای از Supabase لازم نداره. فقط وقتی آلارمی واقعاً منقضی
            # بشه (نادر)، رکوردش با status='expired' علامت‌گذاری می‌شه — نه حذف کامل،
            # تا همچنان با SQL قابل خروجی گرفتن باشه. ──
            now_naive_teh = datetime.now(TEHRAN).replace(tzinfo=None)
            active = []
            for a in all_active_raw:
                exp_str = a.get("expires_at")
                if exp_str:
                    try:
                        exp_dt = datetime.strptime(exp_str, "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        exp_dt = None
                    if exp_dt and now_naive_teh >= exp_dt:
                        _deleted_ids.add(str(a["id"]))
                        print(f"[expiry] ⌛ {a.get('symbol')} ({a['id']}) منقضی شد — status='expired' می‌شه")
                        threading.Thread(target=save_alert_expired, args=(a,), daemon=True).start()
                        continue
                active.append(a)
            if not active:
                time.sleep(CHECK_INTERVAL_SECONDS)
                continue
            forex_open = is_forex_market_open()
            due_forex, due_crypto = [], []
            for a in active:
                sym = a["symbol"]
                atype = a.get("type", "crypto")
                if atype == "forex" and not forex_open:
                    continue
                if atype == "forex":
                    due_forex.append(sym)
                else:
                    due_crypto.append(sym)

            price_map = {}

            # forex: یه batch call با timeout
            if due_forex:
                batch = _fetch_with_timeout(get_forex_prices_batch, due_forex)
                if batch:
                    for sym, p in batch.items():
                        price_map[(sym, "forex")] = p

            # crypto: همزمان، هر کدوم timeout مستقل
            if due_crypto:
                unique_crypto = list(dict.fromkeys(s.upper() for s in due_crypto))
                futures_map = {sym: _price_fetch_executor.submit(get_crypto_price, sym) for sym in unique_crypto}
                for sym, fut in futures_map.items():
                    try:
                        p = fut.result(timeout=PRICE_FETCH_TIMEOUT)
                        price_map[(sym, "crypto")] = p
                    except FuturesTimeoutError:
                        print(f"[check] ⚠️ crypto timeout: {sym}")
                        fut.cancel()
                    except Exception as e:
                        print(f"[check] crypto error {sym}: {e}")

            print(f"[check] loop={_loop_count} forex_open={forex_open} due_f={len(due_forex)} due_c={len(due_crypto)} prices={len(price_map)}")
            fired = []
            now = now_teh()
            for a in active:
                sym = a["symbol"]
                atype = a.get("type", "crypto")
                key = (sym.upper(), atype)
                if key not in price_map: continue
                cur = price_map[key]
                if cur is None: continue
                tgt = float(a["target_price"])
                cond = a.get("condition", "above")
                # ✅ آپدیت قیمت لحظه‌ای برای همه آلارم‌ها
                a["last_price"] = cur
                a["last_checked"] = now
                stale_ts = price_map.get((sym.upper(), atype, "stale"))
                a["price_stale"] = stale_ts if stale_ts else None
                data["last_update"] = now
                triggered = (cond == "above" and cur >= tgt) or (cond == "below" and cur <= tgt)
                print(f"[check] {sym} cur={fmt_price(cur,sym)} tgt={fmt_price(tgt,sym)} cond={cond} → {'🔥 FIRE' if triggered else 'ok'}")
                if triggered and a["id"] not in notified and str(a["id"]) not in _deleted_ids:
                    notified.add(a["id"])
                    _deleted_ids.add(str(a["id"]))  # ← فوری blacklist — جلوگیری از double-fire
                    a["active"] = False
                    a["fired_at"] = now
                    a["fired_price"] = cur
                    alarm_num_tag = _make_alarm_tag(sym)
                    a["tag"] = alarm_num_tag
                    fired.append(a["id"])
                    # 🔒 فوری و سینک قبل از هرکاری (ارسال پیام/تعیین مسئول) تو Supabase ذخیره می‌کنیم —
                    # تا اگه سرور وسط ارسال پیام‌ها کرش/ریستارت کرد، این آلارم already-fired بمونه و دوباره فایر نشه
                    save_alert_fired(a)
                    # ⏸️ موقت: آلارم طلا (XAUUSD) فایر می‌شه و ثبت می‌شه، فقط پیام تلگرام ارسال نشه
                    if sym.upper() in TEMP_MUTED_SYMBOLS:
                        print(f"[FILTER] {sym} در لیست موقت بی‌صدا — پیام تلگرام ارسال نشد")
                    elif token and cids:
                        comment = a.get("comment", "")
                        if str(YOUR_CHAT_ID) in str(comment):
                            notify_cids = [str(YOUR_CHAT_ID)]
                            print(f"[FILTER] comment contains {YOUR_CHAT_ID} → only to you")
                        elif a.get("notify_only"):
                            notify_cids = [str(a["notify_only"])]
                            print(f"[FILTER] notify_only → {a['notify_only']}")
                        else:
                            notify_cids = cids
                            print(f"[FILTER] broadcast → {len(cids)} users")
                        arrow = "📈 ناحیه سل" if cond == "above" else "📉 ناحیه بای"
                        creator = a.get("created_by") or "سیستم"
                        cmt = f"\n💬 <i>{comment}</i>" if comment else ""
                        dist = calc_dist_str(sym, atype, cur, tgt)
                        private_label = "\n\n🔒 <i>آلارم شخصی — فقط برای شما ارسال شده</i>" if a.get("is_private") else ""
                        creator_tag = "#" + re.sub(r'[^\w]', '_', creator).strip('_')
                        # ── تعیین مسئول تریگر — فقط برای آلارم‌های تیمی، آلارم شخصی وارد تقسیم نمی‌شه ──
                        if a.get("is_private"):
                            _assignee, _shift = "", None
                            assignee_line = ""
                        else:
                            _assignee, _shift = _get_assignee_for_alarm(
                                a["id"], alarm_num_tag, now,
                                symbol=sym, target_price=float(tgt), created_by=creator
                            )
                            assignee_line = f"\n\n🎯 مسئول تریگر: <b>{_assignee}</b>" if _assignee else ""
                        created_at_raw = a.get("created_at", "")
                        created_label = f" | 📅 ثبت: <i>{created_at_raw[:16]}</i>" if created_at_raw else ""
                        fired_msg = (
                            f"🚨 <b>آلارم قیمت!</b>\n\n"
                            f"💰 <b>#{sym}</b> — {arrow}\n"
                            f"🔖 {alarm_num_tag}\n"
                            f"👤 {creator_tag}\n\n"
                            f"🎯 هدف: <code>{fmt_price(tgt,sym)}</code>\n"
                            f"📊 قیمت لحظه‌ای: <b>{fmt_price(cur,sym)}</b>\n"
                            f"📏 فاصله: <b>{dist}</b>"
                            f"{cmt}"
                            f"{private_label}"
                            f"{assignee_line}\n\n⏰ {now_pretty()} (تهران){created_label}"
                        )
                        # دکمه هشدار دوره‌ای برای همه — چه شخصی چه عمومی
                        # + دکمه «دیدم» فقط تو PV خودِ مسئول تریگر نمایش داده می‌شه
                        # (هر chat_id همون PV شخصی هر کاربره، پس بقیه اصلاً دکمه رو نمی‌بینن)
                        assignee_ack_id = TEAM_MEMBER_IDS.get(_assignee, "") if _assignee else ""
                        def _fired_kb(cid):
                            rows_kb = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{cid}:{sym}"}]]
                            if assignee_ack_id and str(cid) == str(assignee_ack_id):
                                rows_kb.append([{"text": "✅ دیدم", "callback_data": f"ack_trigger:{a['id']}:{assignee_ack_id}"}])
                            return rows_kb
                        fired_cid_to_mid = {}
                        if a.get("is_private") and a.get("notify_only"):
                            priv_cid = str(a["notify_only"])
                            mid_f = send_tg_keyboard(token, priv_cid, fired_msg, _fired_kb(priv_cid), track=False)
                            if mid_f:
                                fired_cid_to_mid[priv_cid] = mid_f
                        else:
                            for cid in notify_cids:
                                mid_f = send_tg_keyboard(token, str(cid), fired_msg, _fired_kb(str(cid)), track=False)
                                if mid_f:
                                    fired_cid_to_mid[str(cid)] = mid_f
                        # ذخیره map چت→پیام برای /del
                        if fired_cid_to_mid:
                            fired_cid_to_mid["__tag__"] = alarm_num_tag
                            fired_cid_to_mid["__text__"] = fired_msg
                            _fired_msg_ids[a["id"]] = fired_cid_to_mid
                            threading.Thread(target=_sb_save_fired_msgs, args=(a["id"], fired_cid_to_mid), daemon=True).start()
            if fired:
                # save_alert_fired() برای هر آلارم Fired شده، وضعیت نهایی را مستقیم
                # در Supabase ذخیره و cache را هم به‌روز کرده است. بنابراین اینجا
                # نباید save_alerts(data) را اجرا کنیم؛ چون آن تابع همه‌ی آلارم‌های
                # فعال باقی‌مانده را دوباره Upsert می‌کرد و Egress/Write بی‌دلیل می‌ساخت.
                with _alerts_cache_lock:
                    if _cache_alerts is not None:
                        # cache را از data هم همگام نگه می‌داریم، بدون هیچ درخواست شبکه‌ای
                        data = _cache_alerts
            # مهم: وقتی هیچ آلارمی fire نشده، هیچ Write به Supabase انجام نمی‌شود.
            # last_price و last_checked فقط در RAM تغییر می‌کنند و برای مقایسه‌ی
            # قیمت فعلی با target استفاده می‌شوند.
        except Exception as e:
            log_error(f"check_alerts: {e}")
        time.sleep(CHECK_INTERVAL_SECONDS)

def fmt_price(p, sym=""):
    if p is None: return "—"
    v = float(p)
    su = sym.upper()
    if "XAU" in su or "XAG" in su:
        return f"${v:.2f}"
    if "JPY" in su:
        return f"{v:.3f}"
    return f"{v:.5f}"

def calc_dist_str(symbol, atype, cur, tgt):
    if not cur or not tgt: return ""
    diff = abs(float(cur) - float(tgt))
    sym_up = symbol.upper()
    if atype == "crypto": return f"{diff/float(tgt)*100:.2f}%"
    if "XAU" in sym_up or "XAG" in sym_up: return f"{diff:.2f} $"
    if "JPY" in sym_up: return f"{round(diff*100):,} pip"
    return f"{round(diff*10000):,} pip"

# ==================== Routes ====================
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/config", methods=["GET","POST"])
def config():
    data = load_alerts()
    if request.method == "POST":
        body = request.json or {}
        tg = data.get("telegram", {})
        if body.get("bot_token"):
            tg["bot_token"] = body["bot_token"]
        if body.get("chat_id"):
            cid = str(body["chat_id"])
            ids = [str(x) for x in tg.get("chat_ids", [])]
            if cid not in ids: ids.append(cid)
            tg["chat_ids"] = ids
            tg["chat_id"] = cid
        data["telegram"] = tg
        save_alerts(data)
        return jsonify({"ok": True})
    tg = data.get("telegram", {})
    return jsonify({
        "bot_token": tg.get("bot_token",""), "chat_id": tg.get("chat_id",""),
        "chat_ids": tg.get("chat_ids",[]), "user_count": len(data.get("users",[]))
    })

@app.route("/api/alerts", methods=["GET"])
def get_alerts():
    all_alerts = load_alerts().get("alerts", [])
    # آلارم‌های شخصی (is_private=True) رو از لیست عمومی حذف کن
    public = [a for a in all_alerts if not a.get("is_private")]
    return jsonify(public)

@app.route("/api/alerts/my", methods=["GET"])
def get_my_alerts():
    """آلارم‌های شخصیِ یه کاربر (چه عادی چه شخصی/is_private) — با name فیلتر میشه.
    اگه این اسم صاحب کد فعال‌سازی (web_pin) باشه، حتماً باید pin هم درست بیاد،
    وگرنه (کاربرهایی که هنوز کد نگرفتن) طبق روال قبل فقط با تطابق اسم برمی‌گرده."""
    name = request.args.get("name", "").strip()
    cid  = request.args.get("cid",  "").strip()
    pin  = request.args.get("pin",  "").strip()
    if not name and not cid:
        return jsonify([])
    data = load_alerts()
    all_alerts = data.get("alerts", [])
    name_lc = name.lower()
    resolved_cid = cid
    matched_user = None
    if name and not resolved_cid:
        for usr in data.get("users", []):
            if usr.get("custom_name", "").strip().lower() == name_lc:
                matched_user = usr
                resolved_cid = str(usr.get("chat_id", ""))
                break
    if not matched_user and resolved_cid:
        for usr in data.get("users", []):
            if str(usr.get("chat_id", "")) == str(resolved_cid):
                matched_user = usr
                break
    if matched_user and matched_user.get("web_pin"):
        if str(pin) != str(matched_user.get("web_pin")):
            return jsonify([])
    if matched_user and matched_user.get("full_access"):
        # این کاربر از پنل ادمین دسترسی کامل گرفته — کل آلارم‌ها (عادی + شخصی) رو ببینه
        return jsonify(all_alerts)
    my = [
        a for a in all_alerts
        if (
            a.get("is_private") and (
                (resolved_cid and str(a.get("private_cid", "")) == resolved_cid) or
                (name and a.get("created_by", "").strip().lower() == name_lc)
            )
        ) or (
            not a.get("is_private") and name and a.get("created_by", "").strip().lower() == name_lc
        )
    ]
    return jsonify(my)

@app.route("/api/alerts", methods=["POST"])
def add_alert():
    body = request.json or {}
    sym = body.get("symbol","").upper().strip()
    atype = body.get("type","forex")
    tgt = float(body.get("target_price", 0))
    creator = body.get("creator", "").strip()
    pin = body.get("pin", "").strip()
    if not _is_name_approved(creator, pin):
        return jsonify({"ok": False, "error": "این نام هنوز از طرف ادمین تایید نشده. لطفاً ابتدا از طریق ربات تلگرام (/start) درخواست فعال‌سازی بدید."}), 403
    # آلارم شخصی فقط برای کسایی که از پنل ادمین/تلگرام دسترسیش رو گرفتن — وگرنه سایلنت عادی ثبت می‌شه
    is_private = False
    private_cid = None
    if bool(body.get("is_private", False)):
        u_priv = _verified_user_by_name(creator, pin)
        if u_priv and u_priv.get("private_access"):
            is_private = True
            private_cid = str(u_priv.get("chat_id", ""))
    cur = get_price(sym, atype) if (atype!="forex" or is_forex_market_open()) else None
    expire_days = body.get("expire_days")
    expires_at = None
    try:
        expire_days = int(expire_days) if expire_days not in (None, "", 0, "0") else 0
    except (TypeError, ValueError):
        expire_days = 0
    if expire_days > 0:
        expires_at = (datetime.now(TEHRAN) + timedelta(days=expire_days)).strftime("%Y-%m-%d %H:%M:%S")
    a = {
        "id": str(int(time.time() * 1000)), "symbol": sym, "type": atype,
        "target_price": tgt, "condition": body.get("condition","above"),
        "comment": body.get("comment","").strip(), "active": True,
        "created_by": creator or "ناشناس",
        "is_private": is_private,
        "private_cid": private_cid,
        "notify_only": private_cid,
        "expires_at": expires_at,
        "last_price": cur, "last_checked": now_teh() if cur else None,
        "created_at": now_teh()
    }
    # قفل: خوندن، اضافه‌کردن، و ذخیره باید یه عملیات atomic باشه —
    # وگرنه اگه check_alerts هم‌زمان کش رو رفرش کنه، یکی از تغییرات گم می‌شه
    with _alerts_cache_lock:
        data = load_alerts()
        data["alerts"].append(a)
        save_alerts(data)
    return jsonify({"ok": True, "alert": a})

@app.route("/api/alarm-assignments/<aid>/false", methods=["POST"])
def web_false_assignment(aid):
    """فالس کردن یه آلارم مشخص از صفحه‌ی گزارش هفتگی وب (با کامنت/علت) — نیاز به اسم+کد تاییدشده داره.
    فقط برای آلارم‌های تیمی (غیرشخصی) کار می‌کنه، دقیقاً مثل چیزی که تو گزارش هفتگی نمایش داده می‌شه."""
    body = request.json or {}
    name = (body.get("name") or "").strip()
    pin = (body.get("pin") or "").strip()
    reason = (body.get("reason") or "").strip()
    u = _verified_user_by_name(name, pin)
    if not u:
        return jsonify({"ok": False, "error": "هویت تایید نشد. لطفاً اسم و کد فعال‌سازی درست رو وارد کنید."}), 403

    alert_wfa = None
    if SUPABASE_KEY:
        try:
            r_wfa = requests.get(
                f"{SUPABASE_URL}/rest/v1/alerts?id=eq.{aid}&select=id,is_private",
                headers=_sb_h(), timeout=8)
            if r_wfa.status_code == 200:
                found_wfa = r_wfa.json()
                alert_wfa = found_wfa[0] if found_wfa else None
        except Exception as e:
            print(f"[web_false_assignment] targeted fetch error: {e}")
    if alert_wfa and alert_wfa.get("is_private"):
        return jsonify({"ok": False, "error": "این آلارم شخصیه و از این صفحه قابل فالس‌کردن نیست."}), 403

    sender_name = u.get("custom_name") or name
    with _false_in_progress_lock:
        if aid in _false_in_progress:
            return jsonify({"ok": False, "error": "این آلارم همین الان داره پردازش می‌شه، چند ثانیه صبر کنید."}), 409
        _false_in_progress.add(aid)
    try:
        _sb_false_assignment(aid, sender_name, reason)
        _rebuild_active_assign_count(_sb_load_active_assignments())
    finally:
        with _false_in_progress_lock:
            _false_in_progress.discard(aid)
    return jsonify({"ok": True})


@app.route("/api/alerts/<aid>", methods=["DELETE"])
def del_alert(aid):
    data = load_alerts()
    data["alerts"] = [a for a in data.get("alerts", []) if a["id"] != aid]
    global _cache_alerts
    _cache_alerts = data
    threading.Thread(target=_sb_delete_alert, args=(aid,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/users", methods=["GET"])
def get_users():
    return jsonify(load_alerts().get("users", []))

@app.route("/api/users/<cid>", methods=["DELETE"])
def del_user(cid):
    data = load_alerts()
    data["users"] = [u for u in data.get("users",[]) if str(u["chat_id"]) != str(cid)]
    data["telegram"]["chat_ids"] = [x for x in data["telegram"].get("chat_ids",[]) if str(x) != str(cid)]
    save_alerts(data)
    return jsonify({"ok": True})

@app.route("/api/price/<atype>/<symbol>")
def live_price(atype, symbol):
    sym = symbol.upper().replace("-","/")
    p = get_price(sym, atype)
    if p is None:
        return jsonify({"error": "قیمت پیدا نشد"}), 404
    return jsonify({"symbol": sym, "price": p})

@app.route("/api/instant-alert", methods=["POST"])
def instant_alert():
    body = request.json or {}
    sym = body.get("symbol", "").upper().strip()
    if not sym:
        return jsonify({"ok": False, "error": "نماد وارد نشده"}), 400
    atype = body.get("type", "forex")
    condition = body.get("condition", "above")
    comment = body.get("comment", "").strip()
    creator = body.get("creator", "").strip()
    target_price = body.get("target_price")
    only_me = body.get("only_me", False)
    pin = body.get("pin", "").strip()

    if not _is_name_approved(creator, pin):
        return jsonify({"ok": False, "error": "این نام هنوز از طرف ادمین تایید نشده. لطفاً ابتدا از طریق ربات تلگرام (/start) درخواست فعال‌سازی بدید."}), 403

    token, all_cids, data = _get_token_and_cids()
    if not token:
        return jsonify({"ok": False, "error": "توکن تلگرام تنظیم نشده"}), 400

    targets = [YOUR_CHAT_ID] if only_me else (all_cids if BROADCAST_MODE else [YOUR_CHAT_ID])
    if not targets:
        return jsonify({"ok": False, "error": "هیچ chat_id‌ای ثبت نشده"}), 400

    # قیمت لحظه‌ای
    cur = None
    try:
        cur = get_price(sym, atype)
    except:
        pass

    arrow = "📈 ناحیه سل" if condition == "above" else "📉 ناحیه بای"
    cmt = f"\n💬 <i>{comment}</i>" if comment else ""
    price_text = fmt_price(cur, sym) if cur else "—"
    _creator = creator or 'سیستم'
    alarm_num_tag = _make_alarm_tag(sym)
    creator_tag = "#" + re.sub(r'[^\w]', '_', _creator).strip('_')
    out_msg = (
        f"🚨 <b>{'آلارم قیمت' if target_price else 'آلارم فوری'}!</b>\n\n"
        f"💰 <b>#{sym}</b> — {arrow}\n"
        f"🔖 {alarm_num_tag}\n"
        f"👤 {creator_tag}\n\n"
        + (f"🎯 هدف: <code>{fmt_price(target_price, sym)}</code>\n" if target_price else "")
        + f"📊 قیمت لحظه‌ای: <b>{price_text}</b>"
        f"{cmt}\n\n⏰ {now_pretty()} (تهران)"
    )

    # هر کاربر جداگانه با دکمه هشدار دوره‌ای
    sent_count = 0
    if sym.upper() not in TEMP_MUTED_SYMBOLS:
        for cid in targets:
            kb = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{cid}:{sym}"}]]
            mid = send_tg_keyboard(token, str(cid), out_msg, kb)
            if mid: sent_count += 1

    # ذخیره در آرشیو
    try:
        d = load_alerts()
        d.setdefault("archive", []).append({
            "id": str(int(time.time() * 1000)),
            "symbol": sym, "type": atype,
            "condition": condition, "comment": comment,
            "created_by": creator, "active": False,
            "fired_at": now_teh(), "fired_price": cur,
            "target_price": target_price,
            "instant": True, "created_at": now_teh()
        })
        save_alerts(d)
    except Exception as e:
        log_error(f"instant_alert archive: {e}")

    print(f"[INSTANT] {sym} ارسال شد به {sent_count}/{len(targets)} نفر")
    return jsonify({"ok": True, "sent": sent_count, "total": len(targets)})


def test_tg():
    token, cids, _ = _get_token_and_cids()
    if not token or not cids:
        return jsonify({"ok": False, "error": "توکن یا chat_id ست نشده"})
    res = broadcast(token, cids, f"✅ تست موفق\n⏰ {now_pretty()}")
    return jsonify({"ok": any(res), "sent": sum(res), "total": len(cids)})

# ===================== SIGNALS API (وب‌سایت) =====================
SIGNAL_VALID_DIRECTIONS = {"buy_limit", "buy_stop", "sell_limit", "sell_stop"}

def _sb_delete_signal(sig_id):
    """حذف سیگنال از Supabase"""
    if not SUPABASE_KEY: return
    try:
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/signals?id=eq.{sig_id}",
            headers=_sb_h(), timeout=8)
    except: pass

def _build_signal_record(body: dict):
    """از روی دیتای فرم سایت، رکوردی عیناً هم‌ساختار با چیزی که ربات می‌سازه برمی‌گردونه"""
    sym = (body.get("symbol") or "").upper().strip()
    direction = (body.get("direction") or "").lower().strip()
    if not sym:
        return None, "نماد وارد نشده"
    if direction not in SIGNAL_VALID_DIRECTIONS:
        return None, "جهت سیگنال نامعتبره (buy_limit/buy_stop/sell_limit/sell_stop)"
    try:
        entry = float(body.get("entry"))
    except (TypeError, ValueError):
        return None, "قیمت ورود الزامی است"

    def _f(key):
        v = body.get(key)
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    sl  = _f("sl")
    tp1 = _f("tp1")
    tp2 = _f("tp2")
    tp3 = _f("tp3")
    tf  = (body.get("tf") or SIGNAL_DEFAULT_TF).strip()
    note = (body.get("note") or "").strip()
    creator = (body.get("creator") or "وب‌سایت").strip()

    risk_pips = round(abs(entry - sl) * get_pip_multiplier(sym), 1) if sl is not None else None
    rr = None
    if sl is not None and tp1 is not None and abs(entry - sl) > 0:
        rr = round(abs(tp1 - entry) / abs(entry - sl), 2)

    seq = _sb_next_signal_seq()
    sig = {
        "id": f"S{seq:05d}", "seq": seq, "symbol": sym, "direction": direction,
        "entry": entry, "sl": sl, "tp1": tp1, "tp2": tp2, "tp3": tp3,
        "tf": tf, "risk_pips": risk_pips, "rr": rr,
        "sent_by": creator, "sent_at": now_teh(), "channel_msg_id": None,
        "status": "active", "note": note or None,
    }
    return sig, None

@app.route("/api/signals", methods=["GET"])
def get_signals():
    limit = request.args.get("limit", 50)
    try: limit = int(limit)
    except: limit = 50
    return jsonify(_sb_load_signals(limit=limit))

@app.route("/api/signals", methods=["POST"])
def add_signal():
    """ثبت سیگنال در دیتابیس — بدون ارسال به کانال تلگرام"""
    body = request.json or {}
    sig, err = _build_signal_record(body)
    if err:
        return jsonify({"ok": False, "error": err}), 400
    _sb_save_signal(sig)
    sig["text"] = _build_signal_text(sig)
    return jsonify({"ok": True, "signal": sig})

@app.route("/api/signals/<sig_id>", methods=["DELETE"])
def del_signal(sig_id):
    threading.Thread(target=_sb_delete_signal, args=(sig_id,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/send-signal", methods=["POST"])
def send_signal():
    """ثبت سیگنال + ارسال عیناً به کانال تلگرام، با همون فرمتی که ربات می‌سازه"""
    body = request.json or {}
    sig, err = _build_signal_record(body)
    if err:
        return jsonify({"ok": False, "error": err}), 400

    token, _, _ = _get_token_and_cids()
    channel_mid = None
    if token and SIGNAL_CHANNEL:
        try:
            r_ch = requests.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": SIGNAL_CHANNEL, "text": _build_signal_text(sig), "parse_mode": "HTML"},
                timeout=10, headers=H)
            if r_ch.status_code == 200:
                channel_mid = r_ch.json().get("result", {}).get("message_id")
                sig["channel_msg_id"] = channel_mid
        except Exception as e:
            print(f"[signal] web channel send error: {e}")

    _sb_save_signal(sig)
    sig["text"] = _build_signal_text(sig)
    if not SIGNAL_CHANNEL:
        return jsonify({"ok": True, "signal": sig, "sent": 0,
                         "warning": "کانال سیگنال (SIGNAL_CHANNEL) تنظیم نشده — فقط در دیتابیس ذخیره شد"})
    return jsonify({"ok": True, "signal": sig, "sent": 1 if channel_mid else 0})
# ===================================================================

@app.route("/api/status")
def status():
    alerts = load_alerts()
    all_active = [a for a in alerts.get("alerts", []) if a.get("active")]
    team_active = [a for a in all_active if not a.get("is_private")]
    private_active = [a for a in all_active if a.get("is_private")]
    return jsonify({
        "status": "ok", "last_update": alerts.get("last_update"),
        "errors": alerts.get("errors", [])[-5:], "time_tehran": now_teh(),
        "alert_count": len(all_active),          # کل (تیمی + شخصی)
        "team_alert_count": len(team_active),     # فقط تیمی
        "private_alert_count": len(private_active), # فقط شخصی
        "forex_open": is_forex_market_open(),
        "loop_count": _loop_count
    })

@app.route("/api/version")
def version():
    return jsonify({"version": VERSION})

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

@app.route("/api/assignments", methods=["GET"])
def api_assignments():
    """
    لیست کامل assignments با join از alerts جدول.
    پارامترها:
      ?active=true    — فقط فعال‌ها
      ?week=true      — از ابتدای هفته جاری (شنبه تهران)
      ?all=true       — همه (پیش‌فرض)
    """
    try:
        active_only = request.args.get("active") == "true"
        week_only   = request.args.get("week") == "true"

        # لود assignments از Supabase
        url = f"{SUPABASE_URL}/rest/v1/alarm_assignments?select=*&order=fired_at.desc"
        if active_only:
            url += "&is_active=eq.true"
        if week_only:
            now_dt = datetime.now(TEHRAN)
            days_since_sat = (now_dt.weekday() - 5) % 7
            week_start = (now_dt - timedelta(days=days_since_sat)).replace(
                hour=0, minute=0, second=0, microsecond=0)
            url += f"&fired_at=gte.{week_start.strftime('%Y-%m-%dT%H:%M:%S')}"

        r = requests.get(url, headers=_sb_h(), timeout=10)
        if r.status_code != 200:
            return jsonify({"ok": False, "error": r.text[:100]}), 500
        rows = r.json()

        # لود فقط alertهایی که تو این assignmentها هستن (نه کل جدول)
        ids_aa = sorted({str(rw.get("id","")) for rw in rows if rw.get("id")})
        alerts_map = {}
        if SUPABASE_KEY and ids_aa:
            try:
                r_aa = requests.get(
                    f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(ids_aa)})&select=*",
                    headers=_sb_h(), timeout=10)
                if r_aa.status_code == 200:
                    for row in r_aa.json():
                        alerts_map[str(row.get("id",""))] = row
            except Exception as e:
                print(f"[api_assignments] targeted fetch error: {e}")

        result = []
        for row in rows:
            aid    = str(row.get("id", ""))
            alert  = alerts_map.get(aid, {})
            sym    = alert.get("symbol", "")
            tgt    = alert.get("target_price", 0) or 0
            result.append({
                # از alarm_assignments
                "id":           aid,
                "alarm_tag":    row.get("alarm_tag", ""),
                "assigned_to":  row.get("assigned_to", "") or "",
                "shift":        row.get("shift", ""),
                "is_active":    row.get("is_active", True),
                "fired_at":     row.get("fired_at", ""),
                "false_at":     row.get("false_at", "") or "",
                "false_by":     row.get("false_by", "") or "",
                "false_reason": row.get("false_reason", "") or "",
                # از alerts جدول
                "symbol":       sym,
                "condition":    alert.get("condition", ""),
                "target_price": tgt,
                "target_fmt":   fmt_price(float(tgt), sym) if tgt else "",
                "created_by":   alert.get("created_by", "") or "",
                "created_at":   alert.get("created_at", "") or "",
                "comment":      alert.get("comment", "") or "",
                "fired_price":  alert.get("fired_price", "") or "",
                "is_private":   alert.get("is_private", False),
            })

        # فیلتر آلارم‌های شخصی
        result = [r for r in result if not r["is_private"]]
        return jsonify({"ok": True, "count": len(result), "items": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# =====================================================================
# 🌐 پنل ادمین وب — API endpoints
# =====================================================================

@app.route("/admin-panel")
def admin_panel_page():
    """صفحه‌ی HTML پنل ادمین وب — لاگین + داشبورد"""
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), "admin_panel.html")


@app.route("/api/admin-panel/login", methods=["POST"])
def admin_panel_login():
    body = request.json or {}
    pw = body.get("password", "")
    if not ADMIN_PANEL_PASSWORD:
        return jsonify({"ok": False, "error": "رمز پنل تنظیم نشده — باید ADMIN_PANEL_PASSWORD رو تو Railway بذاری"}), 500
    if not secrets.compare_digest(pw, ADMIN_PANEL_PASSWORD):
        return jsonify({"ok": False, "error": "رمز اشتباهه"}), 401
    token = _create_admin_session()
    return jsonify({"ok": True, "token": token})


@app.route("/api/admin-panel/assignments", methods=["GET"])
def admin_panel_assignments():
    """لیست آلارم‌های فعال با مسئول‌هاشون — برای بخش «مسئولین آلارم» پنل وب"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    rows = _sb_load_active_assignments()
    ids_apa = sorted({str(rw.get("id","")) for rw in rows if rw.get("id")})
    alerts_map = {}
    if SUPABASE_KEY and ids_apa:
        try:
            r_apa = requests.get(
                f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(ids_apa)})&select=*",
                headers=_sb_h(), timeout=10)
            if r_apa.status_code == 200:
                for row in r_apa.json():
                    alerts_map[str(row.get("id",""))] = row
        except Exception as e:
            print(f"[admin_panel_assignments] targeted fetch error: {e}")

    items = []
    for row in rows:
        aid = str(row.get("id", ""))
        alert = alerts_map.get(aid, {})
        if alert.get("is_private"):
            continue
        sym = alert.get("symbol", "")
        tgt = alert.get("target_price", 0) or 0
        items.append({
            "id": aid,
            "alarm_tag": row.get("alarm_tag", ""),
            "assigned_to": row.get("assigned_to", "") or "",
            "symbol": sym,
            "condition": alert.get("condition", ""),
            "target_fmt": fmt_price(float(tgt), sym) if tgt else "",
            "fired_at": row.get("fired_at", ""),
        })
    return jsonify({"ok": True, "items": items, "team_members": TEAM_MEMBERS})


@app.route("/api/admin-panel/reassign", methods=["POST"])
def admin_panel_reassign():
    """جابجایی دستی مسئول یک آلارم — همون منطق پنل تلگرام"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    body = request.json or {}
    aid = str(body.get("id", ""))
    new_assignee = body.get("assignee", "")
    if not aid or new_assignee not in TEAM_MEMBERS:
        return jsonify({"ok": False, "error": "پارامتر نامعتبر"}), 400

    try:
        r_get = requests.get(
            f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}&select=*",
            headers=_sb_h(), timeout=8)
        row = r_get.json()[0] if r_get.status_code == 200 and r_get.json() else {}
    except Exception:
        row = {}
    old_assignee = row.get("assigned_to", "")
    tag = row.get("alarm_tag", "")
    already_acked = bool(row.get("ack_at"))

    if old_assignee and old_assignee in _active_assign_count:
        _active_assign_count[old_assignee] = max(0, _active_assign_count[old_assignee] - 1)
    _active_assign_count[new_assignee] = _active_assign_count.get(new_assignee, 0) + 1

    requests.patch(
        f"{SUPABASE_URL}/rest/v1/alarm_assignments?id=eq.{aid}",
        headers={**_sb_h(), "Prefer": "return=minimal"},
        json={"assigned_to": new_assignee}, timeout=8)

    # ریپلای به گروه — دقیقاً مثل جابجایی دستی تو تلگرام
    tg_tok, _, _ = _get_token_and_cids()
    msg_map = _fired_msg_ids.get(aid, {})
    reply_text = (f"🔀 <b>جابجایی دستی (از پنل وب)</b>\n\n{tag}\n"
                  f"👤 مسئول جدید: <b>{new_assignee}</b>"
                  + (f"\n↩️ قبلی: {old_assignee}" if old_assignee else ""))
    for tc, tm in msg_map.items():
        if tc in ("__tag__", "__text__"):
            continue
        try:
            requests.post(f"https://api.telegram.org/bot{tg_tok}/sendMessage",
                          json={"chat_id": tc, "text": reply_text, "parse_mode": "HTML", "reply_to_message_id": tm},
                          timeout=8, headers=H)
        except Exception:
            pass

    # آپدیت دکمه‌ی «دیدم» — منتقلش کن به مسئول جدید (اگه هنوز کسی ack نکرده)
    if not already_acked:
        new_ack_id = TEAM_MEMBER_IDS.get(new_assignee, "")
        sym_r = _extract_symbol_from_tag(tag)
        for tc, tm in msg_map.items():
            if tc in ("__tag__", "__text__"):
                continue
            kb = [[{"text": "⏰ هشدار دوره‌ای", "callback_data": f"set_reminder:{tc}:{sym_r}"}]]
            if new_ack_id and str(tc) == str(new_ack_id):
                kb.append([{"text": "✅ دیدم", "callback_data": f"ack_trigger:{aid}:{new_ack_id}"}])
            try:
                requests.post(f"https://api.telegram.org/bot{tg_tok}/editMessageReplyMarkup",
                              json={"chat_id": tc, "message_id": tm, "reply_markup": {"inline_keyboard": kb}},
                              timeout=8, headers=H)
            except Exception:
                pass

    return jsonify({"ok": True})


@app.route("/api/admin-panel/users", methods=["GET"])
def admin_panel_users():
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    data = load_alerts()
    users = data.get("users", [])
    return jsonify({"ok": True, "users": [
        {"chat_id": str(u.get("chat_id", "")),
         "name": u.get("custom_name", "") or u.get("username", "") or str(u.get("chat_id", "")),
         "private_access": bool(u.get("private_access")),
         "full_access": bool(u.get("full_access"))}
        for u in users
    ]})


@app.route("/api/admin-panel/users/<chat_id>/full-access", methods=["POST"])
def admin_panel_set_full_access(chat_id):
    """فعال/غیرفعال کردن دیدن کل آلارم‌ها (نه فقط آلارم‌های خودش) برای یه کاربر — از پنل وب"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    body = request.json or {}
    value = bool(body.get("value", False))
    with _alerts_cache_lock:
        data = load_alerts()
        found = False
        for u in data.get("users", []):
            if str(u.get("chat_id", "")) == str(chat_id):
                u["full_access"] = value
                found = True
                break
        if not found:
            return jsonify({"ok": False, "error": "کاربر پیدا نشد"}), 404
        save_alerts(data)
    return jsonify({"ok": True, "full_access": value})


@app.route("/api/admin-panel/users/<chat_id>", methods=["DELETE"])
def admin_panel_delete_user(chat_id):
    """حذف کاربر — دقیقاً همون منطق admin:confirmdelete تو تلگرام"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    with _alerts_cache_lock:
        data = load_alerts()
        data["users"] = [u for u in data.get("users", []) if str(u.get("chat_id", "")) != str(chat_id)]
        data["telegram"]["chat_ids"] = [x for x in data["telegram"].get("chat_ids", []) if str(x) != str(chat_id)]
        save_alerts(data)
    return jsonify({"ok": True})


@app.route("/api/admin-panel/bulk-false", methods=["POST"])
def admin_panel_bulk_false():
    """فالس دسته‌جمعی همه‌ی آلارم‌های فعال — دقیقاً همون منطق admin:bulkfalse تو تلگرام"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    body = request.json or {}
    reason = (body.get("reason") or "").strip()

    rows = _sb_load_active_assignments()
    tg_tok, _, _ = _get_token_and_cids()
    done_count = 0
    for row in rows:
        aid = row.get("id")
        tag = row.get("alarm_tag", "")
        _sb_false_assignment(aid, "ادمین (پنل وب)", reason)
        done_count += 1
        msg_map = _fired_msg_ids.get(aid, {})
        reason_line = f"\n📝 علت: {reason}" if reason else ""
        reply_text = f"❌ <b>فالس (دسته‌جمعی از پنل وب)</b>\n\n{tag}{reason_line}"
        for tc, tm in msg_map.items():
            if tc in ("__tag__", "__text__"):
                continue
            try:
                requests.post(f"https://api.telegram.org/bot{tg_tok}/sendMessage",
                              json={"chat_id": tc, "text": reply_text, "parse_mode": "HTML", "reply_to_message_id": tm},
                              timeout=8, headers=H)
            except Exception:
                pass
    _rebuild_active_assign_count(_sb_load_active_assignments())
    return jsonify({"ok": True, "count": done_count})


@app.route("/api/admin-panel/deprioritize-masoud", methods=["GET"])
def admin_panel_get_deprioritize():
    """وضعیت فعلی تاگل اولویت پایین مسعود رو برمی‌گردونه"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    return jsonify({"ok": True, "active": _get_deprioritize_masoud()})


@app.route("/api/admin-panel/deprioritize-masoud", methods=["POST"])
def admin_panel_set_deprioritize():
    """روشن/خاموش کردن اولویت پایین مسعود از پنل وب"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    body = request.json or {}
    value = bool(body.get("active", True))
    _set_deprioritize_masoud(value)
    return jsonify({"ok": True, "active": value})


@app.route("/api/admin-panel/team-availability", methods=["GET"])
def admin_panel_get_team_availability():
    """لیست کل اعضای تیم + وضعیت در دسترس بودنشون برای تقسیم آلارم"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    unavailable = _get_unavailable_members()
    return jsonify({"ok": True, "members": [
        {"name": m, "available": m not in unavailable} for m in TEAM_MEMBERS
    ]})


@app.route("/api/admin-panel/team-availability", methods=["POST"])
def admin_panel_set_team_availability():
    """روشن/خاموش کردن دسترسی یه عضو خاص برای گرفتن آلارم جدید (مثلاً مرخصی)"""
    auth_err = _require_admin_session()
    if auth_err:
        return auth_err
    body = request.json or {}
    name = (body.get("name") or "").strip()
    available = bool(body.get("available", True))
    if name not in TEAM_MEMBERS:
        return jsonify({"ok": False, "error": "عضو ناشناخته"}), 404
    _set_member_availability(name, available)
    return jsonify({"ok": True, "name": name, "available": available})


print("=" * 60)
print(f"[STARTUP] 🚀 سرور در حال راه‌اندازی...")
print("=" * 60)
threading.Thread(target=check_alerts, daemon=True).start()
print("[STARTUP] thread check_alerts شروع شد")

# ── بازیابی reminder‌ها از Supabase بعد از restart ──
def _restore_reminders():
    """همه reminder‌های فعال رو از Supabase بخون و loop‌هاشون رو restart کن"""
    time.sleep(5)  # صبر تا bot token آماده بشه
    rows = _sb_load_reminders()
    if not rows:
        print("[STARTUP] reminder: هیچ reminder فعالی نیست")
        return
    token = BOT_TOKEN_ENV or load_alerts().get("telegram", {}).get("bot_token", "")
    if not token:
        print("[STARTUP] reminder: token نیست، restore نشد")
        return
    count = 0
    for row in rows:
        cid = str(row["chat_id"])
        sym = row["symbol"]
        interval_sec = int(row["interval_sec"])
        tf_sec = int(row.get("tf_sec") or interval_sec)
        _schedule_reminder(token, cid, sym, interval_sec, persist=False, tf_sec=tf_sec)
        count += 1
    print(f"[STARTUP] reminder: {count} هشدار بازیابی شد")
threading.Thread(target=_restore_reminders, daemon=True).start()
threading.Thread(target=daily_news_scheduler, daemon=True).start()
print(f"[STARTUP] thread daily_news_scheduler شروع شد — ارسال ساعت {FF_NEWS_HOUR:02d}:{FF_NEWS_MINUTE:02d} تهران")
threading.Thread(target=poll_telegram, daemon=True).start()
print("[STARTUP] thread poll_telegram شروع شد")

# ── بازیابی fired_msgs و counters از Supabase بعد از restart ──
_sb_load_fired_msgs()
_sb_load_sym_counters()
# بازسازی کامل state از Supabase بعد از هر restart
_sb_restore_on_startup()

# ── بازیابی notified set از Supabase — جلوگیری از double-fire بعد از restart ──
def _restore_notified():
    """آلارم‌هایی که قبلاً fire شدن رو به notified اضافه کن تا دوباره fire نشن"""
    if not SUPABASE_KEY: return
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/alerts?status=eq.fired&select=id&limit=5000",
            headers=_sb_h(), timeout=10)
        if r.status_code == 200:
            for row in r.json():
                aid = row.get("id")
                if aid:
                    notified.add(str(aid))
                    _deleted_ids.add(str(aid))
            print(f"[STARTUP] notified بازسازی شد — {len(notified)} آلارم fired")
    except Exception as e:
        print(f"[STARTUP] notified restore error: {e}")
_restore_notified()

@app.route("/api/gold-alarms")
def api_gold_alarms():
    """
    آلارم‌های XAUUSD (فعال یا فایرشده) رو برمی‌گردونه — بدون فیلتر شخصی/تیمی
    (استثنا: همه چیز طلا برای همه قابل دیدنه).
    Query params: status=active|fired, page (از ۱), per_page (پیش‌فرض ۱۰),
                  from, to (YYYY-MM-DD), sender (substring match روی created_by)
    """
    status = request.args.get("status", "active")
    page = max(1, int(request.args.get("page", 1)))
    per_page = min(50, max(1, int(request.args.get("per_page", 10))))
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()
    sender_q = request.args.get("sender", "").strip().lower()

    data = load_alerts()
    pool = data.get("alerts", []) if status == "active" else data.get("archive", [])
    items = [a for a in pool if str(a.get("symbol","")).upper() == "XAUUSD"]

    # فیلتر بازه‌ی زمانی — روی created_at برای فعال‌ها، fired_at برای فایرشده‌ها
    date_field = "created_at" if status == "active" else "fired_at"
    if date_from:
        items = [a for a in items if str(a.get(date_field) or "")[:10] >= date_from]
    if date_to:
        items = [a for a in items if str(a.get(date_field) or "")[:10] <= date_to]

    # فیلتر فرستنده
    if sender_q:
        items = [a for a in items if sender_q in str(a.get("created_by","")).lower()]

    # جدیدترین اول
    items.sort(key=lambda a: str(a.get(date_field) or a.get("created_at") or ""), reverse=True)

    total = len(items)
    pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, pages)
    start = (page - 1) * per_page
    page_items = items[start:start + per_page]

    out = []
    for a in page_items:
        out.append({
            "id": a.get("id"),
            "condition": a.get("condition"),
            "target_price": a.get("target_price"),
            "created_by": a.get("created_by") or "—",
            "created_at": a.get("created_at"),
            "comment": a.get("comment") or "",
            "fired_at": a.get("fired_at"),
            "fired_price": a.get("fired_price"),
        })

    return jsonify({"ok": True, "items": out, "total": total, "page": page, "pages": pages})


@app.route("/gold-alarms")
def gold_alarms_page():
    return send_from_directory(app.static_folder, "gold-alarms.html")


@app.route("/sva-report")
def sva_report_page():
    """
    گزارش SVA — همون فایل ثابتی که دستی (با تحلیل کامل) ساخته و تأیید شده.
    هر بار که داده‌ی جدید بیاد، این فایل static/sva-report.html جایگزین می‌شه و push می‌خوریم؛
    این مسیر همیشه آخرین نسخه‌ای که منتشر شده رو نشون می‌ده.
    """
    return send_from_directory(app.static_folder, "sva-report.html")


@app.route("/report/weekly")
def report_weekly_html():
    """گزارش آلارم‌های تیم — با بازه تاریخ دلخواه و شماره‌گذاری"""
    from_str = request.args.get("from", "")
    to_str = request.args.get("to", "")
    now_dt = datetime.now(TEHRAN)

    range_start = None
    range_end = None
    try:
        if from_str:
            range_start = TEHRAN.localize(datetime.strptime(from_str, "%Y-%m-%d"))
    except: range_start = None
    try:
        if to_str:
            range_end = TEHRAN.localize(datetime.strptime(to_str, "%Y-%m-%d")) + timedelta(days=1)
    except: range_end = None

    if range_start is None and range_end is None:
        # پیش‌فرض: هفته جاری (شنبه تا الان)
        days_since_sat = (now_dt.weekday() - 5) % 7
        range_start = (now_dt - timedelta(days=days_since_sat)).replace(hour=0, minute=0, second=0, microsecond=0)
        range_end = None
    elif range_start is None:
        range_start = range_end - timedelta(days=7)
    elif range_end is None:
        range_end = now_dt + timedelta(days=1)

    from_value = (range_start).strftime("%Y-%m-%d")
    to_value = (range_end - timedelta(days=1)).strftime("%Y-%m-%d") if range_end else now_dt.strftime("%Y-%m-%d")
    week_label = f"{range_start.strftime('%d/%m/%Y')} — {to_value[8:10]}/{to_value[5:7]}/{to_value[0:4]}"

    range_start_str = range_start.strftime("%Y-%m-%dT%H:%M:%S")
    PAGE_SIZE = 50
    try:
        page = max(1, int(request.args.get("page", "1")))
    except (TypeError, ValueError):
        page = 1
    offset = (page - 1) * PAGE_SIZE
    rows = []
    total_count = 0
    if SUPABASE_KEY:
        try:
            url = (f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                   f"?fired_at=gte.{range_start_str}&select=*&order=fired_at.desc"
                   f"&limit={PAGE_SIZE}&offset={offset}")
            if range_end:
                url += f"&fired_at=lt.{range_end.strftime('%Y-%m-%dT%H:%M:%S')}"
            r = requests.get(url, headers={**_sb_h(), "Prefer": "count=exact"}, timeout=10)
            if r.status_code in (200, 206):
                rows = r.json()
                cr = r.headers.get("Content-Range", "")  # مثلاً "0-49/213"
                if "/" in cr:
                    try: total_count = int(cr.split("/")[-1])
                    except ValueError: total_count = len(rows)
                else:
                    total_count = len(rows)
        except: pass
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    alert_ids = sorted({str(r.get("id","")) for r in rows if r.get("id")})
    alerts_map = {}
    if SUPABASE_KEY and alert_ids:
        try:
            r2 = requests.get(
                f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(alert_ids)})&select=*",
                headers=_sb_h(), timeout=10)
            if r2.status_code == 200:
                for row in r2.json():
                    alerts_map[str(row.get("id",""))] = row
        except Exception as e:
            print(f"[report_weekly] targeted alerts fetch error: {e}")
    rows = [r for r in rows if not alerts_map.get(str(r.get("id","")), {}).get("is_private")]
    rows_html = ""
    false_by_set = set()
    for idx, row in enumerate(rows, 1):
        aid       = str(row.get("id",""))
        tag       = row.get("alarm_tag","—")
        assignee  = row.get("assigned_to","") or "⏳ منتظر"
        fired     = row.get("fired_at","")[:16]
        false_by  = row.get("false_by","") or ""
        false_at  = row.get("false_at","")[:16] if row.get("false_at") else ""
        false_rsn = row.get("false_reason","") or ""
        is_active = row.get("is_active", True)
        if false_by:
            false_by_set.add(false_by)
        # تاریخچه کامل false
        false_history = row.get("false_history") or []
        if isinstance(false_history, str):
            try: false_history = json.loads(false_history)
            except: false_history = []
        alert     = alerts_map.get(aid, {})
        sym       = alert.get("symbol","") or row.get("symbol","") or ""
        tgt_raw   = alert.get("target_price",0) or row.get("target_price",0) or 0
        target    = fmt_price(float(tgt_raw), sym) if tgt_raw else "—"
        creator   = alert.get("created_by","") or row.get("created_by","") or "—"
        created   = str(alert.get("created_at",""))[:16]
        status_cls = "active" if is_active else "false"
        status_txt = "فعال" if is_active else f"False — {false_by}"
        cond_html = alert.get("condition", "") or row.get("condition", "")
        is_buy = (cond_html == "below")
        direction_icon = "📉" if cond_html == "above" else ("📈" if cond_html == "below" else "❓")
        dir_zone_label = "ناحیه سل" if cond_html == "above" else ("ناحیه بای" if cond_html == "below" else "")
        candle_cls = "candle-up" if is_buy else "candle-down"
        false_detail = ""
        false_action = ""
        if not is_active:
            if false_history:
                hist_items = ""
                for idx_h, h in enumerate(false_history, 1):
                    h_by  = h.get("by","")
                    h_at  = str(h.get("at",""))[:16]
                    h_rsn = h.get("reason","")
                    rsn_span = f'<span class="reason">{h_rsn}</span>' if h_rsn else ""
                    hist_items += f'<span class="hist-entry"><b>{idx_h}.</b> {h_by} — 🕐 {h_at}{(" · "+rsn_span) if rsn_span else ""}</span>'
                false_detail = f'<div class="false-detail false-history">{hist_items}</div>'
            else:
                false_detail = f'<div class="false-detail"><span>🕐 {false_at}</span>{("<span class=reason>"+false_rsn+"</span>") if false_rsn else ""}</div>'
        else:
            false_action = f'''<div class="false-action" data-aid="{aid}">
              <input type="text" class="false-reason-inp" placeholder="علت/کامنت (اختیاری)">
              <button type="button" class="false-btn" onclick="markFalse('{aid}', this)">❌ فالس کردن</button>
            </div>'''
        rows_html += f"""
        <div class="card card-{status_cls}" data-search="{(assignee + ' ' + sym + ' ' + tag + ' ' + creator + ' ' + false_by).lower()}" data-falseby="{false_by.lower()}" data-status="{status_cls}">
          <div class="card-num">{idx}</div>
          <div class="card-inner">
          <div class="card-glow"></div>
          <div class="card-header">
            <div class="card-icon">{direction_icon}</div>
            <div class="card-title">
              <span class="tag">{tag}</span>
              <span class="sym">{sym}{(' • ' + dir_zone_label) if dir_zone_label else ''}</span>
            </div>
            <span class="badge badge-{status_cls}">{status_txt}</span>
          </div>
          <div class="card-target">
            <span class="target-lbl">🎯 قیمت هدف</span>
            <span class="target-val">{target}</span>
          </div>
          <div class="card-body">
            <div class="info-grid">
              <div class="info-cell"><span class="lbl">📅 ثبت</span><span class="val">{created}</span></div>
              <div class="info-cell"><span class="lbl">⏰ فایر</span><span class="val">{fired}</span></div>
              <div class="info-cell"><span class="lbl">👤 سازنده</span><span class="val">{creator}</span></div>
              <div class="info-cell"><span class="lbl">🙋 مسئول</span><span class="val highlight">{assignee}</span></div>
            </div>
            {false_detail}
            {false_action}
          </div>
          <div class="card-rail">
            <div class="rail-dot {candle_cls}"></div>
            <div class="mini-candles">
              <span class="mc {candle_cls}" style="height:40%"></span>
              <span class="mc {candle_cls}" style="height:65%"></span>
              <span class="mc {candle_cls}" style="height:30%"></span>
              <span class="mc {candle_cls}" style="height:85%"></span>
              <span class="mc {candle_cls}" style="height:50%"></span>
            </div>
            <span class="rail-label">{dir_zone_label if dir_zone_label else ''}</span>
          </div>
          </div>
        </div>"""

    false_by_options = "".join(
        f'<option value="by:{name.lower()}">👤 {name}</option>' for name in sorted(false_by_set)
    )
    # چون rows الان فقط همین صفحه‌ست، برای آمار سربرگ (کل/فعال/False) به‌جای
    # فچ کامل، یه کوئری سبک count-only (بدون داده‌ی واقعی) برای کل بازه می‌زنیم.
    def _sb_count_assignments(extra_filter=""):
        if not SUPABASE_KEY: return 0
        try:
            u = (f"{SUPABASE_URL}/rest/v1/alarm_assignments"
                 f"?fired_at=gte.{range_start_str}&select=id&limit=1{extra_filter}")
            if range_end:
                u += f"&fired_at=lt.{range_end.strftime('%Y-%m-%dT%H:%M:%S')}"
            rc = requests.get(u, headers={**_sb_h(), "Prefer": "count=exact"}, timeout=10)
            cr = rc.headers.get("Content-Range", "")
            if "/" in cr:
                return int(cr.split("/")[-1])
        except Exception:
            pass
        return 0
    active_count = _sb_count_assignments("&is_active=eq.true")
    false_count  = total_count - active_count

    pagination_html = ""
    if total_pages > 1:
        base_url = f"/report/weekly?from={from_value}&to={to_value}"
        prev_link = f'<a href="{base_url}&page={page-1}" class="page-btn">← قبلی</a>' if page > 1 else '<span class="page-btn disabled">← قبلی</span>'
        next_link = f'<a href="{base_url}&page={page+1}" class="page-btn">بعدی →</a>' if page < total_pages else '<span class="page-btn disabled">بعدی →</span>'
        pagination_html = f'''<div class="pagination">
          {prev_link}
          <span class="page-info">صفحه {page} از {total_pages} ({total_count} مورد)</span>
          {next_link}
        </div>'''
    html = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>گزارش آلارم‌های تیم — {week_label}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  *{{box-sizing:border-box;margin:0;padding:0}}
  :root{{
    --bg:#070a12;--surface:#0d1424;--surface2:#10192e;
    --border:#1e293b;--border2:#2d3f5c;
    --text:#e2e8f0;--muted:#64748b;--subtle:#334155;
    --blue:#3b82f6;--blue-dim:#1e3a5f;--blue-glow:rgba(59,130,246,.35);
    --green:#22c55e;--green-dim:#052e16;--green-border:#166534;--green-glow:rgba(34,197,94,.3);
    --red:#ef4444;--red-dim:#1c0a0a;--red-border:#7f1d1d;--red-glow:rgba(239,68,68,.3);
    --gold:#fbbf24;--purple:#8b5cf6;
  }}
  [data-theme="light"]{{
    --bg:#eef2f7;--surface:#ffffff;--surface2:#ffffff;
    --border:#e2e8f0;--border2:#cbd5e1;
    --text:#1e293b;--muted:#64748b;--subtle:#94a3b8;
    --blue:#2563eb;--blue-dim:#dbeafe;--blue-glow:rgba(37,99,235,.15);
    --green:#16a34a;--green-dim:#dcfce7;--green-border:#86efac;--green-glow:rgba(22,163,74,.15);
    --red:#dc2626;--red-dim:#fee2e2;--red-border:#fca5a5;--red-glow:rgba(220,38,38,.15);
    --gold:#d97706;--purple:#7c3aed;
  }}
  html{{scroll-behavior:smooth}}
  body{{font-family:'Inter',Tahoma,sans-serif;background:var(--bg);color:var(--text);min-height:100vh;direction:rtl;transition:background .3s,color .3s;position:relative;overflow-x:hidden}}

  /* ── Animated background particles ── */
  .bg-grid{{position:fixed;inset:0;z-index:0;opacity:.4;pointer-events:none;
    background-image:linear-gradient(var(--border) 1px,transparent 1px),linear-gradient(90deg,var(--border) 1px,transparent 1px);
    background-size:40px 40px;mask-image:radial-gradient(ellipse 60% 50% at 50% 0%,#000 0%,transparent 100%)}}
  .bg-orb{{position:fixed;border-radius:50%;filter:blur(80px);pointer-events:none;z-index:0;opacity:.35;animation:float 16s ease-in-out infinite}}
  .bg-orb.o1{{width:300px;height:300px;background:var(--blue);top:-100px;right:-80px}}
  .bg-orb.o2{{width:260px;height:260px;background:var(--purple);top:30%;left:-100px;animation-delay:-5s}}
  .bg-orb.o3{{width:240px;height:240px;background:var(--green);bottom:0;right:10%;animation-delay:-10s}}
  @keyframes float{{0%,100%{{transform:translate(0,0) scale(1)}}50%{{transform:translate(30px,-40px) scale(1.1)}}}}

  /* ── Scroll progress ── */
  .scroll-bar{{position:fixed;top:0;right:0;left:0;height:4px;background:var(--border);z-index:999;overflow:hidden}}
  .scroll-bar-fill{{height:100%;width:0%;background:linear-gradient(90deg,var(--green),var(--blue),var(--purple));transition:width .08s;position:relative}}
  .scroll-bar-fill::after{{content:'';position:absolute;inset:0;
    background:repeating-linear-gradient(90deg,transparent 0 6px,rgba(255,255,255,.3) 6px 8px)}}
  .scroll-dot{{position:fixed;top:0;width:16px;height:16px;border-radius:50%;
    background:radial-gradient(circle,#fff,var(--blue));box-shadow:0 0 14px var(--blue);z-index:1000;
    transform:translate(-50%,-6px);transition:left .08s;display:flex;align-items:center;justify-content:center;font-size:9px}}

  /* ── Theme toggle ── */
  .theme-toggle{{position:fixed;top:14px;left:14px;z-index:998;width:44px;height:44px;border-radius:50%;border:1px solid var(--border2);background:var(--surface);color:var(--text);font-size:19px;cursor:pointer;display:flex;align-items:center;justify-content:center;box-shadow:0 4px 16px rgba(0,0,0,.25);transition:.25s}}
  .theme-toggle:hover{{transform:scale(1.1) rotate(15deg)}}

  /* ── Header ── */
  .hero{{padding:42px 20px 28px;text-align:center;position:relative;z-index:1}}
  .hero h1{{font-size:26px;font-weight:800;margin-bottom:8px;letter-spacing:-.5px;
    background:linear-gradient(135deg,var(--blue),var(--purple));-webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;
    animation:fadeDown .5s ease}}
  .hero .period{{font-size:13px;color:var(--muted);margin-bottom:24px;animation:fadeDown .6s ease}}

  /* ── Stats bar ── */
  .stats{{display:flex;gap:14px;justify-content:center;flex-wrap:wrap;animation:fadeUp .6s ease}}
  .stat{{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:14px 26px;text-align:center;min-width:90px;transition:.25s;position:relative;overflow:hidden}}
  .stat::before{{content:'';position:absolute;inset:0;background:linear-gradient(135deg,var(--blue-glow),transparent);opacity:0;transition:.3s}}
  .stat:hover::before{{opacity:1}}
  .stat:hover{{transform:translateY(-4px) scale(1.03);box-shadow:0 8px 24px var(--blue-glow)}}
  .stat .num{{font-size:26px;font-weight:800;position:relative}}
  .stat .lbl2{{font-size:11px;color:var(--muted);margin-top:4px;position:relative}}
  .stat.green .num{{color:var(--green)}}
  .stat.red .num{{color:var(--red)}}
  .stat.green:hover{{box-shadow:0 8px 24px var(--green-glow)}}
  .stat.red:hover{{box-shadow:0 8px 24px var(--red-glow)}}

  /* ── Date range nav ── */
  .range-nav{{max-width:680px;margin:0 auto;padding:14px 20px 4px;
    display:flex;flex-direction:column;gap:10px;align-items:center}}
  .range-presets{{display:flex;gap:8px;flex-wrap:wrap;justify-content:center}}
  .preset-btn{{padding:8px 18px;border-radius:12px;border:1px solid var(--border2);background:var(--surface);
    color:var(--muted);font-size:12px;font-weight:600;cursor:pointer;transition:.25s;font-family:'Inter',Tahoma,sans-serif}}
  .preset-btn:hover{{border-color:var(--blue);color:var(--blue);transform:translateY(-2px)}}
  .pagination{{display:flex;align-items:center;justify-content:center;gap:14px;margin:20px 0}}
  .page-btn{{padding:8px 18px;border-radius:12px;border:1px solid var(--border2);background:var(--surface);
    color:var(--blue);font-size:12px;font-weight:700;cursor:pointer;transition:.25s;text-decoration:none;
    font-family:'Inter',Tahoma,sans-serif}}
  .page-btn:hover{{border-color:var(--blue);transform:translateY(-2px)}}
  .page-btn.disabled{{color:var(--muted);opacity:.4;pointer-events:none}}
  .page-info{{font-size:12px;color:var(--muted);font-weight:600}}
  .range-inputs{{display:flex;gap:8px;align-items:center;flex-wrap:wrap;justify-content:center;
    background:var(--surface);border:1px solid var(--border2);border-radius:14px;padding:8px 14px;backdrop-filter:blur(10px)}}
  .range-inputs label{{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--muted);font-weight:600}}
  .range-inputs input[type="date"]{{padding:7px 10px;border-radius:10px;border:1px solid var(--border2);
    background:var(--surface2);color:var(--text);font-family:'Inter',Tahoma,sans-serif;font-size:12px;outline:none}}
  .range-inputs input[type="date"]:focus{{border-color:var(--blue)}}
  .range-go{{padding:8px 20px;border-radius:10px;border:none;cursor:pointer;font-size:12px;font-weight:700;
    background:linear-gradient(135deg,var(--blue),var(--purple));color:#fff;box-shadow:0 6px 16px var(--blue-glow)}}
  .range-go:hover{{transform:translateY(-1px)}}

  /* ── Search ── */
  .search-wrap{{max-width:680px;margin:0 auto;padding:0 20px 14px;position:relative;z-index:1}}
  .search-input{{width:100%;padding:13px 18px;border-radius:14px;border:1px solid var(--border2);
    background:var(--surface);color:var(--text);font-size:13px;font-family:'Inter',Tahoma,sans-serif;
    direction:rtl;outline:none;transition:.25s}}
  .search-input:focus{{border-color:var(--blue);box-shadow:0 0 0 3px var(--blue-glow)}}
  .search-input::placeholder{{color:var(--muted)}}
  .search-select{{width:100%;margin-top:10px;padding:13px 18px;border-radius:14px;border:1px solid var(--border2);
    background:var(--surface);color:var(--text);font-size:13px;font-family:'Inter',Tahoma,sans-serif;
    direction:rtl;outline:none;transition:.25s;cursor:pointer}}
  .search-select:focus{{border-color:var(--blue);box-shadow:0 0 0 3px var(--blue-glow)}}
  .search-count{{display:block;text-align:center;font-size:11px;color:var(--muted);margin-top:8px}}

  /* ── Cards ── */
  .list{{padding:10px 20px 40px;max-width:680px;margin:0 auto;position:relative;z-index:1}}
  .card{{border-radius:20px;border:1px solid var(--border);margin-bottom:18px;overflow:hidden;transition:.3s;
         opacity:0;transform:translateY(20px) scale(.98);animation:cardIn .5s ease forwards;
         background:var(--surface);position:relative;display:flex;align-items:stretch}}
  .card:hover{{transform:translateY(-5px) scale(1.01);box-shadow:0 16px 40px var(--blue-glow);border-color:var(--blue)}}
  .card-false:hover{{box-shadow:0 16px 40px var(--red-glow);border-color:var(--red)}}
  .card-glow{{position:absolute;top:-50%;left:-20%;width:60%;height:200%;
    background:radial-gradient(circle,var(--blue-glow),transparent 70%);pointer-events:none;opacity:.5}}
  .card-false .card-glow{{background:radial-gradient(circle,var(--red-glow),transparent 70%)}}
  .card-inner{{flex:1;min-width:0;position:relative}}
  .card-num{{flex:0 0 40px;display:flex;align-items:center;justify-content:center;
    background:var(--surface2);border-left:1px solid var(--border);
    color:var(--muted);font-family:'Inter',monospace;font-size:14px;font-weight:800}}

  .card-header{{display:flex;align-items:center;gap:12px;padding:18px 18px 14px;position:relative;z-index:1}}
  .card-icon{{font-size:26px;width:46px;height:46px;display:flex;align-items:center;justify-content:center;
    background:var(--surface2);border-radius:12px;border:1px solid var(--border)}}
  .card-title{{display:flex;flex-direction:column;gap:3px;flex:1}}
  .tag{{font-weight:800;font-size:17px;color:var(--blue)}}
  .sym{{font-size:11px;color:var(--muted);font-weight:600;letter-spacing:.5px}}
  .badge{{font-size:11px;font-weight:700;padding:6px 14px;border-radius:24px;white-space:nowrap}}
  .badge-active{{background:var(--green-dim);color:var(--green);border:1px solid var(--green-border)}}
  .badge-false{{background:var(--red-dim);color:var(--red);border:1px solid var(--red-border)}}

  /* ── Side rail: timeline dot + mini candles ── */
  .card-rail{{display:flex;align-items:center;justify-content:flex-start;gap:10px;padding:10px 18px;
    border-top:1px solid var(--border);background:var(--surface2)}}
  .rail-label{{font-size:11px;color:var(--muted);font-weight:600;margin-right:auto}}
  .rail-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0;box-shadow:0 0 8px currentColor}}
  .rail-dot.candle-up{{background:var(--green);color:var(--green)}}
  .rail-dot.candle-down{{background:var(--red);color:var(--red)}}
  .mini-candles{{display:flex;align-items:flex-end;gap:3px;height:26px}}
  .mc{{width:4px;border-radius:2px;opacity:.55;transition:.3s}}
  .mc.candle-up{{background:var(--green)}}
  .mc.candle-down{{background:var(--red)}}
  .card:hover .mc{{opacity:1;transform:scaleY(1.15)}}

  /* ── Target highlight bar ── */
  .card-target{{display:flex;justify-content:space-between;align-items:center;
    margin:0 18px 14px;padding:14px 18px;border-radius:14px;
    background:linear-gradient(135deg,var(--blue-dim),transparent);
    border:1px solid var(--border);position:relative;overflow:hidden}}
  .card-false .card-target{{background:linear-gradient(135deg,var(--red-dim),transparent)}}
  .target-lbl{{font-size:12px;color:var(--muted);font-weight:600}}
  .target-val{{font-family:'Inter',monospace;font-size:20px;font-weight:800;color:var(--gold);letter-spacing:.5px}}

  .card-body{{padding:0 18px 18px;position:relative;z-index:1}}
  .info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
  .info-cell{{display:flex;flex-direction:column;gap:4px;padding:10px 12px;background:var(--surface2);border-radius:10px;border:1px solid var(--border)}}
  .lbl{{font-size:11px;color:var(--muted);font-weight:500}}
  .val{{font-size:13px;color:var(--text);font-weight:600}}
  .val.highlight{{color:var(--blue);font-weight:800}}
  .false-detail{{margin-top:10px;padding:10px 14px;background:var(--red-dim);border-radius:10px;border:1px solid var(--red-border);display:flex;flex-wrap:wrap;gap:8px;align-items:center;font-size:12px;color:var(--red);font-weight:500}}
  .false-detail .reason{{font-style:italic}}
  .false-history{{flex-direction:column;gap:4px}}
  .hist-entry{{display:block;font-size:11px;color:var(--red);padding:2px 0}}
  .false-action{{margin-top:10px;display:flex;gap:8px;align-items:center}}
  .false-reason-inp{{flex:1;background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:8px 10px;font-size:12px;color:var(--text);font-family:inherit}}
  .false-reason-inp:focus{{outline:none;border-color:var(--red)}}
  .false-btn{{background:var(--red-dim);color:var(--red);border:1px solid var(--red-border);border-radius:8px;padding:8px 14px;font-size:12px;font-weight:700;cursor:pointer;white-space:nowrap;transition:.15s}}
  .false-btn:hover{{background:var(--red);color:#fff}}
  .false-btn:disabled{{opacity:.5;cursor:default}}

  /* ── Empty ── */
  .empty{{text-align:center;padding:90px 20px;color:var(--muted);animation:fadeUp .5s ease;position:relative;z-index:1}}
  .empty .icon{{font-size:54px;margin-bottom:14px}}

  /* ── Footer ── */
  .footer{{text-align:center;padding:28px;font-size:11px;color:var(--subtle);position:relative;z-index:1}}

  /* ── Animations ── */
  @keyframes fadeDown{{from{{opacity:0;transform:translateY(-12px)}}to{{opacity:1;transform:translateY(0)}}}}
  @keyframes fadeUp{{from{{opacity:0;transform:translateY(12px)}}to{{opacity:1;transform:translateY(0)}}}}
  @keyframes cardIn{{to{{opacity:1;transform:translateY(0) scale(1)}}}}
</style>
</head>
<body data-theme="dark">
<div class="bg-grid"></div>
<div class="bg-orb o1"></div>
<div class="bg-orb o2"></div>
<div class="bg-orb o3"></div>
<div class="scroll-bar"><div class="scroll-bar-fill" id="scrollFill"></div></div>
<div class="scroll-dot" id="scrollDot">📈</div>
<button class="theme-toggle" id="themeToggle" onclick="toggleTheme()">🌙</button>
<a href="/report/export?from={from_value}&to={to_value}" target="_blank" rel="noopener" style="position:fixed;top:16px;left:64px;z-index:50;background:var(--surface,#1a1f2e);border:1px solid var(--border,#2a3142);border-radius:10px;padding:8px 14px;font-size:13px;color:inherit;text-decoration:none;display:flex;align-items:center;gap:6px">📥 دانلود دیتا</a>
<div class="hero">
  <h1>📋 گزارش آلارم‌های تیم</h1>
  <div class="period">{week_label}</div>
  <div class="stats">
    <div class="stat"><div class="num">{total_count}</div><div class="lbl2">کل آلارم</div></div>
    <div class="stat green"><div class="num">{active_count}</div><div class="lbl2">فعال</div></div>
    <div class="stat red"><div class="num">{false_count}</div><div class="lbl2">False شده</div></div>
  </div>
</div>
<div style="text-align:center;margin:-8px 0 14px">
  <a href="/report/export?from={from_value}&to={to_value}" target="_blank" rel="noopener" class="preset-btn" style="text-decoration:none;display:inline-block">📥 دانلود دیتا</a>
</div>
<form class="range-nav" id="rangeNav" method="get" action="/report/weekly">
  <div class="range-presets">
    <button type="button" class="preset-btn" onclick="setRange(0)">📅 این هفته</button>
    <button type="button" class="preset-btn" onclick="setRange(1)">📅 هفته قبل</button>
    <button type="button" class="preset-btn" onclick="setRange(2)">🗓️ این ماه</button>
    <button type="button" class="preset-btn" onclick="setRange(3)">🗓️ ماه قبل</button>
  </div>
  <div class="range-inputs">
    <label>از <input type="date" name="from" id="fromInput" value="{from_value}"></label>
    <label>تا <input type="date" name="to" id="toInput" value="{to_value}"></label>
    <button type="submit" class="range-go">نمایش 🔍</button>
  </div>
</form>
<div class="search-wrap">
  <input type="text" id="searchBox" class="search-input" placeholder="🔍 جستجو بر اساس مسئول، نماد، تگ یا سازنده..." oninput="filterCards()">
  <select id="filterSelect" class="search-select" onchange="filterCards()">
    <option value="">📂 همه آلارم‌ها</option>
    <optgroup label="وضعیت">
      <option value="status:active">✅ فقط فعال</option>
      <option value="status:false">❌ فقط False شده</option>
    </optgroup>
    <optgroup label="False شده توسط">
      {false_by_options}
    </optgroup>
  </select>
  <span class="search-count" id="searchCount"></span>
</div>
<div class="list" id="cardList">
  {'<div class="empty"><div class="icon">📭</div>آلارمی ثبت نشده</div>' if not rows else rows_html}
  <div class="empty" id="noResults" style="display:none"><div class="icon">🔍</div>چیزی پیدا نشد</div>
</div>
{pagination_html}
<div class="footer">آخرین بروزرسانی: {now_dt.strftime('%H:%M — %d/%m/%Y')}</div>
<script>
  // فالس کردن یه آلارم مستقیم از همین صفحه (با کامنت اختیاری)
  async function markFalse(aid, btnEl) {{
    const wrap = btnEl.closest('.false-action');
    const reasonInp = wrap ? wrap.querySelector('.false-reason-inp') : null;
    const reason = reasonInp ? reasonInp.value.trim() : '';
    let name = localStorage.getItem('trader_name') || '';
    let pin = localStorage.getItem('trader_pin') || '';
    if (!name) {{
      name = prompt('نام شما (همون نامی که تو سایت آلارم استفاده می‌کنید):', '') || '';
      if (!name) return;
    }}
    if (!confirm('این آلارم به عنوان False ثبت بشه؟')) return;
    btnEl.disabled = true;
    btnEl.textContent = '⏳ ...';
    try {{
      const r = await fetch(`/api/alarm-assignments/${{aid}}/false`, {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{name, pin, reason}})
      }});
      const d = await r.json();
      if (d.ok) {{
        location.reload();
      }} else {{
        alert('❌ ' + (d.error || 'خطا در ثبت'));
        btnEl.disabled = false;
        btnEl.textContent = '❌ فالس کردن';
      }}
    }} catch (e) {{
      alert('❌ خطا در ارتباط با سرور');
      btnEl.disabled = false;
      btnEl.textContent = '❌ فالس کردن';
    }}
  }}

  // miyanbor‌های بازه تاریخ
  function _fmtDate(d) {{
    const y = d.getFullYear(), m = String(d.getMonth()+1).padStart(2,'0'), day = String(d.getDate()).padStart(2,'0');
    return `${{y}}-${{m}}-${{day}}`;
  }}
  function setRange(type) {{
    const now = new Date();
    let from, to;
    if (type === 0) {{ // این هفته: شنبه تا امروز
      const diffToSat = (now.getDay() - 6 + 7) % 7;
      from = new Date(now); from.setDate(now.getDate() - diffToSat);
      to = now;
    }} else if (type === 1) {{ // هفته قبل
      const diffToSat = (now.getDay() - 6 + 7) % 7;
      const thisWeekStart = new Date(now); thisWeekStart.setDate(now.getDate() - diffToSat);
      to = new Date(thisWeekStart); to.setDate(thisWeekStart.getDate() - 1);
      from = new Date(thisWeekStart); from.setDate(thisWeekStart.getDate() - 7);
    }} else if (type === 2) {{ // این ماه
      from = new Date(now.getFullYear(), now.getMonth(), 1);
      to = now;
    }} else if (type === 3) {{ // ماه قبل
      from = new Date(now.getFullYear(), now.getMonth() - 1, 1);
      to = new Date(now.getFullYear(), now.getMonth(), 0);
    }}
    document.getElementById('fromInput').value = _fmtDate(from);
    document.getElementById('toInput').value = _fmtDate(to);
    document.getElementById('rangeNav').submit();
  }}
  // scroll progress — chart line style
  window.addEventListener('scroll', () => {{
    const h = document.documentElement;
    const pct = (h.scrollTop / (h.scrollHeight - h.clientHeight)) * 100;
    document.getElementById('scrollFill').style.width = pct + '%';
    document.getElementById('scrollDot').style.left = pct + '%';
  }});
  // stagger card animations
  document.querySelectorAll('.card').forEach((c,i) => {{
    c.style.animationDelay = (i * 0.06) + 's';
  }});
  // search/filter
  function filterCards() {{
    const q = document.getElementById('searchBox').value.trim().toLowerCase();
    const sel = document.getElementById('filterSelect').value;
    const cards = document.querySelectorAll('#cardList .card');
    let visible = 0;
    cards.forEach(c => {{
      let match = !q || (c.dataset.search || '').includes(q);
      if (match && sel) {{
        if (sel.startsWith('status:')) {{
          match = c.dataset.status === sel.slice(7);
        }} else if (sel.startsWith('by:')) {{
          match = c.dataset.falseby === sel.slice(3);
        }}
      }}
      c.style.display = match ? '' : 'none';
      if (match) visible++;
    }});
    document.getElementById('noResults').style.display = (visible === 0) ? '' : 'none';
    document.getElementById('searchCount').textContent = (q || sel) ? `${{visible}} نتیجه پیدا شد` : '';
  }}
  // theme toggle with localStorage
  function toggleTheme() {{
    const body = document.body;
    const isDark = body.getAttribute('data-theme') === 'dark';
    body.setAttribute('data-theme', isDark ? 'light' : 'dark');
    document.getElementById('themeToggle').textContent = isDark ? '☀️' : '🌙';
    try {{ localStorage.setItem('reportTheme', isDark ? 'light' : 'dark'); }} catch(e) {{}}
  }}
  try {{
    const saved = localStorage.getItem('reportTheme');
    if (saved) {{
      document.body.setAttribute('data-theme', saved);
      document.getElementById('themeToggle').textContent = saved === 'dark' ? '🌙' : '☀️';
    }}
  }} catch(e) {{}}
</script>
</body></html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/report/export")
def report_export_html():
    """صفحه‌ی دانلود دیتا — بازه‌ی تاریخ + فیلترهای ترکیبی + دکمه‌ی خروجی اکسل"""
    assignee_rows = "".join(
        f'<label class="chk-row"><input type="checkbox" value="{m}" class="assigneeChk"> {m}</label>'
        for m in TEAM_MEMBERS
    )
    today_str = datetime.now(TEHRAN).strftime("%Y-%m-%d")
    month_ago_str = (datetime.now(TEHRAN) - timedelta(days=30)).strftime("%Y-%m-%d")
    default_from = request.args.get("from", "") or month_ago_str
    default_to = request.args.get("to", "") or today_str
    html = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>دانلود دیتا — گزارش آلارم‌ها</title>
<style>
  :root{{--bg:#070a12;--surface:#0d1424;--surface2:#10192e;--border:#1e293b;--border2:#2d3f5c;
    --text:#e2e8f0;--muted:#64748b;--blue:#3b82f6;--green:#22c55e;--red:#ef4444}}
  *{{box-sizing:border-box}}
  body{{font-family:'Inter',Tahoma,sans-serif;background:var(--bg);color:var(--text);min-height:100vh;
    direction:rtl;margin:0;padding:24px 16px;display:flex;justify-content:center}}
  .card{{max-width:560px;width:100%;background:var(--surface);border:1px solid var(--border2);
    border-radius:18px;padding:26px 22px}}
  h1{{font-size:19px;margin:0 0 4px;font-weight:800}}
  .sub{{font-size:12px;color:var(--muted);margin-bottom:22px}}
  .fld{{margin-bottom:16px}}
  label{{display:block;font-size:12px;color:var(--muted);font-weight:600;margin-bottom:6px}}
  input[type=date],input[type=text]{{width:100%;padding:10px 12px;border-radius:10px;
    border:1px solid var(--border2);background:var(--surface2);color:var(--text);font-size:13px;
    font-family:inherit;outline:none}}
  input:focus{{border-color:var(--blue)}}
  .row2{{display:flex;gap:10px}}
  .row2 .fld{{flex:1}}
  .chk-group{{display:flex;flex-wrap:wrap;gap:10px;background:var(--surface2);
    border:1px solid var(--border2);border-radius:10px;padding:12px 14px}}
  .chk-row{{display:flex;align-items:center;gap:6px;font-size:13px;margin:0;font-weight:400;color:var(--text)}}
  .chk-row input{{width:16px;height:16px}}
  .dl-btn{{width:100%;padding:13px;border:none;border-radius:12px;background:linear-gradient(135deg,var(--blue),#2563eb);
    color:#fff;font-size:14px;font-weight:800;cursor:pointer;margin-top:8px;transition:.2s}}
  .dl-btn:hover{{filter:brightness(1.1)}}
  .hint{{font-size:11px;color:var(--muted);margin-top:6px;line-height:1.7}}
  a.back{{color:var(--blue);font-size:12px;text-decoration:none;display:inline-block;margin-bottom:14px}}
  .modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:100;
    align-items:center;justify-content:center;padding:16px}}
  .modal-overlay.show{{display:flex}}
  .modal-box{{max-width:420px;width:100%;max-height:85vh;overflow-y:auto;background:var(--surface);
    border:1px solid var(--border2);border-radius:16px;padding:22px}}
  .modal-box h2{{font-size:16px;margin:0 0 4px;font-weight:800}}
  .modal-sub{{font-size:11px;color:var(--muted);margin-bottom:14px}}
  .col-actions{{display:flex;gap:8px;margin-bottom:12px}}
  .col-actions button{{flex:1;padding:7px;border-radius:8px;border:1px solid var(--border2);
    background:var(--surface2);color:var(--text);font-size:11px;cursor:pointer;font-family:inherit}}
  .col-list{{display:flex;flex-direction:column;gap:2px;background:var(--surface2);
    border:1px solid var(--border2);border-radius:10px;padding:8px 12px;margin-bottom:16px}}
  .col-list label{{display:flex;align-items:center;gap:8px;font-size:13px;font-weight:400;
    color:var(--text);margin:0;padding:6px 0}}
  .col-list input{{width:16px;height:16px}}
  .modal-btns{{display:flex;gap:8px}}
  .modal-btns button{{flex:1;padding:11px;border-radius:10px;border:none;font-size:13px;font-weight:700;
    cursor:pointer;font-family:inherit}}
  .btn-cancel{{background:var(--surface2);color:var(--text);border:1px solid var(--border2) !important}}
  .btn-confirm{{background:linear-gradient(135deg,var(--blue),#2563eb);color:#fff}}
</style>
</head>
<body>
<div class="card">
  <a class="back" href="/report/weekly">→ برگشت به گزارش هفتگی</a>
  <h1>📥 دانلود دیتا</h1>
  <div class="sub">یه بازه و فیلتر انتخاب کن، فایل اکسل (xlsx) دانلود کن.</div>

  <div class="row2">
    <div class="fld"><label>از تاریخ</label><input type="date" id="expFrom" value="{default_from}"></div>
    <div class="fld"><label>تا تاریخ</label><input type="date" id="expTo" value="{default_to}"></div>
  </div>

  <div class="fld">
    <label>وضعیت (خالی = همه)</label>
    <div class="chk-group">
      <label class="chk-row"><input type="checkbox" value="fired" class="statusChk"> ✅ فایرشده</label>
      <label class="chk-row"><input type="checkbox" value="false" class="statusChk"> ❌ False شده</label>
      <label class="chk-row"><input type="checkbox" value="expired" class="statusChk"> ⌛ منقضی‌شده</label>
      <label class="chk-row"><input type="checkbox" value="active" class="statusChk"> 🟢 فعال</label>
    </div>
  </div>

  <div class="fld">
    <label>مسئول (خالی = همه)</label>
    <div class="chk-group">{assignee_rows}</div>
  </div>

  <div class="row2">
    <div class="fld">
      <label>ثبت‌کننده (اختیاری — چندتا با کاما)</label>
      <input type="text" id="expCreator" placeholder="مثلاً مسعود, علی">
    </div>
    <div class="fld">
      <label>نماد/ارز (اختیاری — چندتا با کاما)</label>
      <input type="text" id="expSymbol" placeholder="مثلاً XAUUSD, EURUSD" dir="ltr">
    </div>
  </div>

  <button class="dl-btn" id="dlBtn" onclick="openColumnModal()">📥 دانلود اکسل</button>
  <div class="hint">بازه‌ی تاریخ بر اساس تاریخ فایر شدن آلارم اعمال می‌شه. فیلترها با هم AND می‌شن.<br>
  ⓘ آلارم‌های فایرشده‌ی قبل از ۱۱ ژوئن ۲۰۲۶ («فایرشده (قدیمی)» تو ستون وضعیت) مربوط به قبل از راه‌اندازی سیستم «مسئول آلارم» هستن — برای این‌ها ستون‌های «مسئول» و «علت/تاریخ False» همیشه خالیه، چون اون‌موقع اصلاً همچین چیزی ثبت نمی‌شد. اگه فیلتر «مسئول» رو فعال کنید، این آلارم‌های قدیمی (چون مسئول ندارن) تو خروجی نمیان.</div>
</div>

<div class="modal-overlay" id="columnModal">
  <div class="modal-box">
    <h2>ستون‌های اکسل</h2>
    <div class="modal-sub">هر ستونی رو نمی‌خواید تو خروجی باشه، تیکش رو بردارید.</div>
    <div class="col-actions">
      <button type="button" onclick="setAllColumns(true)">✅ انتخاب همه</button>
      <button type="button" onclick="setAllColumns(false)">❌ پاک کردن همه</button>
    </div>
    <div class="col-list">
      <label><input type="checkbox" class="colChk" value="symbol" checked> نماد</label>
      <label><input type="checkbox" class="colChk" value="alarm_tag" checked> هشتک</label>
      <label><input type="checkbox" class="colChk" value="direction" checked> جهت</label>
      <label><input type="checkbox" class="colChk" value="target_price" checked> قیمت هدف</label>
      <label><input type="checkbox" class="colChk" value="fired_price" checked> قیمت فایر</label>
      <label><input type="checkbox" class="colChk" value="status" checked> وضعیت</label>
      <label><input type="checkbox" class="colChk" value="created_by" checked> ثبت‌کننده</label>
      <label><input type="checkbox" class="colChk" value="assigned_to" checked> مسئول</label>
      <label><input type="checkbox" class="colChk" value="created_date" checked> تاریخ ثبت</label>
      <label><input type="checkbox" class="colChk" value="created_time" checked> ساعت ثبت</label>
      <label><input type="checkbox" class="colChk" value="fired_date" checked> تاریخ فایر</label>
      <label><input type="checkbox" class="colChk" value="fired_time" checked> ساعت فایر</label>
      <label><input type="checkbox" class="colChk" value="false_reason" checked> علت False</label>
      <label><input type="checkbox" class="colChk" value="false_date" checked> تاریخ False</label>
      <label><input type="checkbox" class="colChk" value="false_time" checked> ساعت False</label>
      <label><input type="checkbox" class="colChk" value="expired_date" checked> تاریخ انقضا</label>
      <label><input type="checkbox" class="colChk" value="expired_time" checked> ساعت انقضا</label>
      <label><input type="checkbox" class="colChk" value="is_private" checked> خصوصی</label>
      <label><input type="checkbox" class="colChk" value="comment" checked> کامنت</label>
    </div>
    <div class="modal-btns">
      <button class="btn-cancel" onclick="closeColumnModal()">انصراف</button>
      <button class="btn-confirm" onclick="doExport()">📥 دانلود</button>
    </div>
  </div>
</div>
<script>
function openColumnModal() {{
  document.getElementById('columnModal').classList.add('show');
}}
function closeColumnModal() {{
  document.getElementById('columnModal').classList.remove('show');
}}
function setAllColumns(val) {{
  document.querySelectorAll('.colChk').forEach(c => c.checked = val);
}}
function doExport() {{
  const from = document.getElementById('expFrom').value;
  const to = document.getElementById('expTo').value;
  const statuses = Array.from(document.querySelectorAll('.statusChk:checked')).map(c => c.value).join(',');
  const assignees = Array.from(document.querySelectorAll('.assigneeChk:checked')).map(c => c.value).join(',');
  const creator = document.getElementById('expCreator').value.trim();
  const symbol = document.getElementById('expSymbol').value.trim();
  const columns = Array.from(document.querySelectorAll('.colChk:checked')).map(c => c.value).join(',');
  const params = new URLSearchParams();
  if (from) params.set('from', from);
  if (to) params.set('to', to);
  if (statuses) params.set('statuses', statuses);
  if (assignees) params.set('assignees', assignees);
  if (creator) params.set('creator', creator);
  if (symbol) params.set('symbol', symbol);
  if (columns) params.set('columns', columns);
  window.location.href = '/api/report/export.xlsx?' + params.toString();
  closeColumnModal();
}}
</script>
</body></html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/api/report/export.xlsx")
def report_export_xlsx():
    """
    خروجی اکسل واقعی (xlsx) — هدفمند و محدود به بازه/فیلتر انتخاب‌شده، نه کل جدول.
    وضعیت «False شده» از جدول alarm_assignments (is_active=false) تشخیص داده می‌شه،
    نه از خود alerts — چون False یه ویژگی از assignment هست، نه خود آلارم.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return jsonify({"ok": False, "error": "پکیج openpyxl روی سرور نصب نیست — باید به requirements.txt اضافه بشه."}), 500

    from_str = request.args.get("from", "")
    to_str = request.args.get("to", "")
    statuses = [s.strip() for s in request.args.get("statuses", "").split(",") if s.strip()]
    assignees = [a.strip() for a in request.args.get("assignees", "").split(",") if a.strip()]
    creators = [c.strip() for c in request.args.get("creator", "").split(",") if c.strip()]
    symbols = [s.strip().upper() for s in request.args.get("symbol", "").split(",") if s.strip()]

    want_all = not statuses
    want_fired = want_all or "fired" in statuses
    want_false = want_all or "false" in statuses
    want_expired = want_all or "expired" in statuses
    want_active = want_all or "active" in statuses

    if not SUPABASE_KEY:
        return jsonify({"ok": False, "error": "Supabase تنظیم نشده"}), 500

    now_dt = datetime.now(TEHRAN)
    try:
        range_start = TEHRAN.localize(datetime.strptime(from_str, "%Y-%m-%d")) if from_str else (now_dt - timedelta(days=30))
    except Exception:
        range_start = now_dt - timedelta(days=30)
    try:
        range_end = TEHRAN.localize(datetime.strptime(to_str, "%Y-%m-%d")) + timedelta(days=1) if to_str else now_dt + timedelta(days=1)
    except Exception:
        range_end = now_dt + timedelta(days=1)
    range_start_str = range_start.strftime("%Y-%m-%dT%H:%M:%S")
    range_end_str = range_end.strftime("%Y-%m-%dT%H:%M:%S")

    assignments_map = {}  # alert_id -> {assigned_to, is_active, false_by, false_at, false_reason, alarm_tag}

    # جدول alarm_assignments (مسئول/False) از این تاریخ به بعد ساخته و پر شده — آلارم‌های
    # فایرشده‌ی قدیمی‌تر از این، فقط تو خود alerts (status=fired) هستن و مسئول/False ندارن.
    ASSIGN_SYSTEM_START = TEHRAN.localize(datetime(2026, 6, 11, 0, 0, 0))
    assign_start_str = ASSIGN_SYSTEM_START.strftime("%Y-%m-%dT%H:%M:%S")

    def _fetch_assignments(ids_list=None, assignee_list=None):
        base_select = "select=id,assigned_to,is_active,false_by,false_at,false_reason,alarm_tag"
        if ids_list:
            # چون ممکنه لیست id خیلی طولانی باشه، تکه‌تکه (هر بار ۲۰۰ تا) می‌گیریم
            out = []
            ids_list = list(ids_list)
            for i in range(0, len(ids_list), 200):
                chunk = ids_list[i:i+200]
                u = (f"{SUPABASE_URL}/rest/v1/alarm_assignments?{base_select}"
                     f"&id=in.({','.join(chunk)})&limit={len(chunk)}")
                r = requests.get(u, headers=_sb_h(), timeout=15)
                if r.status_code == 200:
                    out.extend(r.json())
            return out
        # حالت بازه‌ای: صفحه‌به‌صفحه همه‌ی رکوردهای این بازه رو می‌گیریم تا چیزی جا نمونه
        out = []
        page_size = 1000
        offset = 0
        while True:
            u = (f"{SUPABASE_URL}/rest/v1/alarm_assignments?{base_select}"
                 f"&fired_at=gte.{range_start_str}&fired_at=lt.{range_end_str}"
                 f"&order=fired_at.asc&limit={page_size}&offset={offset}")
            if assignee_list:
                u += f"&assigned_to=in.({','.join(assignee_list)})"
            r = requests.get(u, headers=_sb_h(), timeout=15)
            if r.status_code != 200:
                break
            batch = r.json()
            out.extend(batch)
            if len(batch) < page_size:
                break
            offset += page_size
        return out

    try:
        # دقیقا عین منطق گزارش هفتگی: همیشه از alarm_assignments بر اساس fired_at شروع می‌کنیم،
        # بعد alertهای متناظرشون رو می‌گیریم. چه فیلتر مسئول زده باشه چه نه.
        assign_rows = _fetch_assignments(assignee_list=assignees if assignees else None)
        for x in assign_rows:
            assignments_map[str(x.get("id",""))] = x
        ids = sorted(assignments_map.keys())
        if not ids:
            rows = []
        else:
            rows = []
            for i in range(0, len(ids), 200):
                chunk = ids[i:i+200]
                u2 = f"{SUPABASE_URL}/rest/v1/alerts?id=in.({','.join(chunk)})&select=*&limit={len(chunk)}"
                if creators:
                    u2 += f"&created_by=in.({','.join(creators)})"
                if symbols:
                    u2 += f"&symbol=in.({','.join(symbols)})"
                r2 = requests.get(u2, headers=_sb_h(), timeout=15)
                if r2.status_code == 200:
                    rows.extend(r2.json())
            rows = [x for x in rows if x.get("id") != "__config__"]

        # منقضی‌شده‌ها هیچ‌وقت فایر نشدن، پس تو alarm_assignments نیستن — مستقیم از alerts
        # بر اساس همون بازه‌ی expired_at میاریم. (چون آلارم منقضی‌شده مسئول نداره، فقط وقتی
        # فیلتر مسئول زده نشده اضافه‌شون می‌کنیم.)
        if want_expired and not assignees:
            ue = (f"{SUPABASE_URL}/rest/v1/alerts"
                  f"?expired_at=gte.{range_start_str}&expired_at=lt.{range_end_str}"
                  f"&status=eq.expired&select=*&limit=5000")
            if creators:
                ue += f"&created_by=in.({','.join(creators)})"
            if symbols:
                ue += f"&symbol=in.({','.join(symbols)})"
            re_exp = requests.get(ue, headers=_sb_h(), timeout=15)
            expired_rows = re_exp.json() if re_exp.status_code == 200 else []
            existing_ids = {str(x.get("id","")) for x in rows}
            rows += [x for x in expired_rows if str(x.get("id","")) not in existing_ids]

        # ── دیتای قدیمی (قبل از ساخته‌شدن alarm_assignments) — اون‌موقع هنوز مفهوم
        # «مسئول»/False اصلاً نبود، پس این آلارم‌ها هیچ‌وقت تو alarm_assignments ثبت
        # نمی‌شن. اگه بخشی از بازه‌ی انتخابی قبل از اون تاریخ باشه، مستقیم از خود
        # alerts (status=fired) میاریمشون — بدون فیلتر مسئول (چون اصلاً معنی نداره). ──
        if (want_fired or want_false or want_all) and not assignees and range_start < ASSIGN_SYSTEM_START:
            legacy_end = min(range_end, ASSIGN_SYSTEM_START)
            ul = (f"{SUPABASE_URL}/rest/v1/alerts"
                  f"?fired_at=gte.{range_start_str}&fired_at=lt.{legacy_end.strftime('%Y-%m-%dT%H:%M:%S')}"
                  f"&status=eq.fired&select=*&limit=5000")
            if creators:
                ul += f"&created_by=in.({','.join(creators)})"
            if symbols:
                ul += f"&symbol=in.({','.join(symbols)})"
            r_legacy = requests.get(ul, headers=_sb_h(), timeout=15)
            legacy_rows = r_legacy.json() if r_legacy.status_code == 200 else []
            existing_ids = {str(x.get("id","")) for x in rows}
            for x in legacy_rows:
                if str(x.get("id","")) not in existing_ids:
                    x["_legacy"] = True
                    rows.append(x)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    # فیلتر نهایی — دقیقا مثل استت‌بار گزارش هفتگی (کل آلارم / فعال / False شده):
    # «فایرشده» یعنی همه‌ی آلارم‌های فایرشده تو این بازه (چه فعال چه False)،
    # «فعال» فقط زیرمجموعه‌ی فالس‌نشده، «False شده» فقط زیرمجموعه‌ی فالس‌شده.
    final_rows = []
    for a in rows:
        st = a.get("status", "")
        asg = assignments_map.get(str(a.get("id","")), {})
        if a.get("_legacy") and st == "fired":
            # آلارم قدیمی از قبل سیستم مسئول — مفهوم مسئول/False براش وجود نداره
            if want_fired or want_false or want_all:
                final_rows.append((a, {}, "فایرشده (قدیمی)"))
        elif st == "fired":
            is_false = asg.get("is_active") is False
            if is_false:
                if want_fired or want_false:
                    final_rows.append((a, asg, "False شده"))
            else:
                if want_fired or want_active:
                    final_rows.append((a, asg, "فعال"))
        elif st == "expired" and want_expired:
            final_rows.append((a, asg, "منقضی‌شده"))

    # مرتب‌سازی بر اساس تاریخ فایر (برای منقضی‌شده‌ها که فایر نشدن، بر اساس تاریخ انقضا)
    final_rows.sort(key=lambda item: item[0].get("fired_at") or item[0].get("expired_at") or "")

    def _split_dt(s):
        """'2026-08-01 10:00:00' یا '2026-08-01T10:00:00' رو به (تاریخ, ساعت) جدا می‌کنه"""
        s = (s or "").strip()
        if not s:
            return "", ""
        s = s.replace("T", " ")
        parts = s.split(" ", 1)
        if len(parts) == 2:
            return parts[0], parts[1][:8]
        return parts[0], ""

    # هر ستون یه کلید داره — پارامتر columns تو URL تعیین می‌کنه کدوما تو خروجی باشن
    COLUMN_DEFS = [
        ("symbol",       "نماد",         10),
        ("alarm_tag",    "هشتک",         9),
        ("direction",    "جهت",          6),
        ("target_price", "قیمت هدف",     10),
        ("fired_price",  "قیمت فایر",    10),
        ("status",       "وضعیت",        11),
        ("created_by",   "ثبت‌کننده",    12),
        ("assigned_to",  "مسئول",        9),
        ("created_date", "تاریخ ثبت",    11),
        ("created_time", "ساعت ثبت",     9),
        ("fired_date",   "تاریخ فایر",   11),
        ("fired_time",   "ساعت فایر",    9),
        ("false_reason", "علت False",    22),
        ("false_date",   "تاریخ False",  11),
        ("false_time",   "ساعت False",   9),
        ("expired_date", "تاریخ انقضا",  11),
        ("expired_time", "ساعت انقضا",   9),
        ("is_private",   "خصوصی",        7),
        ("comment",      "کامنت",        26),
    ]
    requested_cols = [c.strip() for c in request.args.get("columns", "").split(",") if c.strip()]
    all_keys = [k for k, _, _ in COLUMN_DEFS]
    selected_keys = [k for k in all_keys if (not requested_cols or k in requested_cols)]
    label_by_key = {k: lbl for k, lbl, _ in COLUMN_DEFS}
    width_by_key = {k: w for k, _, w in COLUMN_DEFS}

    wb = Workbook()
    ws = wb.active
    ws.title = "آلارم‌ها"
    ws.sheet_view.rightToLeft = False
    ws.append([label_by_key[k] for k in selected_keys])
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    direction_col = selected_keys.index("direction") + 1 if "direction" in selected_keys else None
    buy_fill   = PatternFill("solid", fgColor="D9F2E3")  # سبز کم‌رنگ
    sell_fill  = PatternFill("solid", fgColor="FBDEDE")  # قرمز کم‌رنگ
    zebra_fill = PatternFill("solid", fgColor="F5F5F5")  # رنگ آروم برای ردیف‌های یکی‌درمیان

    for a, asg, status_lbl in final_rows:
        created_d, created_t = _split_dt(a.get("created_at",""))
        fired_d, fired_t = _split_dt(a.get("fired_at",""))
        false_d, false_t = _split_dt(asg.get("false_at",""))
        expired_d, expired_t = _split_dt(a.get("expired_at",""))
        direction_lbl = "بای" if a.get("condition") == "below" else "سل"
        values_by_key = {
            "symbol":       a.get("symbol",""),
            "alarm_tag":    asg.get("alarm_tag","") or "",
            "direction":    direction_lbl,
            "target_price": a.get("target_price",""),
            "fired_price":  a.get("fired_price","") or "",
            "status":       status_lbl,
            "created_by":   a.get("created_by",""),
            "assigned_to":  asg.get("assigned_to","") or "",
            "created_date": created_d, "created_time": created_t,
            "fired_date":   fired_d,   "fired_time":   fired_t,
            "false_reason": asg.get("false_reason","") or "",
            "false_date":   false_d,   "false_time":   false_t,
            "expired_date": expired_d, "expired_time": expired_t,
            "is_private":   "بله" if a.get("is_private") else "",
            "comment":      a.get("comment",""),
        }
        ws.append([values_by_key[k] for k in selected_keys])
        r = ws.max_row
        ws.row_dimensions[r].height = 24
        # ردیف‌های یکی‌درمیون یه رنگ آروم بگیرن
        if (r % 2) == 0:
            for cell in ws[r]:
                cell.fill = zebra_fill
        # فقط سلول جهت (بای/سل) رنگ خودشو بگیره
        if direction_col:
            dc = ws.cell(row=r, column=direction_col)
            dc.fill = buy_fill if direction_lbl == "بای" else sell_fill

    from openpyxl.utils import get_column_letter
    for i, k in enumerate(selected_keys, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_by_key[k]

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"alarms_export_{from_str or 'all'}_{to_str or 'all'}.xlsx"
    return buf.read(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f'attachment; filename="{fname}"'
    }


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"[STARTUP] Flask روی پورت {port} اجرا میشه")
    app.run(host="0.0.0.0", port=port, debug=False)
